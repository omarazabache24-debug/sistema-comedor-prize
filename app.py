# -*- coding: utf-8 -*-
"""
Sistema Comedor PRIZE - Interfaz PRO
Archivo único app.py para Render / local.

Usuarios demo:
- adm1 / adm1
- adm2 / adm2
- admin / admin123
- comedor / comedor123

Dependencias recomendadas en requirements.txt:
Flask
pandas
openpyxl
gunicorn
"""

import os
import re
import base64
import sqlite3
import smtplib
try:
    import psycopg2
    import psycopg2.extras
except Exception:
    psycopg2 = None
from io import BytesIO
from datetime import datetime, date
from zoneinfo import ZoneInfo
from functools import wraps
from email.message import EmailMessage

import pandas as pd
from openpyxl import load_workbook
from flask import (
    Flask, request, redirect, url_for, session, send_file,
    render_template_string, flash, jsonify
)
from werkzeug.security import generate_password_hash, check_password_hash


# =========================
# CONFIGURACIÓN
# =========================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# Carpeta persistente: en Render usa /data si existe o define PERSIST_DIR=/data.
# Esto evita que usuarios/consumos se pierdan cuando el servicio duerme o reinicia.
PERSIST_DIR = os.getenv("PERSIST_DIR", "/data" if os.path.isdir("/data") else BASE_DIR)
os.makedirs(PERSIST_DIR, exist_ok=True)
STATIC_DIR = os.path.join(BASE_DIR, "static")
UPLOAD_DIR = os.path.join(PERSIST_DIR, "uploads")
REPORT_DIR = os.path.join(PERSIST_DIR, "reportes_cierre")
CONCESIONARIA_DIR = os.path.join(PERSIST_DIR, "consumos_concesionaria")
ENTREGAS_DIR = os.path.join(PERSIST_DIR, "reportes_entrega")
DB_PATH = os.path.join(PERSIST_DIR, "comedor_prize.db")
DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
USE_POSTGRES = bool(DATABASE_URL)
APP_TZ = ZoneInfo(os.getenv("APP_TIMEZONE", "America/Lima"))

os.makedirs(STATIC_DIR, exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(REPORT_DIR, exist_ok=True)
os.makedirs(CONCESIONARIA_DIR, exist_ok=True)
os.makedirs(ENTREGAS_DIR, exist_ok=True)

# =========================
# LOGO PRIZE EMBEBIDO
# =========================
# El archivo app.py genera automáticamente el logo real en /static.
# Así el logo NO se rompe en Render aunque solo subas este app.py.
LOGO_PRIZE_B64 = """/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAMCAgICAgMCAgIDAwMDBAYEBAQEBAgGBgUGCQgKCgkICQkKDA8MCgsOCwkJDRENDg8QEBEQCgwSExIQEw8QEBD/2wBDAQMDAwQDBAgEBAgQCwkLEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBD/wAARCAEyAWEDASIAAhEBAxEB/8QAHQABAAIDAQEBAQAAAAAAAAAAAAcIBQYJBAMCAf/EAFcQAAEDAwMCBAMEBAcJCw0BAAEAAgMEBREGByESMQgTQVEUImEVMnGBI0JSkQkWGDNiobEXJFVyk5SVs9E0Q1dzdHWCksHS0zY3OERFRmSDhKKjsrTh/8QAHAEBAAIDAQEBAAAAAAAAAAAAAAUGAgMEBwEI/8QAOhEAAQMCAwUFBgYCAgMBAAAAAQACAwQRBRIhBhMxQVFhcYGhsRQikcHR8AcVMkJS4SPxFlMzNGKy/9oADAMBAAIRAxEAPwDp6iIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiKOdwfEDthtu59Ler82qr2HDqGgAmmb8xaerkNYQQchzgfotcs0cDc8jgB2rCSVkTczzYKRkVYNL+IvcvevWUek9t9O0tntuC6ur53edNT0/WAZWk4Y1+Dhrel/zEegJVnYmGONkbpHPLWgFzsZd9Tj1Wmlq46wF0VyBztYHuWqCpZUguj4Dmv0iIupdCIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiLB6w1tpbQVoffNWXmnt9K3IaZHfNI4AnpY3u52B2C8G5W5Wm9rdNTaj1FUcDLKama4eZUy44Y0f2nsAueO6u6upd2dSvv1+n6Yo8x0dIwnyqaPP3Wj3PGT3P5ACHxTFmYe3I3V55dO0qNr8RbRjK3V3T6qTd3vFxq3WoqbJo9r7HZ5AYy9j/75mbjnqcPug5xhv7yoNtluuepLvT2uhjkqq2umEcbRy573H/tJXgAycBXL8IOx32RQM3R1TRQPqa+MOs8bx1Ogi5/TewL/ANX1Dec/MqnAyoxmpDXuv1PQKuxNmxOcB5v17Apa2J2it+0mjYaB1NB9tVrWyXSojPV1vGelgJ/VaDj2yXH1UkIiv8MLKeMRRiwCuEUbYWBjBoEREW1bERERERFq+rNztB6IhdJqTUtHSvaHkQh/XK4tGS0Mbk55Hf3WTWOebNFysXvbGMzzYdq2hFWnVvjX03Rian0fp2euk8tphqKx/lx9ZPILG5JAH9IKKNR+Lvdi8zS/ZtdSWinli8vyqWnYS045cHvDnA/geFIRYTUyakW71Fy43SRaA5u4f6V7HOawZc4AfUrC3vXOjNNTMp9Q6rtNtllHUxlVWRxOcPcBxC5xXzcXXGpYGU1/1XdbhFG7rbHU1kkjQffDiQsC+plkPU+Qk++V3MwP+b/gFHSbRf8AXH8SulP913az/hF05/pKH/vL4zb0bTwNLn7h2E4/YrWOP9RXNrzn/tn96ec79s/vW38jj/kVp/5FN/Aea6Ku8QWzjXFp13buDjhxP/Yv7H4gNnZHhjdeW0E/tOIH9i50+af2j+9PNd+0f3r7+SQ/yPksf+Qz/wAR5/VdKI94tqZWhzdxNPAHn5rhG3+0r3Wvcfb+91jLdZ9a2Stqn8thgro3vd+AByuZPnP/AGz+9f1tRI1wc15BHY5WJwOPk8rMbRS31YPiuqzXsf8AceHY9jlf1cxNPa/1jpSR8unNS3C3GUtMnw9Q6MP6TkdQBwe57qf9uPGbdaSSK37j0DK6BzsGupmBkrASOSwDpcAM8AA/iuOfBpoxeM5vIrvp8eglOWUFvmPvwVvEWM05qaw6utUd703dIK+ilJa2WF2QHA4LSO4IPoVk1EEFpsVONcHC7eCIiL4vqIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiIiLEas1XY9E2Cr1LqOuZS0NGzrke7kk+jWgclxPAA5JWXVLvGRu6+9XsbZ2WpPwNreHV7mPOJaj9gj+h/bn2XDiNaKGAynjwHeuStqhSQmTny71C+626Wot1tU1F/vVS/4cOcyipAcR00GflYB74wXH1PPsBpaLZNu9C3fcjV9v0jZQGz10mHSuaSyGMDL5HY9AAfxOB3IXnRMlTJc6ucfiVSyXzPudXFSR4Y9lJNztUtvV2jiNgs0rJKtkjcipd3EIH1xyfbPuFf+KKKCJkEEbI442hjGMADWtHAAA7BYLQei7Rt9pWg0pZYmtgoogx0gYGumf8ArSOx3cTyVn16DheHtoIcv7jxPy8FcaCjFHFl/ceKIiKSXciIiIiwWtdbad2/0/Uak1NXtpqSAcDu+V/oxjf1nH2/M4AJWbmmip4XzzPDI42l73HsGgZJXOff/du5boa4q5WVsjrLb5XQW2DPyNYOC/A7lxGc98YHou/D6I1kljo0cVHYlXChjuNXHh9VuG6Xi21trF9Ta9JuNgtEnVGDEc1UrCC09Un6uQc4byCPvKDqq4VddPJVVtTLPNK4vfJI8uc5x7kk9yvD1FfalpqyulEFHTSzyO7MjaXOP4AK2w08dO3LGLBUueeWpdmldcr9eYnmLJam0dqrRs9PT6osVbbJKqLzoW1MLmdbM4yMjlYXqK2CzhcFaXMLTZ2hXo8xb3onZTczcO1m9aU01JV0ImMHnuniib1gAn77gSBkcgFaDSRSVVTFTRjqdK8NAz3JK6e7Y6Ui0RoCxaYZTCCSjo4xUMD+vFQ4dUvzevzuco/Eaw0TAWC5PVSeF4e2ue4PNmjoqlU/gq3RmhZLLd9Pwvc0ExvqZSWn2OIiP3FfT+RNub/h7Tn+czf+ErsIoP8AOKnqPgrB+RUfQ/FUn/kTbm/4e05/nM3/AISfyJtzf8Pac/zmb/wldhE/OKrqPgn5FR9D8VQS7+E3em21r6ak0/T3KJvaelrIgx3+ULHf1KPtSbc690g6o/jHpK6UMdM8RyTyUzvJDj2xKAWH8iV08X5liimYY5o2SMPdrgCD+RW6PG5R+toPktEmz0Dh/jcR5rlB5ieYrd+K3YfSNDpefcjTMNPaKqhewVdNE0NiqmySBocGjs8OcDx3GfYKnnUVP0tQyrj3jFXKujfRybt6mDw67vV+2utaenqJaiay3R7aerpWPHSHOIDZQDxlv4jIz9F0JBBAIIIPIIXJmB7hPGQf1x/aupmhyXaLsDiSSbXSHJOc/omqExyFrXNlHE6HwU/s/M5zXwngNR4rNIiKBVjREREREREREREREREREREReX7Wtf2kbP8AaVL8eI/O+F85vm9H7XRnOPrhepRrvXsva917OHwzCgv1FG4UNczg88+W/HJYSB+B5986p3SMYXRNzEcr2v4rXK57WF0YueikpFQ/SniJ3a2W1BPo3WLpLrT2yUUs1JWuLnxBoaB5cnfHSMjktPVnCuFtxuno7dK0m6aVuTZXRAfEUz/lmgcfRzfbIOCODhcVFikFacg9145HiuWlr4qo5Ro7oVtyIikl3LV9ztaQbfaDvOrpi3qoKYmAOaXNdO75YmkDnBe5oP0K5h3K4Vd1r6i5V075qipkdLJI9xJc4nJJJ57q1fjl1yHSWTb2ll+5m5Vbekgg4LIhnsQQZDj3AVS1RtoarfVO6HBnqeP0VTxmo3s+7HBvqv3BBLUzMp4I3SSSODWNa0kuJ7AAcldC/DdsxBtTpL4u4dMt8vLGTVTzHgwMxlsIzzxnLu2T6cBQ54PtkXVtQ3dXUlPLHFSyFlphezAldjDpue7QSWj6g+wVwlI4BhuQe1yjU8O7r4ruwehyj2iQanh9UREVoU8iIiIiIiItH3vnr6baTVU1sknZUtt0nQ6HPWO2cY57ZXMcy88k5XW2eCGqgkpqiNskUzCx7HDIc0jBB/JQVdPBltBcLqy4wMuVJCZHyTU0U4LH5OQ0EjLQPzU1heIRUjXMk56qDxbDpqx7XxEaC2qp9tTtJq3d2+fZOnKdrIYh11VZMCIadnuSO5PYNHJP0BIvptHsdo7aS1wx22ljq7wY+mqukrP0sjj3DRz0N5wAPTuSeVtOjtE6Y0DZYrBpS1RUNHHyQzl0jscue48uJ9ys4tFdiUlWcrdGdOvet+H4XHRjO/V/Xp3fVUv8deo55tW6f0uYgIaGgfWB/q500haR+Xkj96q95g91JHiW1bHq3eXUVZTVs1RSUk4oYBKCPLETQx7Wg9h5gkP559VF/WFZ6GLdUzG9nrqqriEu+qnvHX00UkbDaW/jrurYLLLQPrKU1TZqpgd04hYep5JyOMA9l0xVJ/ArpVtdq286uqKOo6bbTCCnnGRF5knDmn0J6fT0yrsKu43LnqAz+I9VZMChyUxef3H00+qIiKHU2iIiIi81zudvstvqLrdayKlo6WMyzTSu6WsaO5JWtblbp6Q2qsL77qm4Bg+7BTRkOmnfjhrG/wBp7D1VCN6PELq/dy6TxvqZbfYGvHw1sjkPQA3OHSY++85JJ7DgDspCiw6SsN+Dev0UbX4lFRC3F3T6rZfEn4hZd1rmzT1g64NNW2YviD2gPqpgC3zXeoABIa36knnGIO8we68/WFuO2W1ert1r/HY9MUPVjDqipky2Gnjzy57vT6DuTwFbo44qSLKNGhU6SSWtlzHVxWV2T26ue5uvbdZaSlkfRxStmrpxF1shhByS7kd8YHOcldL6engpKeKlpomxQwsbHGxowGtAwAPoAtA2V2asWzWmfsm3yGruFWRLX1rm4Mzx2a0ejG9gPxJ5KkNVPEq0Vcnu/pHD6q4YXQ+xRe9+o8foiIijVJoiIiIiIiIiIiIiIiIiIiIiIiKCvEtsBbdxrNUausNP5OpbdAXfo25FdE0Z8twHd4Gek9/1TnjFLdCa51Nthqun1DYamSmqqR/TNC7IbNHn5opG+rTj8jgjkBdR1U3xb7CS1Hmbn6MtkfTDG514p4W4dgc+eGjvjnqx+PuVWcaw1wPttNo4am3r3jmoLFKEg+1QaEcbeqsloLW9k3E0tRassEr30tY3lr29Lo3jhzCPcHjgkexK2BUX8IG6k+lNZjRNzqnfZl9cI4muJIiqezSAOB1EgEn6eyuBujqCn0tt3qG+1TpWsp6CUAxffD3joZj/AKThypPD8QbV0u/dxHHwXfR1gqKfeniOPgue++uq36y3U1BeOqo8kVToIGTOyY2M+UNHsMgnH1Xt2H2ir92tZwW50MrbPSuEtyqGODTHF7NPPzO7Dj1z6LQ7dbq/UN5p7XboX1FZcKhsMLM5c973YAyfqV0i2W2st20+i6WxQxxOuErRLcKln+/TfieekZwFVMMonYpUmWT9N7n6ffJV6gpTXzmR/wCm9z9Futtt9FaLfS2q3QNgpKOFlPBE3OGRsaGtaM+wAC9CIr6AALBW8C2gRERfURERERERERERERYPXGpKXSGj7xqaskeyK3Uck5cxnU4EN4wPXnCzigXxn6v/AItbPT2ynuTqWrvdSylYxrcmaIcytzjgYxn8VvpYt/M2PqVoqpdxC6ToFQG5XOqutxqrnXTumqauZ880h4L3vcXOPHuSSvN5i83m/UL60zZameOnha58kjg1rWjJcT2AHqV6BYBefaldCfBbpZti2fjvD21DJ77VyVT2yjADWnoYWcdi0A59cqfFg9C6dg0joyyaYpp5JorXQQUrJJBhzgxgGSPQ8LOLz+ql38zpOpXoNLFuIWx9AiL+Oc1gLnOAA7klQDu54xNvNAMktulpotTXgYHTTS/3rFkA5dKMg8Hs3JyMHC+QU8tQ7LELlfZqiOnbmlNgp3uFxoLTRTXK51kNLS07C+WaV4axjfckqr+8XjYslphnsu1cTbjWuD4zc52YhiOOHRsPLzzkF2Bx2cCqtbj757kbp1Jfqm/yOpgcsoqceVTx8Nzhg78sB5JOcr+bdbK7lbpS50npuonpQ4h9bKBFTtILcjzHYaSA4HpGTj0Vhp8Hipxvapw+X9qu1OMTVB3VI0j1/pYPVesdQ62vdRqLU90mr6+qd1SSyH09GgDAa0egAAC9mjtvta6+rRQ6S07WXGTnJjZ8jcDPLj8o4I7lW82x8Dmm7FJDdNxrx9s1Ubg/4KlBjpgQ7IDnn5nggAEYb6qyVi0/Y9M26K0aetNLb6OFrWMhp4gxoDWhozjucADJ54WdRjUMIyU4v5BYU+CSzHPUG3mVVLbLwMCN8Vz3SvnWGkO+zbe7g4d2fKeSCO4aAR+0rVac0zYNI2mGxaatNNbqGnGGQwMDW/Un3J9SeSsmir9TWTVRvIfDkrDTUUNILRN8eaIiLlXUiIiIiIiIiIiIiIiIiIiIiIiIiIiIi+dRT09XTyUlVCyaGZhjkjkaHNe0jBBB7ghfRERc/wDxF7MzbQ6xjvmnqOSPTlbK2SieX9fkyjl0XuMHtnPGOSQVN29W59RqXws0GqKENL9QupqOs64+npe3qM3SM8fpISB9FNm42g7RuRpGv0rd4mFtTGfJlc3JglwemRv1BVVdGaO1LqOzW/w+3uoeG27VVZNWtPDoKOCKIh7A7pPQ907nNPr1ZwqtUUj6GSRkI92UWHY7p3an7CgJqd1JI9kX6ZBYdh6ea2XwbbNuoYX7qagpMSzNdDaGuIIDDxJNjGQe7Bz+3x2Vq15rbbaKz2+ntdtp2QUtLG2KKNgADWgYAwF6VPUNI2igELfHtKlqSnbSxCNvj3oiIutdKIiIiIiIiIiIiIiIiKj/AI+9Ymq1TYtF09fBJDb6U1c8DQC+OaQnHUfTLGsIH1V4CQBkrlr4ktXSax3o1Rc3TU0kUNc+jgkpzlj4Yf0THZycktYDn6qZwOLeVOc/tH9KHxqTJT5B+4/2o462+/8AUpR8M2mTq3e3S9vFR5LaasFe53l9QIpwZuk+3UYw3P1UUdQ9/wCpWR8Kl2O22lNeb1ttkt2fY6WCgZQRsx1GeVpMhk56WtDOcA8OJVmrXFkDsvE6DvOgVco4w6dubgNT3DVdAqiop6SF9RVTxwxRjL5JHBrWj3JPAUE7qeMTbLQBmttjn/jHdoyWGKldiCNwLmnrk7HDm8hueCOVUHU24G/3iErpXRQXu5URlEHwVqppBRwl2MMd0/L6A5eT+OFIm3HgS19fJo63cK5U2n6PLuunjcKiqcQ4cYaehoIzz1HHHyqvsw2mpRmrH69B93U8/EaipOWkZ4n7so43M8Rm6W71Q6guF1lprdM8NjtdAHRwnLmlocBzIepoILySCeMLIba+FXdjcV8FU6zus1rlDHmtr29AMZJBLGfecRjtx6cq8O3Xh12m2ydHU2DTTKivj7V1eRPOPmDgRkBrSCBgtaDwpKX2XGWRN3dIyw7fp9VjHg7pXbyrfc/fP6KAdtvBjtXoyOGr1JTO1PcmgF7q0f3sHdGHBsI+UtzkjqyRwp5paWmoqeOlo6eOCGJoYyONoa1oAwAAO3C+qKFmqJag5pXXUzDTxU4yxNsiIi0rciIiIi8tXdrVQSCKvudJTPcOoNmmawke+CVrW7etX7ebc33WEUbJJrfTF0LHP6QZHENbzg9i7Pb0XMbUG42oNUXWa9X68VNbWTnLpZZCTjJOB7Dk8DjlSmH4Y6uBfewCi8QxIUJDQ25K6zQzRVETZ4JWSRvHU17HAtcPcEd1+lQTwnb8Xyy7hWzQ92us9RZr3I2iZBITJ5MzuIizJ+UdRAOOMHtwr9rnraN1FJkcb8wuihrG1sWcCx5hERFxrsRERERERERERERERERERERfCtr6G2wGpuNZBSwg4Mk0gY3OM9z+BUB71+LOwaEqarS2io4rvfoCGSTnDqSmdz1NJBBe8YAIGAM8nIIVT9Wbla13ArfjdVX+prCHAsiLumKPvjpYOBwSM9+e6qWMbXUuGuMUI3jx00A7z9FecE2FrsUYJ6g7qM8Li7iOoGmnaSFeC+eJjaSyzNgZfn3Bz4/MDqKLrZ64HUSBnj+sKLKDxM6No9a3PV9HoieOa50kFNMROwOe6Nz/AJ3EM5Ja5je/ZgVXqf0WSp/RUGr23xWZwLC1tujQfW6ujNgcHhH+Rrnkcy63/wCbK2rPF7YX/wDulVD/AOqH/cWy6f8AEvoC7xQ/aLKu2zyydBY9oka3nAcXDnHOe3CptAD/AFrI03YLmj27xmJ2Zzw4dC0fKxWio2LwlzbMYWnqHH53C6A2XUtg1FD59ku9NWNyQfKeCRjvx3WSVCLNd7pZ52Vdrr56WZnZ8Ty0hTrt/wCIevZLHbtaQioicQ0VkTcSMyTy5vZw5HbBwOxKt2D/AIiUtU4RYg3du/kNW+PMeY6lU3FNjZ6UGSkdnHTg76Hy7lYJF8KGvobnSsrbdVxVMEgy2SJ4c0/mF916K1weA5puCqYQWmx4oiIvq+IiIiIiIiLAbgakOjtDX/VTYGTutNuqKxsT39IkMcZcG5+uMLkLUTmaeSV3Bc4kroh46tUtsWyb7P5Mjn3+4QUjZGuwI+g+cc+4IjLcfVc5esftK2YBDlhdIeZ9FWsafnlawch6r7dX1XTPwiaRj07sVaGzOZOby6S4SB0WMB5DQw578M7n3XNrS1ofqPUtqsEUpjdcayGlDwwu6Ot4b1Y9cZzhdg7NbmWez0NpjcHNoqaOnDg3GQxobnHp2WOPzZY2RDmb/D/a+4LDeR0h5C3xX3p6SlpGllJTRQtcckRsDQT78L6oiqqsiIiIiIiIiIiIiIiIi1HdvRLtxduL/o2KYRTXKkcyF5AOJQQ5nfsC5oBPoCVyc1hofcPQl8m09qfStyoqyL5uh8DiHs6iA9pHDmnpOCOCuyKKUw/E3UALcuYHwXBWUDashxNiFz18FmwertRa1t25+qLNNQ2CzSfE0pqQ+N9VUNz5ZjHGWteA4k8fLjnK6FIi562sfWy7xwt0C3UtM2lZkaiIi410oi+VVV0tDTSVldUxU8ETeqSWV4Yxg9yTwAol1x4ptqtGvkpKe5S3utYJB5dvaHRtkaBhrpHENwScZb1ditsUEkxtG0laZqiKnGaVwCl9fySRkTDJK9rGNGS5xwAPxVKdZeNDXV486m0pbqKxwPDeiXHxFQwjvhzsM59ujt6qI9Q7q7gaqlqJL7rG61LasYmh+JcyFw9vLaQzHHbCk4sFmfq8geaiJsfp2aRgu8h9fJdKW3W1ucGtuVKSeABM3J/rXqXKeOrfDI2WKQsew5a5pwQfcEdlKG1fiB1xonU1HUXHUlwuNpkljZW01VK6oBhzhxZ1nLXAEkYI5AzkcLbLgj2tvG657lph2hY5wEjLDre/0XQhF/I3tkjbI3OHgOGfYr+qCVjRUq8WnicuEt3q9rNAXLyKSm/Q3Wvp5Pnlk/WhY4dmt7OI7nI7DmT/ABmb4S7WaBbpywVRi1BqZr6eGRkjmSUtPjEkzSB3/VHIILsjsuddC5z3dbiSScklVLaTFHRNNLCbE8T8vqvVNgdl2VAGK1jbtB9wHmR+7wOg7deQWzUL3yO8yRxc5xySTySt30TonU2ubrHZ9M2metqHEdXQ35WNJA6nHs0DPc8cFe/YXZm8bsXl56xQ2G24ludxkwGRM79LSeC4gH6AAk+mdu3T8VVi27pqvbHw326mt1LC5sdXqEASTVL2n5jH1A5Hcdbs/ePSG8FVnZ/Yyr2kmz/piHE9e76/AHVWLbDbmk2abuW+/MeXId/0+JGl90ZsBoTb+giuO825tvsz5GhzaOGRpkLmjqezJ7nGBwD3WJrt4fCNpeWnrdO6QvWopXOIfFMXRxxjHch5w79yqBXXa6Xyvlud5uNTXVlQ7qlnqJXSSPPuXOJJX0ga5x+VpP4BexYf+HWB0LRmiD3DmdfW/lZeCYn+IWOV7id6Wg8m6elvO6uDTeIDwxaorIaK9bZXGw07Q5xqqVwOD3Ac1hyc/mtps+nPD5uW4t293BFurpP0cNBcD0Oc5oySA7BII/Hsqe2/QOua+mZWUOjr3UQSt6mSxUEr2OHuCG4IXkEU9LI6GeN8UjHFrmuBBaRwQQfXKzrtgdn8QaQYQD1Gh8rHzUfS7cY5QuB3pI6HX108lbDV21OrdEl01dRGehJ/R1cHzxPBLunkdiQ3ODg8ha/TcHBWv7PeJPVuggzT+oSNQ6amcxk9FX5ldFGBjETnHgYx8py35eAO6m/WehrHebNDuVtlM2ssNW3qmhj+/SO9QR3AHYg/d/A5Hh+2f4bVGAMNXREvi5jmO3tA+I46i9vVNmdu4McIp6kBkvkfv7svNtpuRctFVwYS6egmcPOgJ7/Uezvr+/jtaC13OivNup7pbphLTVLBJG7GMg+/sfQhUsgBDsEYIUw7I64fa6/+LdyqHmkq3DyOp46YZD379g7j17+nJUdsNtQ+gnbh1U7/ABPNm3/a4/I+R16rftNgzalhq4R7449o+o9PBT4iIvb15yiIiIiIiIqEfwhOsGV2urFo6E1LfsmidUzAu/RPdMR0kD3AaQT9fxVTOsKQ/EdrRmuN59T3ymuE9VSfGOgpXTZBZEz5Q0D0HBwPqo08z6r0Chi3FMxnZ66qm1b97O5/arB+CXS0ep997dVSzGNljpZ7njy+oSOAEYYc9v53Of6K6Xqon8HbpP4PR2pdavmcXXOujoGROjwGtgZ19Yd65MxHH7Ct2qrjM29qyP46Kw4XFu6cHrqv5JIyKN0sjw1jAXOcTgADuVA2pvGNtvZLnNbrVQ3C8thy34mnDWROcCQQ0uOSOM5AwQRhTdeoJqqz11NTl3my00rGdOM9RaQMZ4XKy5RVlsuFTbq6J0NTSzPgmjd3Y9pIc0/UEEL7hVFFV5jJysubGK6ajDBDzvquh203iD0ju3cKm0Wykq7fX08XnCGp6T5jAQHFpbxwSOO/r6FSPc7rbbLRTXK719PR0kDC+WaeQMYxo7kk8Bc0Np9wH7d6/s+rC3zIqOoHntxkmF4LJMDjnoc7H1ws5vfvfet2dSSz/ESQ2Wlkc230f3QxmeHOGTl5GCT+Q4C6ZcFzT2j0Zb7C5IcdLae8ou+/d4q9X92naP8A4SdOf6Ri/wBq9lp3Q25vtYy32bXNjramT7sMFdG97vwAK5f/ABT/ANtf0Vbwch/IW04FHbR5WkbQy31YPNdY454JsiGZj8d+lwOF+1yusetdUaZlkn09qC4W2SZobI6kqXwl4ByAekjI/FS9pPxk7qWF/ReZqO+wdLGBlXCGuYG9yHx9JJI7l3UuWXA5W6xuB8l2Q4/C7SVpHmr5ooY2p8U2g9zLlHYJ4ZbHdpy4QQVEgfFNgDAbJgfMcnDSBnpOCeFM6iZoZIHZJBYqZhnjqG54jcI5zWgucQAOST6KLrz4mNnLHdqizVmqC+emf5cjoKaSWMO9QHtBBx2OD3BUJeLLf01NXPthpGumjjpJDHd54z0iSQf7yD36W/re549FVg1TyckqYosH30e8mJF+AUHX426GTd04BtxJXVu13S3Xu3wXW0VsNXR1LBJDNC8OY9p7EEL1Kk/g03Dv9LrkaDdUGa1XKKaXynuJEMrGF4ez2z0kEeufore621XRaH0ndNW3GOSSntlO6dzIxlzvQAfiSFHVdG6mm3I1vw8VKUVa2qg3x0tx8FkLrdrXY6Ca6Xm4U9DR07eqWeokDGMHuSeAq3bmeNXT1odNa9ubf9rVDepnx1QCynacEZa37z8OwfQEepVWdzd6Nb7p3SSr1FeJnUgkLqehY4tghbk9OGDguAcR1EdR91o/nn3U5SYKxnvT6npy/tQVZjkknu0+g68/6Uia53m3B3Dmc/Uuo6iWEuc5tNG7ohZnGQGDj9ULTDN1HLnEk+pK8VO2oqpmU9NFJLLI4NaxjSS4k4AA/FThtj4Sty9exw3K7xt07bJQHCasYTM9pGQWxcH6fMQpV74aRnvENCh2Rz1j/dBcfvmoda4vcGsBJJwAF+6iGopHBlVBJE5zQ4B7S3IPYjPor6WvaXY/w4acm1re6cVc9AA4V9eGyzukzljYWcNa8ngdIB9zjJVLd4N1bjuxrWs1TWROpoHkR0lL5hc2nhaMNaPTPqcAAuLj6rnpaz2t53bfdHM9ewLoq6D2Ng3jhnPIdO9ax5o91uO0Wkn683FsemRTPnhqquM1LWv6SKdp6pTn0wxritBhM1RKyCBj5JJHBrGNGS4nsAFfXwobGT7d2R+r9U0Iiv8AdGdMUbzl1LTnB6SOwc4gE+oAA9Ssq+pbSwlx4nh3/wBLHDqR1XOG20Gp7v7VgWMbGxsbRw0AD8Av6SACSQAOSSij3xCarborZPWOonUr6jybXLA2Nr+g9U2IQc+mDID+Soz3hjS88l6JS07qqdkDOLiAO8my5r+Izc+q3W3hvd+kD2UlJM6goongAshiJaM4JGScnP1WvaPslw1JeqGxWqmkqKuunZBFHG0uc5ziAAAPxWmUJy4nGM8q1fge07bKnW141vdvh3Q6Utr6xsc7W9JkOQ1we7hhGCQV5s+F+JVbYjxef9+S/S1fPDs7hLpIx7sTNB3aD4m11lPE3rizbJbe0Hhq0A+JtwqYI6zVFbTzEudK4AmI+vz4DsHGIxGOQ4qocPJKymvdYV2v9cXzWVxkmdLd66WqDZpTI6ONzj0R9R7hjOlo+jQty8Oe3dBunu/p/Rt2l6KGqndLVN+b9JDEx0r4wWkEdQYW5B46s+i/Q1DRw4RRCJgsGjXwC/HlfWT4tVunldmc88+0qWPDv4O9S7ow0mrdXySWbTMwEkLgB8RWN/oNP3Wn9s9wTjKvDofYzavb2lFNpvR1Ax5ibFJUTxiaaUAkjqc7OTk+mFnbxqbRO3tnaL1ebbZaCip8xxyytj6YWDHyM7kDgYAKgjcPx07baafNQaMoajUlWzLRMD5FMCWZa4OILnjJwR0jseVXJZ63E3WjBy9Bw8SpNkVJh7bvIv1PHwVk2Mhp4gyNjIoo24AADWtA/sCp34ytVbMXKjbY7VFTVWsKOoAdPQtAbA3qPWyZw4ce/wAvcHk49YR3H8Um7W5jJqGuvptdrmzmhtoMEZaWlpa5wJe9pBOQ5xH0Wl6U0Xq3Wlayg0zYa25Tyv6QIYi4AkE8u7DseSVIUWEupXCaZ1iOQ+ZUVX4qKlpghbe/X5BeWn7q5Xgbr7tVWfU1lqmOltMToJGh7stZK8PDgG+vU1oz/iN91gtrfBFcZ3QXXc27ili4f9m0R6pHfdID5Dw39ZpABPs5Wy01pbT2j7VHZNM2imt1FFyIoGBoJwB1OPdziAMk8nC1YriEEsRgZ71/gvuE4bPFKKiT3QOXMqt26Oh26L1EY6XqNFVfpIC4gnp9j+ByPyB9VrlunfS1EVRG5zXRu6gWnBVg987JHctHm4ASGW3yiQdDQctPBzxnA7qu8C/J+2GEtwbFXxRCzHe83sB5eBBt2WXvuC1zsQoWvk1cND225+IVtdIXkX7TtFci8OkfH0Snqz87eCT+OM/mswov2HuM09lrre9rRHTyRytI75cC05/yY/epQXuezlecTwqCqdxLde8aHzBXnGKUwpKySEcAdO46jyKIiKaXAiwut7zPpzRl+v8ASujbPbbbU1cRk+75jInObn6ZAWaUI+M3UFBYvD1qSGsq3QS3XyKGmDc5kkMrXlvH9CN+foCt1PHvpmx9SAtcz93G5/QLl7crhPcLhU19S5rpqiZ8shAwC5xJPC+MXXNKyJjS5z3BoAGSSfovJ1lbtsrpmPWm62l9MTVbqZlwucMTpWjJaOoHgL0V5DGlx4BU5rC4gdV1K8P+jf4hbN6U0090plit7KifzYvLe2WYmV7HN9C10hbzz8qkFEXm8jzI8vdxJurmxoY0NHJFz68YO3zdDbnyXm3Uzo7dqOP45mI+mNs+cTMBycnOHn6yroKoT8XO3D9fbTVVZb6Rstz0+/4+AiPqkMQGJmNPoC3Dj/xYXdhdR7PUi/A6FR+K03tNMbcRqPvuXO/4k+4T4k+4WKMpaS05BHBTzvqrtYKlZCsr8S73C2a17dbj3yhiuln0Lf66jnBMc9PbZpI3gHHDmtIPKz/he27tm6G7NDZb2I5LfRwyXCpgfnE7Iy0dHHbJeD7YBHqumdNTU9FTxUlJBHDBCwRxxxtDWsaBgAAdgofEcTFE8Rsbc8VLYfhPtjDI82C5ct2i3ce4NbtrqbJ97VOP7WrZLR4ZN9bzT/E02g6uJp9KmWKB37nvBXSZFGux2U/pYPNSTdn4QfecfJUk2e8Im51PrG26g1i2lsdHaK2Cr6TOyeWfy3dYawRktHLWglxHDsgHGFYvxFbqt2n26qrtSvZ9q15NHb2uz/OOHL/+iMu/IKUFQbx0a5qrruVT6QbK0UdkpWEBkpIdLIOpxc3sCMge/Cwp3vxarbvrWGth98yts8UeE0jtze55/fRQFPcJqqZ9RUSvlllcXve92XOcTkkk9ySvn8SsV54/aWR07aLjqi+0GnbRH5tZcaiOmhb6dT3AAk+gGck+gBKthAaLlVIMJNgrc+BnQMlTXXTce4UY8qnaaKge9jgfMd/OPYexAblp7/eKtlqfTtt1bp+v03d4RLSXCB0MjT9ex/I4P5LG7b6Htm3Oi7XpC1RtEVBA1sjw0NMsp5e92OC4nJJWyqi1lUaioMrfDw4K9UVIKanEJ8fHiqC6q8EG7Vprms05Pa71SyyuaJI6gQuiZn5XPbJj07hpd2Wz6Z8A+pJqhx1fre30lP5eWi3xvnk68jg9YYAMZ5yeccK6aLqdjVWW5bgdtlzNwWkDrkE9l1Hu3ewu2W2LRJp2wMlrASfjazE0/fIwcANxgfdAK3LUGoLNpazVeoNQXCGht9DGZZ55XYa1o/tJOAAOSSAOV+dSaksmkbLVah1FcIqKgo2GSWaQ4AHsPcnsAO65x+IzxH3rea9G32+SWi0vQyE0dHnBmcOPOlx3cfQdmg8ckk4UlJNiUuZ5NuZPottVUw4dHlYBfkAv34ivEBX7y6kDKJslNp+3Pc2hp3Hl/oZXjt1H+ocKIhUEnA9V4RIScAnJ+iuJ4VvCo6tbR7lbm28/Du6Z7Za52fzg7tllaf1fUNPfueO9pkkgw2DoBwHVVZkM+Iz9SeJ6LOeEDw8OoYId1NdWwiolAdZqSdvMbD/6w5p9T+oD6fN6gq2yABoDWgADgALWdWbm6A0LG2TVmrbbbep5jDZZh19WM46Rk9voqbUTy10uci55AK4U8EVDEGDQcytmUJeM+sFJ4btXtIcfiI6aIYaTyaiM847DjutZ1F48dpLbSslsFuu93nc/BiMQpw1v7XU7P7sKPtd+NTbLcbS1do3VW2F1qrXcWtbURMuYiLg1wcMOa0EctHZJcIrZonNaziCNdF14bjdFQ10NRK+7WOa4210BB7lRS3+it/sXFE3wgbxVDY2iUwyMMgA6unyW8Z745PC17T1H4UdxL9b9FW7bnUelay91UVHT3KC7OqhBK9wDMxP+UhziGk446s+i3vbbTNXpfw8777cSvZVXi1S1MEkEB63vDGdHW0DkglpVWocGq8IxinNUAA46ar1faLa/DdqtnaoYeTdmW4IsbXGvPRUdhWw6T1Rf9HXqDUOmLrPbrlS9RhqYXYezqaWnH4gkfmteiWx6LsH8atVWjTPx8ND9q10FF8TN/Nw+Y8M63fQZyfoF7XJbKc3BfnIXv7vFeivv98v07ai9Xitr5Wt6Q+pndI4DOcZcTxlb/ttsRujufLE7S2lKuSjkJzXzt8mmADulx8x2GuwTyG5PB4V5dtfBjszoIxV1bbH6luDMkT3PD4+cdoR8nGOCQTz3U7xxxxNDImNY0dg0YAVbqcea0Zadt+0/RSsODOfrO74fVVa228CGlLGYLhuHfZbzVMLXupKQGKnDmvzgvPzvaQADwz1VlNPaY09pO3R2nTVmpLdSRNa0RwRhuQBgZPdx+pJKyaKAqKuapN5XX9PgpiCkhph/jbb1+KIi1rUm5WgtH1UdFqXVVvt9RKCWxSy/NgYzkDJHcd1oaxzzZouVue9sYzPNgvruEcaHvp/+Bm//AFKqnF94/mrPa0vdouW2tzu9FXwzUVbQP+Hmafll6hhuPxKrDFy4kfVeJ/if/wC/C3mGH1K9B2Q1pZHDhm+QUu7C1dQ29VlEJD5MlKXub/Sa9uD/APe5TeoY2DtwdV3C6eZzFCyEN/x3E5//ABj96mdXX8P2vbgUZfwJdbuzH53Vd2kLTiD8vQX+CIiK6qCRVE/hHtU0lBtzpzSMkEjqi6XR1bHIPutZBGWuB+pM7cfgVbtUG/hLdSUE+odG6WjD/i6Ckqq2U4+XomfG1mPrmF6k8HZnrWdlz5LixB2Wmd981THrCsD4GLLbb54hLR9oQ+b8BTVFbCD+rKxnyu/InKrv5g9yre/wbtgt9x3K1Df6gONTabW1tPzwPNf0uP7h/WrfiT93SSO7PXRQFG3NOwdq6IoiLz1WxF+KiCGqgkpamJssMzDHIx4y1zSMEEeoIX7REXI7d/TLtDbmaj0s6SFwoK6RjTECG9JPU0AfgQtP876qxXj+07UWbeGkvooI4KS822MxyswPOljJEhI9x1M5/BVk88e69CpJd/AyTqFSqmARTOZ0KtF4CZyd7p4xjDrHVZ/68S6ILlH4W9YxaS300tcKq6OoqSeqNJUvHZzJGloY76F3R/UurirGPMLakO6hT+Dkbgt6FERFCKWRcpfETVMfvfrR0bwW/bFTg/8AzCurS5G7/Q/Zm9GsqASFwhvNU3J9f0hVg2eF5n93zULjQvG3v+S1Lz/6Std4CNuZrzrO4bj3ChDqKxwGno5Xh7T8XKMFzD913TH1hwzx5reOVT74k/tLp/4LbfY6LYGyzWWq859bPUVFd8/V0VPX0lv0wxkfH1z6qVxmYwUpDf3aKOwuASVAJ5aqc0RFSla0WL1RqiwaLsNZqbU1zhoLbQxmSaeQ8AegA7ucTwAOSSAF9r7fbTpmz1l/vtdFR0FBE6aonldhrGAcn/8AxcxvEh4kb7vZqF9LSSSUemKCVwoKLOPMxx50nu8j9wOB65kMPoH1z7cGjifvmuOsrG0jL8XHgF7PEV4kr7vXfHUtIZbfpiieRRUPV80np5suOC8+3Zo4GeSYZEuTgFeHzj7hTl4dpNltINn3R3bvMNbPbZ/KtGn4W+bNNUNYX+dIz0YDhrSeOo89ublZlHDaNug4AcT99VVrOqpbvOp5lTV4UPCf5wo9zN0bZmP5Z7XaZ2ffPds0zT6erWHvwTxwbEbr+IbbHZ+B8epL2youYHyWujIlqSePvNziMYIOXEcdsqmW7Xjo3C1n51q0Mz+K9qJw18L81b2hwIzIPu9sENxwSOVWyor6mrmdUVVQ+WR5y573Ekn6kqJ/LJq6TfVhsOTRy8VJCtio2bqlF+pKsnuf439z9a+db9LeVpe2SAs6aY+ZUvaWlpDpSOM5z8rQQfVV+r7tXXSrlr7nXT1dTO7rkmmkL3vd7knklYnzne4W/wC2eyO527VV5OjtOTz07TiSslHl00Z6SfmkPAJA4HJ+ilmRQUTLtAaPvmo57pqp/vXcVp3mj3W9ba7L7k7sVbYdHabqKim6w2SulHl0sXIBzI7gkZ5a3Lvorj7TeBPQmkjFdNw6sanuLD1fDhpZRMILv1TzJkEZ6uMjsrMW+3UFpo4rfbKKGkpoWhkcULAxjWgYAAH0Ch6vHWM92nFz1PBSNPgzne9MbDpzXNyz6q2m8O832hZamDX24EIkijkML47ZZ5xlpeOsB1RI1w4IDRgnBBwVnvBfuAG7q3Sy6kqy92taaSmlqpZHGV05JcME5y5xJGT6lYDxubU1Ohd3JtV0sJ+ytVA1cbwAGsqBgSMwOw5aRnvk+yhvTN0rrJcqW722pfT1VJK2aGVhw5jmnII/NePY1jVa7EBPUm5YdByt2d4X6YwDZbCP+Pugw4WE7NXE3ObtOn6XdAOC+O7m3tbtbuXf9EVkEsbbdWPFKZHBzpKZx6oXkgAEmMtzgd8j0WtQd1eTdDRFF4xNrLduLohkLNd6Yg+FulGYmskrgGhxaCPr1Pj9Pne089qQzUdVb6uahrqeSCogeY5I5GlrmOBwQQexyvcMIxSHFqRtRE69xr9/fRfmfGMMnwmqfSztsWm3399q6b+FDxE2vdnSVJpm+XDp1daqcR1McpANaxvHnMxgE4x1ADg/TlT8uMOnb7eNM3amvlguM9BX0jxJBUQPLXscOxBVwdufH/cqG2Ch3H0ubnUxMAZWUDmxPlPU777T8vbpGRjsT6qIxDBX5zJTC4PLp3di6KPFWZQyfQ9Vd1FVMfwgmiHA9Og71nHAdURD+zKi3WfjY3R1PTmksMNJp6GSMxvNKC+UnOeoSO5acccYXDHg9W82Lbd5W+XGKWMXDr9wVr96d+NLbSWeVrqqnrr/ACAtpbayQF7XEAh8oHLGDIPPJzx6kc+rneLlqW9VF3uMr56uumdK8klxLnHsM5P0CxNTcrheK6a43OrlqqqpkMkssri5z3k8kk9yrH+G/ZltN5W8O4DI6Ww2zM1FBOzLqyYcNcAf1Q7kftOAxwCpfJT4JTOmldwFyTpw+X+1AySVGOVLYYm89ANfFSxqaprtJbM6S0PXRtir3UUJqonDqLA0ZwHA4HOffso1g9PwXu1bqut1hfp7tV9Les4Yxv3WNHAA/AAf2+q+2kbNPfr3SWyBnUZXjIPbC/Ie02Kv2kxd9RELhxDWDs5fEm/iv0ThOHtwbD2QPP6Rdx9fgNPBTzstZfs7S766WnMctbL94uz1MYMD8PmL1IC8tqt1PaLdT2ykaGxU0YjbgYzjufxJyfzXqXv2DUAwvD4aMfsaAe/n53XmddU+11L5/wCR8uXkiIik1yIua38I/wBTd8LWSCAdO02D7/p510pVUfH5shqPcrR9n1ho21zXG5abklZU0lPEHSy0svSS8frO6HMHyjPD3H0Upg0zYatpebA3C4sQjMkBDeWq5r9X1V1P4Ms51drb/m2m/wBa5Uudbbk15idQTh4OC0xnIPthdMPANs3fdtdvbnqDV1ifbbvqCryyKohdHUx00Y6Wte1wGAXdTh9CFZ8alZHRuaTq6wHxBUNh0bnVAI5K0aIioisyIiIio5/CV2uuDtEX9sLvg4hWUb5McCV3lvaM+5DHfuVHPiR7rrn4jtjKHf3b1+k5bk+319JMK621HJjZUta5o8xo5LC1zmnHIzkZxg80dU+F7f3SN0ltdw2yvdUYun++KCldVQP6gCOmSMFp74xnIPcBXLBKuF1MInOAc2/HTmq7iNM8TGQC4K1jQNRM7XOnm0zw2U3SlDC7sHea3BK7Ts6+hvmEF2B1Y7ZXPrwd+EfXkWvKLcfc3Tk1mtlmcZaSjr4+iepqBkNJicMtY372XAZPTjPOOgyi8eqI5pWsjN8vHxXdhcLomFzha6IiKBUoi5GeKGOqo9/dbsqqeWAyXeolYJGFvUxzyWuGe4I5B7Fdc1VnxjeFG4bxNg1xt/DTDU1IzyqqCSTo+OhAw0Bx+UPbjAzgEceymMFqo6WoO80Dha/RR+IwOni9ziFzg8/+kr6fwcGvhWWTU23dVWQB1HNHdKOANxI5rx0TOz6gFsP/AFlRrUOg9d6TucVl1NpC8Wuvna10VPVUckcjwTgFrXDJyeOFffwF+HnU+3cFz3K13bJ7ZcrrTihoKKYuZLHT9YdI+VhHy9TmR9IPOGkkfMFYMZdEaN2Y8bW77qJw5kgqBYd6t+vLdLpbrJbqm73etho6KkjdLPPM8NZGwDkklepc7/Hpv5c75rWXaXTt2fHZLK1guLYXt6aqrI6iC5pyQwODek9nB30xU6CjdXTCNug4k9inqqoFNHnPgtd8V3iprd4Ls/SWk55aXSFvl+UZLX3CUf77IPRo/Vb6dzzgNrn531C8fnD3Tzh7q+QU7KaMRxiwCqssjp3l7+K9nn/UJ5/1C8fnD3Wz7f7da13QvsenND2CpudY8dThG3DImjGXPecNY3kcuIHIW1xDBmcbBYBmY2AWE8/6hSDtbsbudu/Wtp9G6cmlpuoNlr58xUsQzgkvI5x6hoJ+iuBsz/B+6X026G9bsXBl+rW4eLdTktpGHg4e44dJ6gjAH4q2drtVsslDFbLPb6eipIG9McEEYYxo+gHCr9Zjscfu04zHry/tStPhLn+9NoOnNVj2e8Bug9H/AA953Gqzqe6M6X/CgGKiid8jsdOeqUhzXDLiGuB5YrP0FvoLVRQ2610UFHSU7BHDBBGI442jsGtbgAfQL7oq1UVU1U7NK66moYI4BaMWRERc63LRd6dprNvPoKt0XdpRTSSls1HViNr3U07Tlrxn0P3XAEEtLhkZXLW+aVveitQ12mNQ0MtJX2+Z0MscjcHIPce4PcEcEEEd12EUQeIHw82DeezmqgbFRako4yKOt6cCQDtFLjkt9j3GfbhV7HsINfHvYf1jzHTv6K+7F7VjBZDSVZ/wvPH+J69x5/Ec70X2h3P1PtVqFl/03U46miOpppCTDUx/svA/qPcenrmc9bbc7VeLWBmqNH3al0nuD5T/AIqgnH6KveB8uTxkk4HW3LsE9TTgKveodDaq0BeH2HVlmqLdVx5IbK3Ae3JHU09nNJBwQSF+6CpqKSZk9NM+KRjstcw4IKoWE7T4hszUkR/p5tP3p6di9E2j2Tw3auESOsHkaPbrccr9R59ttFhdwNjdz9rK2Sm1bpaqigjJ6a2BplpntDi0OEjeACRwHYPI4Woxgg4IIP1VrNI+JLcWx0jLVdKmC+27LA+nuMfm9UbcfJ1HnBA+vus2d0dnNS3iW6aw2IsskkwHXLSu6XuIGBkcDsvVKD8WMJmYBVAsd3E+l/VeI4n+FWL0zz7MWvb329bHyVQou/5LadKaL1XrGtZbtM2CtuM75Gx4hiJDS44HU7s0fUkBWerdZeHuARzae2JoHTseHA1bsNGD7AnKydd4idQyRSUulbHa9PwTRhrhSQND+oZ+bqwPQ4HHCyr/AMWMDpmnckvd2A/Ow81H0v4X4xO4b+zB2kf78lgtvvDRZ9AMptZb4XWmjEXTNDYKdwkmmJALRIcjsQ8Fo4+X73os1uBuPW65r44KeP4O0UQ6KOiZwxjQMAkDjOOPYDgeudErbvcrzUyV10rZameVxc98ji4kkk+v1J/evVQQS1ErYYY3Pe44DWjJJK8R2t27r9qDuT7kX8Rz7/p66L0/Z/Y+i2dbnb70nNx5d319FkKVrnvDWjJPorHbMaAksFB/GG7QSR19U0iKNxx5cJxyR7nvz2GOxysZtHtA22xw6i1RS/3ycPgpZG/c9nOB9fYfv9lMSsWxGyD6dzcUr22d+xp5f/R7eg5cTrwg9pMdbNejpjcfuPXsHZ1KIiL1RUlERERERERfjyIO/kx/9UL9oiIiIiIiIiIiIiIiIiIiIiIiIiIvy6ON5Dnsa4jsSMr9IiIiot4yPB7qrUGqardPaq2/aLro5rrnaYGNZIyYNAM0YGOsOx1OB+bqJPOcC9KLqo6ySik3kf8AtaKinZUsyPXDS62i9WOqNDerVWUNSB1eVUwujfjJGcOAOMg/uXq0zpTVOsrpFZdLWKuulbNIyNsVNC556nu6W5wMNBPqcBdrbrpXS9+nbU3zTdruMzG9DZKujjmc1uc4BcCQMk8L92nTenbAZDYrBbrcZsCQ0lLHD147Z6QMqfO0gy6R69+noov8n1/Xp3KiWx/8Hjebi6l1BvTcfgKUhsostG/M7wWtcGyyjhndzXNbk8cOCvFo/RGktAWaOwaN0/RWigiwfKpYgzqcGhvW4jlziGty45JxyVnEUFV189abynTpyUnBSxU49weKIiLjXQiIiIiIiIiIiItb1zt3pDcW1m1ass0FYwNc2GVzQJYCcEujf3actHbvjlVY134NNVWaWat0Lc47vRta+RtPORHUgADDR+q8nn9n0Vy0UNimA0OLi9Q33v5DQ/343U9g+0uI4J7tM/3P4nUf14WXNW8aF1fpWeWmv+nq2jdAQ15kiIaCQMc9jnIXnhhmbguiePxaV0tqqSlroHUtbTRVEL/vRysD2u/EHgrC1+gNEXOIQVuk7VIwcgClY3+sAKk1P4duJJgn07W/MH5K6w/iO1zbVEGv/wAu+RHzXPynjkdjpjcefQLNWu0XSvcY6K31EzmDqc1kZJA91eG2bYbe2dwfbtH2uJwzgmAO7/42VnaK1Wu2lxt1tpaXr+95ELWdX44HK5Yvw2mcf81QB3NJ9SFqqPxAhIO6gJ7yB6Aqqmj/AA/64vzo5bhTNtVK4tJkqPvlpGchvc/nhT9oPaPS+hQyqp4TWXADmrnAJaSMHoHZvr9cHGVu6K34PsfhmDuErG55B+52tu4cB32v2qoYntJXYmCxxysPJunxPE+nYiIitSgERERERERERERERERERERERfmWaKFvXNKyNvu5wAWLi1dpOe4stEGp7TJXSHDKVlbEZnHvgMDuo9j6LEuDeJXwuA4lZZfmWWKCN000jY42Auc5xwGgdySey/ShTeirt9z3I0lorWlxloNI10M1RUuE76aKqqB1dMEsoIBYcDLcj73fstVTP7OzPbmB04m2p6LVPLuWZu4fHRS3/GLT/wDh23/50z/aveyRkjBJG9rmOGQ5pyCPfKjL+TXsdI0EaEpy0jgisqOR+UnK3fTelbFo2zCx6co3UtDGS5kRnklDcgDAL3EgcDjOF8idOT/laAOwk/IJGZif8gAHYSfkFk6eppquPzqWojmjyR1RvDhkd+Qv7NNDTxOmqJWRRsGXPe4NaPxJUReFTjaGm/5wrP8AWlZ3xCf+ZvU//Jo/9cxa21RdSe02/bmt4XssG1F6ff25Xt4XW/S1dLBT/Fz1MUcGAfMe8BmD2OTwvK2/2J7gxl6oHOccACpYST+9Q5vkB/JYkGP/AGbaf9bAs5bPDbsnNbqSok0LCZHwxvLvjKkHqLQc/wA57rE1Ezpd3E0HQHUkcb9AeixM0jpMkbRwB1NuN+w9FKy8c96s1NK6Cpu1FFKzhzH1DGuH4glRHolt0273rrds475X3SxXi1/bFBHW1JlfQOD3MfGHOy5wJHGTwAO5yTr+mNsNC7h7zbov1lYGXI0NbQinLp5Y+jrif1fccM56W9/ZYOrXnK2NozFxaQToLAniAeQ6c1iapxsGN94kjU9ATxsVPcV8sk8jYYbxQySPIa1jahhLj7AA8r6VV0tlC8R1txpad7h1BsszWEj3wStGtHh+2fsN0pbzadFwwVlFMyogl+Knd0SMcHNdhzyDggHkei0ncTRmmdc+JCx2TVdrbX0R0vJL5TpHsHW2eXBywg+p9VnJPPFHdzRmJAGptr1Nvks3yyxsu5ouSANTz8Pkpvpbnba5xZQ3CmqHNGSIpWvIHvwV95ZoYGGSeVkbB3c9wA/eVE188O219usdyqtL0c+l7i2le6O6UlyqI3QdI6gXEvILMgdQI7Z7d1G24Or77qvwgvvd5dUCvZU09NJUvOHVPl1LW+cCAOHY9lhLWyU7Xb1ouGlwsbg25cBb4LCSqfCDvG6gE6Hp4BWb+1LZ/hGl/wAs3/avpFWUk7XuhqoZAwZcWvBDR9cdlTTZrwyQbn7fW/WdXry50UtZJOwwxxh7W+XK5g5Jzz05/NTHprZSHZvQGvHU2p6y6m52eY5nYGGLy4JsYwT36/6gsKetqZmiR0VmEXvmHS40WENVPKA8x2aRe9+y6mP7Vth5Fypf8s3/AGr1Nc1wDmkEEZBHqFQ3YHY1u81vvNfW6wuFtNtqI4mMib1hwcHH1PGOlZuuuW4fho3ZtWn366qdRWmsbA6SmklI64XEtLSxxf5ZBDiOk8jHvhc0eMv3TaiWK0btL3B7OHFaWYm/dtmkjsw8738ldKSspIZBDNVQskdjDXPAJ/JfVU98S9TUxeJPR8Uc8jWOitmWh5AP99yeiuC77p/BSFNV+0SyR2tkNu9dkFTvpHstbKbL5x1lJNIYYaqF8g7ta8Ej8k+MpPP+G+Kh87t5fWOr93dU68NFVUy+IrUMctRK9gbX4a55I/nfZfapqan+XG2D4iTy/j2fJ1nH+4B6Ljbi+aJkmT9T8nHz4LlbiOaNr8vF2Xj5q3j7jb4nujkrqdj2nDmulaCD9Rlfz7Utn+EaX/LN/wBqr7uN4T59ZarvOsm7jVFCK+V1SKZtEXCP5e3V5gz274Cr9sXtBXbx3m8WmTWVTaxaoWSh4iMvmdTsYx1twvk2JVMMwhMOribe8NbeiS108Uoi3WpvbUa2XQqKWKeMSwyskY7s5rgQfzC8hvllEpgN4ohID0lhqGdWfbGe6q1v7War2K2m0jtrp7VErhVSVnxdxjYYp3tbK17Wtw4lgzNzg5PSOcZB8mjfCrt/r/S1NebDuxNWXaZjKqeaJrJBC9/OHx9XW12cjJcCcZWT8SmMu4iju8AEguA4i9h1WTq6UybqNl3AAnUDjyHVW9RaDtBtpeNsrTW22762rdSOqpmyRy1LXgwtDcdIDnv4/MLfJZYoInzzyMjjjaXve8gNa0DJJJ7AKUie57A57cp6cbLvjc5zQ54sei/SLE0Gr9J3R3TbNUWisIPTiCtikOfb5XFZYEOGWkEH1Cza4O1abrIODuCIiL6vqIiIiIiIiIiIiIiIi+FXSCsj8t1RPEPeKQtK0W/7JaX1K/rut71M4Zz0MvErG/uBUgotckMcos8XWD42SCzxdRF/Jc2vPJdfSfc3N6fyXNr/AHvn+k3qXUXP+XUn/WPgtHsVP/AfBQVefCJoa4xuZQal1DQk9j8SJcfk4L56B8JentBaxt2saXV9yq57fKZWxSwtDXktLeSD9VPKLD8row8PEYuO9Y/l9NmDgwXCLGah0xp/VlvNq1LZ6W40hcH+VUMDm9Q9R7FZNapqrQ9w1Dd6W92zXV+sU9JA6BsdE+J0EmTnqfHIxzXO+v0XXLq22XN2f7XTJ+m1r9i1B3hz05azJVaG1TqPTVc84E1LXvewMzkxljjy3t65GO6yWxOrtQ6x0TVS6lqY6qvtlyqLW6oYzo88RBuHuHoT1fn3WMftXu/XzvpLzv5cZLTNlssVJaaenqSw+jZmjLT9QFIektJ2TRNjg0/YaXyaaHLnOccvlkP3pHu7uc48kn+zC4aeAtmD42ZG2Nxca8LaAkddVywxFsgcxuUa3GmvgCQo38KjmnaOnYHAubcazqAPI/Snv7LL+IyspKPZzUIqqhkRqI4oYg44L5DKwho9ycHj6Lz1myD7Vdbtf9s9b3PSVZeHiWop4YoqiiMmcuf5MjThx9wRj04Sg2UrbpXW+67na/u2rJ7a/wA2GlcyOkovMByx5hjA6nN5wST3WtsU7aX2TJrbLe4t0v14a2ssBHKIPZ8utrXuLdL9fJYDfaN8XhcnilY5j2W61Nc1wwQRLBwVtFPHvpPZIYaCbQ9K59MxsU2Kt5j+UYPSRgn6FbNuFoe27jaSrdHXepqKelrjEXyU5AkHRI14x1Aju0enZZ6lgbS00VKwkthY2ME9yAMLoFK4zF1yBlaND0Lvqtu4JlLr6WA07L/VaDtxtbU6Xu9y1pq29/bmqrzhtTWBnRFDE0ANihb+q0AD2z7KM9PbY6e3C3n3QkvlZd4HUVZQiP4C4SUwPXE7PV0H5vujGe3PurHKPHbRS0mqb5qvTuvr5ZqjUMsUtbFTx00kZMbS1uPMicRgE+vqtc9G20bWtu0EkjrcHXXiblYS0zbMa1twDc/A9e0rI6G2o01t9W1FfZK28zy1MXkv+OuMlS0NyDwHng5Hf8VoOt6C9XHxK2OmsWoDZ6n+KsjviBSMqPlE8mW9L+OfdSDbdFaooq+Crqt0r9XQxSNe+mmpqMMlAPLXFsIcAe3BBXqqdB2yq3BpNxX1VSK+jtzrYyEFvlGMvc7qIxnOXH1wspKfPG2NjMoDgeml7ngVk+HOwMY2wBB8+xRJvzpvcKh0XT1d93CqbtpyO50ov1NBQRUjzQuf0vIdH8zhkty3659F6PExRWa3+G2po9PQRw22I0ApWM7CPzmEf2qb7nbqO8W6qtNxhE1JWwPp54ySOuN7S1wyORkE9lodz2SsV32v/uUVt7ub7Wx7HRTZj86NjHh7WA9OCARjJBOPVaqiicRII9c7banUHpc8jf4ha5qVxDwzXM22p5+PIqveyG1+9epNtrbd9HbsustqmkqBDRCWZvllszw7hoxy4E/mrDzWXUen9krzaNWX03m6QWa4+fWkuPm5ZKW8u54aQPyWc240Da9s9I0mjrNVVNTSUb5XskqS0yEySOec9IA7uPosze7VBfbLX2SpkeyG4UstJI5n3mtkYWkjPrgr7R4eKaEcc2Wx1JF7chwX2moxBEOOa1jqSPoqPeHTQW6er7bfJ9v9xDp2GnqY2zxCSRvmuId0n5Ae2D+9ZHY20N1F4gKqm3iu1XV6ltUhbTMnk8xk1RAektc49wAwFoHBx++z+0uzen9nqO40VguNfVsuUrJZDVlhLS0OxjpaP2isbrHw+6T1br6m3IbdbrarzTuieZKKVoEj48BriHNOD0hoOMZA/FRkWDyRQxOtdzTctJ0OvLlcLgjw2SOON3FwOoJ08OSgbxUSMt3iH0hda4mCjjgoHuneMMAZVPLzn6Agn8QrdvvdnFA64/alJ8L5Zk87zm9HTjOc5x2Wubm7U6R3YsjLNqmlfmB/XTVUBDZ6dxx1dBIIwQMEEEHjjIBEHfyHqL4/yP7o9w+w/N6/hPhx5mP8bq6Or69C7RFVUdRI+FgeHkHja3eund1FLM98TMwcb8bWWqeEmKS/b1ak1PRtDaRkNRI4OPzASy/KvrU/+nQ3/l7P/wCAKyO1mz2jtorXLbtMU8sk1S7qqKypLXTy88NJAADR6AAD178rEP2B0s/dr+7Cbrc/tXzhP8P1R+RkQiLGOnqx0jPfutDMMnZTxR6Zg8OK1NoJWwxs5h2YqR6//cNR/wAU/wDsKqJ4Hv8Ayx1f/wAki/1hVwJoxNE+FxID2lpI+oUa7T7C6Y2hudyulhutyq5LnE2KVtU5ha0NdnjpaOfxUhVU0ktVDK3g3NfxC7aiB8lRFI3g29/ELZtf6M0bryyGwa1poJqV7xJGXyBj2PH6zHdwcHBx6HCrbqzwhsstFPeNvNyWQPibJK6OsmEQw0FzGiVpwOeMnA9eFO27myemt5IrbFqK43GlFrMxh+DewdXmdGerra7t0DGMdyobHgcphWfD/wB0mu+xzL1upfhR1luffq6Or69K5MRpXVEn/gDujs1j9/Fc1bA6Z/8A4Q7oc1isr4Rt2tZ63dedL6trXXL7LjbPDWyOLpT1OwWOd6j1Hr39FYW821l5tFdaJZXRsrqaWmc9oBLQ9paSAfbK1bavaXS20Vkms2mTVS/Ey+dPUVTw6WU+mekBoABwMAfvW6rvoIZYaZsdQbu5/S67KOKSKAMmNyq1ReCDS8c/njXl5bl3UeiFgPftnJW2W3wobd0kQZWXXUFW4frGuLP6mhTSi1swmij/AExj19Vg3DqVnBgURfyXNr/e+f6Tev63wvbZMPVHLf2OHZzLrI0j8wpcRbfy6k/6x8Fs9ip/4D4LUNPbZWjTLAy333UMgaMD4m6SS4H/AEltsbPLY1nU53SMZcck/iV+kXSyNsYs0WW9rGsFmhERFmskRERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERERF/9k="""

def ensure_prize_logo():
    try:
        os.makedirs(STATIC_DIR, exist_ok=True)
        logo_path = os.path.join(STATIC_DIR, "logo_prize.jpeg")
        data = base64.b64decode(LOGO_PRIZE_B64)
        if (not os.path.exists(logo_path)) or os.path.getsize(logo_path) != len(data):
            with open(logo_path, "wb") as f:
                f.write(data)
    except Exception as e:
        print("No se pudo crear logo_prize.jpeg:", e)

ensure_prize_logo()


def logo_prize_data_uri():
    """Logo PRIZE embebido para que se vea incluso en login y Render."""
    return "data:image/jpeg;base64," + LOGO_PRIZE_B64.strip()


app = Flask(__name__, static_folder="static")
app.secret_key = os.getenv("SECRET_KEY", "prize-comedor-pro-2026")


@app.errorhandler(500)
def internal_error(e):
    """Evita bucles de redirección cuando una ruta falla en Render.
    Antes se hacía redirect al referrer/dashboard y Chrome terminaba en ERR_TOO_MANY_REDIRECTS.
    Ahora se muestra una pantalla segura con botones para limpiar sesión o volver a iniciar.
    """
    app.logger.exception("Error interno controlado: %s", e)
    try:
        html = f"""
        <div class="login-page">
          <div class="login-card" style="max-width:520px">
            <div class="login-inner">
              <div class="prize-wordmark prize-wordmark-login"><div class="prize-script">Prize<span class="prize-e">e<i></i></span></div><div class="prize-super">SUPERFRUITS</div></div>
              <h2 class="login-title">Sistema Comedor PRIZE</h2>
              <p class="login-subtitle" style="color:#991b1b;font-weight:900">Se detectó un error interno controlado.</p>
              <p style="font-size:13px;color:#cbd5e1;line-height:1.45">No se perdió información. Se corrigió para no generar redirecciones infinitas.</p>
              <div style="display:grid;gap:10px;margin-top:18px">
                <a class="btn btn-blue" href="/logout">Limpiar sesión / volver a ingresar</a>
                <a class="btn" href="/consumos">Ir a Consumos</a>
              </div>
            </div>
          </div>
        </div>
        """
        return render_template_string(BASE_HTML, content=html), 500
    except Exception:
        return "Error interno controlado. Usa /logout para limpiar sesión y vuelve a ingresar.", 500


# =========================
# BASE DE DATOS PERSISTENTE
# Render: PostgreSQL con DATABASE_URL. Local: SQLite de respaldo.
# =========================
def _sql(sql):
    return sql.replace("?", "%s") if USE_POSTGRES else sql

def get_conn():
    if USE_POSTGRES:
        if psycopg2 is None:
            raise RuntimeError("Falta psycopg2-binary en requirements.txt")
        return psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def q_all(sql, params=()):
    with get_conn() as conn:
        if USE_POSTGRES:
            with conn.cursor() as cur:
                cur.execute(_sql(sql), params)
                return cur.fetchall()
        return conn.execute(sql, params).fetchall()

def q_one(sql, params=()):
    rows = q_all(sql, params)
    return rows[0] if rows else None

def q_exec(sql, params=()):
    with get_conn() as conn:
        if USE_POSTGRES:
            with conn.cursor() as cur:
                cur.execute(_sql(sql), params)
                conn.commit()
                return None
        cur = conn.execute(sql, params)
        conn.commit()
        return cur.lastrowid

def audit_event(accion, tabla='', registro_id='', detalle=''):
    try:
        q_exec("INSERT INTO auditoria(usuario,accion,tabla,registro_id,detalle) VALUES(?,?,?,?,?)",
               (session.get('user','sistema'), accion, tabla, str(registro_id or ''), detalle or ''))
    except Exception:
        pass

def init_db():
    if USE_POSTGRES:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                CREATE TABLE IF NOT EXISTS usuarios (
                    id SERIAL PRIMARY KEY,
                    username TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    password_plain TEXT DEFAULT '',
                    role TEXT NOT NULL DEFAULT 'comedor',
                    active INTEGER NOT NULL DEFAULT 1
                );
                CREATE TABLE IF NOT EXISTS trabajadores (
                    id SERIAL PRIMARY KEY,
                    empresa TEXT DEFAULT 'PRIZE',
                    dni TEXT UNIQUE NOT NULL,
                    nombre TEXT NOT NULL,
                    cargo TEXT DEFAULT '',
                    area TEXT DEFAULT '',
                    activo INTEGER NOT NULL DEFAULT 1,
                    creado TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    actualizado TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
                CREATE TABLE IF NOT EXISTS consumos (
                    id SERIAL PRIMARY KEY,
                    fecha TEXT NOT NULL,
                    hora TEXT NOT NULL,
                    dni TEXT NOT NULL,
                    trabajador TEXT DEFAULT '',
                    empresa TEXT DEFAULT 'PRIZE',
                    area TEXT DEFAULT '',
                    tipo TEXT DEFAULT 'Almuerzo',
                    cantidad INTEGER DEFAULT 1,
                    precio_unitario REAL DEFAULT 10,
                    total REAL DEFAULT 10,
                    observacion TEXT DEFAULT '',
                    estado TEXT DEFAULT 'PENDIENTE',
                    creado_por TEXT DEFAULT '',
                    entregado_por TEXT DEFAULT '',
                    entregado_en TEXT DEFAULT '',
                    comedor TEXT DEFAULT 'Comedor 01',
                    fundo TEXT DEFAULT 'Kawsay Allpa',
                    responsable TEXT DEFAULT '',
                    adicional INTEGER DEFAULT 0
                );
                CREATE TABLE IF NOT EXISTS auditoria (
                    id SERIAL PRIMARY KEY,
                    fecha_hora TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    usuario TEXT DEFAULT '',
                    accion TEXT DEFAULT '',
                    tabla TEXT DEFAULT '',
                    registro_id TEXT DEFAULT '',
                    detalle TEXT DEFAULT ''
                );
                CREATE TABLE IF NOT EXISTS cierres (
                    id SERIAL PRIMARY KEY,
                    fecha TEXT UNIQUE NOT NULL,
                    cerrado_por TEXT DEFAULT '',
                    cerrado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    total_consumos INTEGER DEFAULT 0,
                    total_entregados INTEGER DEFAULT 0,
                    total_pendientes INTEGER DEFAULT 0,
                    total_importe REAL DEFAULT 0,
                    archivo_excel TEXT DEFAULT '',
                    correo_destino TEXT DEFAULT '',
                    correo_estado TEXT DEFAULT ''
                );
                CREATE TABLE IF NOT EXISTS importaciones (
                    id SERIAL PRIMARY KEY,
                    fecha_hora TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    archivo TEXT DEFAULT '',
                    total INTEGER DEFAULT 0,
                    creados INTEGER DEFAULT 0,
                    errores INTEGER DEFAULT 0,
                    usuario TEXT DEFAULT ''
                );
                CREATE TABLE IF NOT EXISTS configuracion (
                    clave TEXT PRIMARY KEY,
                    valor TEXT DEFAULT ''
                );
                """)
                for stmt in [
                    "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS password_plain TEXT DEFAULT ''",
                    "ALTER TABLE consumos ADD COLUMN IF NOT EXISTS comedor TEXT DEFAULT 'Comedor 01'",
                    "ALTER TABLE consumos ADD COLUMN IF NOT EXISTS fundo TEXT DEFAULT 'Kawsay Allpa'",
                    "ALTER TABLE consumos ADD COLUMN IF NOT EXISTS responsable TEXT DEFAULT ''",
                    "ALTER TABLE consumos ADD COLUMN IF NOT EXISTS adicional INTEGER DEFAULT 0",
                ]:
                    cur.execute(stmt)
                cur.execute("""
                    DELETE FROM consumos c USING consumos d
                    WHERE COALESCE(c.adicional,0)=0 AND COALESCE(d.adicional,0)=0
                      AND c.fecha=d.fecha AND c.dni=d.dni AND c.id>d.id
                """)
                cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_consumo_unico_dni_fecha ON consumos(fecha, dni) WHERE COALESCE(adicional,0)=0")
                conn.commit()
    else:
        with get_conn() as conn:
            conn.executescript("""
            CREATE TABLE IF NOT EXISTS usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                password_plain TEXT DEFAULT '',
                role TEXT NOT NULL DEFAULT 'comedor',
                active INTEGER NOT NULL DEFAULT 1
            );
            CREATE TABLE IF NOT EXISTS trabajadores (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                empresa TEXT DEFAULT 'PRIZE',
                dni TEXT UNIQUE NOT NULL,
                nombre TEXT NOT NULL,
                cargo TEXT DEFAULT '',
                area TEXT DEFAULT '',
                activo INTEGER NOT NULL DEFAULT 1,
                creado TEXT DEFAULT CURRENT_TIMESTAMP,
                actualizado TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS consumos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT NOT NULL,
                hora TEXT NOT NULL,
                dni TEXT NOT NULL,
                trabajador TEXT DEFAULT '',
                empresa TEXT DEFAULT 'PRIZE',
                area TEXT DEFAULT '',
                tipo TEXT DEFAULT 'Almuerzo',
                cantidad INTEGER DEFAULT 1,
                precio_unitario REAL DEFAULT 10,
                total REAL DEFAULT 10,
                observacion TEXT DEFAULT '',
                estado TEXT DEFAULT 'PENDIENTE',
                creado_por TEXT DEFAULT '',
                entregado_por TEXT DEFAULT '',
                entregado_en TEXT DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS auditoria (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha_hora TEXT DEFAULT CURRENT_TIMESTAMP,
                usuario TEXT DEFAULT '',
                accion TEXT DEFAULT '',
                tabla TEXT DEFAULT '',
                registro_id TEXT DEFAULT '',
                detalle TEXT DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS cierres (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT UNIQUE NOT NULL,
                cerrado_por TEXT DEFAULT '',
                cerrado_en TEXT DEFAULT CURRENT_TIMESTAMP,
                total_consumos INTEGER DEFAULT 0,
                total_entregados INTEGER DEFAULT 0,
                total_pendientes INTEGER DEFAULT 0,
                total_importe REAL DEFAULT 0,
                archivo_excel TEXT DEFAULT '',
                correo_destino TEXT DEFAULT '',
                correo_estado TEXT DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS importaciones (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha_hora TEXT DEFAULT CURRENT_TIMESTAMP,
                archivo TEXT DEFAULT '',
                total INTEGER DEFAULT 0,
                creados INTEGER DEFAULT 0,
                errores INTEGER DEFAULT 0,
                usuario TEXT DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS configuracion (
                clave TEXT PRIMARY KEY,
                valor TEXT DEFAULT ''
            );
            """)
            user_cols = [x["name"] for x in conn.execute("PRAGMA table_info(usuarios)").fetchall()]
            if "password_plain" not in user_cols:
                conn.execute("ALTER TABLE usuarios ADD COLUMN password_plain TEXT DEFAULT ''")
            cols = [x["name"] for x in conn.execute("PRAGMA table_info(consumos)").fetchall()]
            for col, sqltype, default in [("comedor", "TEXT", "'Comedor 01'"), ("fundo", "TEXT", "'Kawsay Allpa'"), ("responsable", "TEXT", "''"), ("adicional", "INTEGER", "0")]:
                if col not in cols:
                    conn.execute(f"ALTER TABLE consumos ADD COLUMN {col} {sqltype} DEFAULT {default}")
            try:
                conn.execute("""
                    DELETE FROM consumos
                    WHERE id NOT IN (
                        SELECT MIN(id) FROM consumos WHERE COALESCE(adicional,0)=0 GROUP BY fecha,dni
                        UNION
                        SELECT id FROM consumos WHERE COALESCE(adicional,0)=1
                    )
                """)
                conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_consumo_unico_dni_fecha ON consumos(fecha, dni) WHERE adicional=0")
            except Exception:
                pass
            conn.commit()

    defaults = {"bloqueo_activo": "0", "hora_inicio": "00:00", "hora_fin": "23:59", "clave_quitar": "1234"}
    for k, v in defaults.items():
        if not q_one("SELECT clave FROM configuracion WHERE clave=?", (k,)):
            q_exec("INSERT INTO configuracion(clave,valor) VALUES(?,?)", (k, v))

    for username, password, role in [("adm", "@123", "admin"), ("adm1", "adm1", "admin"), ("adm2", "adm2", "admin"), ("admin", "admin123", "admin"), ("comedor", "comedor123", "comedor")]:
        existe = q_one("SELECT id FROM usuarios WHERE username=?", (username,))
        if not existe:
            q_exec("INSERT INTO usuarios(username,password_hash,password_plain,role,active) VALUES(?,?,?,?,1)", (username, generate_password_hash(password), password, role))
        elif username in ("adm", "adm1", "adm2"):
            q_exec("UPDATE usuarios SET role='admin', active=1, password_hash=?, password_plain=? WHERE username=?", (generate_password_hash(password), password, username))

    demos = [
        ("PRIZE", "74324033", "AZABACHE LUJAN, OMAR EDUARDO", "OPERARIO", "PRODUCCION"),
        ("PRIZE", "45148597", "CONCEPCION ZAVALETA, VICTOR", "OPERARIO", "PRODUCCION"),
        ("PRIZE", "47625779", "HUAYLLA NACARINO, RAUL", "OPERARIO", "PRODUCCION"),
        ("PRIZE", "41678684", "TANTALLEAN PINILLOS, ERNESTO", "OPERARIO", "PRODUCCION"),
        ("PRIZE", "80503598", "LLANOS VASQUEZ, SEGUNDO", "OPERARIO", "PRODUCCION"),
    ]
    for emp, dni, nom, cargo, area in demos:
        if not q_one("SELECT id FROM trabajadores WHERE dni=?", (dni,)):
            q_exec("INSERT INTO trabajadores(empresa,dni,nombre,cargo,area,activo) VALUES(?,?,?,?,?,1)", (emp, dni, nom, cargo, area))


# =========================
# HELPERS
# =========================
def now_app():
    """Fecha y hora real del sistema en Perú/Lima por defecto.
    En Render evita desfase UTC configurando APP_TIMEZONE=America/Lima.
    """
    return datetime.now(APP_TZ)

def hoy_iso():
    return now_app().date().isoformat()


def fecha_peru_txt(fecha_iso=None):
    f = datetime.strptime(fecha_iso or hoy_iso(), "%Y-%m-%d")
    return f.strftime("%d/%m/%Y")


def periodo_sql(periodo, fecha_iso):
    """Compatibilidad anterior."""
    fecha_iso = fecha_iso or hoy_iso()
    if periodo == "anio":
        return "substr(fecha,1,4)=?", (fecha_iso[:4],)
    if periodo == "mes":
        return "substr(fecha,1,7)=?", (fecha_iso[:7],)
    return "fecha=?", (fecha_iso,)


def rango_sql(fecha_inicio=None, fecha_fin=None):
    """Filtro por rango: fecha desde / hasta."""
    fecha_inicio = fecha_inicio or hoy_iso()
    fecha_fin = fecha_fin or fecha_inicio
    return "fecha BETWEEN ? AND ?", (fecha_inicio, fecha_fin)


def filtro_bar(action, fecha_inicio=None, fecha_fin=None, buscar="", extra_html=""):
    fecha_inicio = fecha_inicio or hoy_iso()
    fecha_fin = fecha_fin or fecha_inicio
    return f"""
    <div class="card filter-card">
      <form method="get" action="{action}" class="filter-grid">
        <div>
          <label>Desde</label>
          <input type="date" name="fecha_inicio" value="{fecha_inicio}">
        </div>
        <div>
          <label>Hasta</label>
          <input type="date" name="fecha_fin" value="{fecha_fin}">
        </div>
        <div>
          <label>Buscar</label>
          <input name="buscar" value="{buscar}" placeholder="DNI, trabajador, área, fundo, comedor...">
        </div>
        <button class="btn-blue">🔍 Filtrar</button>
        <a class="btn" href="{action}">Actualizar</a>
        {extra_html}
      </form>
    </div>
    """


def hora_now():
    return now_app().strftime("%H:%M:%S")


def money(v):
    try:
        return "S/ {:,.2f}".format(float(v or 0))
    except Exception:
        return "S/ 0.00"


def clean_text(v):
    if v is None or (hasattr(pd, "isna") and pd.isna(v)):
        return ""
    return str(v).strip()


def extract_dni(v):
    """Extrae un DNI peruano de 8 digitos desde texto manual, QR o codigo de barras.
    Prioriza numeros asociados a DNI/documento y evita devolver cadenas largas completas.
    """
    raw = str(v or "").strip()
    if not raw:
        return ""
    digits_only = re.sub(r"\D", "", raw)
    if len(digits_only) == 8:
        return digits_only
    if 1 <= len(digits_only) < 8:
        return digits_only.zfill(8)

    txt = raw.upper()
    m = re.search(r"(?:DNI|DOC(?:UMENTO)?|NRO|NUM(?:ERO)?)\D{0,12}(\d{8})(?!\d)", txt)
    if m:
        return m.group(1)
    m = re.search(r"(?<!\d)(\d{8})(?!\d)", txt)
    if m:
        return m.group(1)
    if len(digits_only) > 8:
        return digits_only[-8:]
    return ""

def clean_dni(v):
    return extract_dni(v)


def cfg_get(clave, default=""):
    r = q_one("SELECT valor FROM configuracion WHERE clave=?", (clave,))
    return r["valor"] if r else default

def cfg_set(clave, valor):
    existe = q_one("SELECT clave FROM configuracion WHERE clave=?", (clave,))
    if existe:
        q_exec("UPDATE configuracion SET valor=? WHERE clave=?", (str(valor), clave))
    else:
        q_exec("INSERT INTO configuracion(clave,valor) VALUES(?,?)", (clave, str(valor)))

def registro_bloqueado():
    if cfg_get("bloqueo_activo", "0") != "1":
        return False, ""
    ahora = now_app().strftime("%H:%M")
    inicio = cfg_get("hora_inicio", "00:00")
    fin = cfg_get("hora_fin", "23:59")
    if inicio <= ahora <= fin:
        return False, ""
    return True, f"Registro bloqueado por horario. Horario permitido: {inicio} a {fin}."

def require_remove_key(clave):
    return str(clave or "").strip() == cfg_get("clave_quitar", "1234")

def opciones_comedor():
    return [f"Comedor {i:02d}" for i in range(1, 11)]

def opciones_fundo():
    return ["Kawsay Allpa", "Ayllu Allpa", "Vivadis", "Arena Azul"]


def normalize_columns(cols):
    out = []
    for c in cols:
        x = str(c).strip().upper()
        for a, b in [("Á","A"),("É","E"),("Í","I"),("Ó","O"),("Ú","U"),("Ñ","N")]:
            x = x.replace(a, b)
        x = re.sub(r"[^A-Z0-9]+", "_", x).strip("_")
        out.append(x)
    return out

def col_value(row, *names):
    """Obtiene valores aunque el Excel venga con nombres de columnas distintos.
    Se aceptan reportes de consumo, plantillas y bases exportadas desde otros sistemas.
    """
    aliases = {
        "DNI": [
            "DNI", "DOCUMENTO", "DOCUMENTO_IDENTIDAD", "NUMERO_DOCUMENTO",
            "NUMERO_DE_DOCUMENTO", "NRO_DOCUMENTO", "NRO_DNI", "DOC", "CEDULA",
            "IDENTIFICACION", "NUM_DOC", "DOCUMENTO_NACIONAL_DE_IDENTIDAD"
        ],
        "NOMBRE": [
            "NOMBRE", "NOMBRES", "APELLIDOS_Y_NOMBRES", "APELLIDOS_NOMBRES",
            "APELLIDOS_Y_NOMBRE", "NOMBRE_COMPLETO", "TRABAJADOR", "COLABORADOR",
            "APELLIDOS", "PERSONAL", "EMPLEADO", "NOMBRE_ACTIVIDAD", "NOMBRE__ACTIVIDAD_",
            "NOMBRE_Y_APELLIDOS", "APELLIDOS_Y_NOMBRES_COMPLETOS"
        ],
        "EMPRESA": ["EMPRESA", "RAZON_SOCIAL", "COMPANIA", "CIA", "ORGANIZACION"],
        "CARGO": ["CARGO", "PUESTO", "OCUPACION", "FUNCION", "LABOR", "ACTIVIDAD"],
        "AREA": ["AREA", "AREA_TRABAJO", "SEDE", "FUNDO", "UNIDAD", "DEPARTAMENTO", "CENTRO_COSTO"],
    }
    for name in names:
        for key in aliases.get(name, [name]):
            try:
                val = row.get(key, "")
            except Exception:
                val = ""
            if clean_text(val):
                return val
    return ""


def _normalizar_fila_trabajador(row):
    dni = clean_dni(col_value(row, "DNI"))
    nombre = clean_text(col_value(row, "NOMBRE")).upper()
    if len(dni) != 8 or not nombre:
        return None
    return {
        "empresa": (clean_text(col_value(row, "EMPRESA")) or "PRIZE").upper(),
        "dni": dni,
        "nombre": nombre,
        "cargo": clean_text(col_value(row, "CARGO")).upper(),
        "area": clean_text(col_value(row, "AREA")).upper(),
    }


def _buscar_cabecera_excel(rows_preview):
    """Detecta la fila de cabecera aunque el Excel tenga títulos arriba."""
    mejor_idx = 0
    mejor_score = -1
    for idx, row in enumerate(rows_preview[:25]):
        cols = normalize_columns(row)
        joined = "|".join(cols)
        score = 0
        if any(x in joined for x in ["DNI", "DOCUMENTO", "NRO_DNI", "NUMERO_DOCUMENTO"]):
            score += 3
        if any(x in joined for x in ["NOMBRE", "TRABAJADOR", "COLABORADOR", "APELLIDOS"]):
            score += 3
        if "EMPRESA" in joined:
            score += 1
        if "AREA" in joined or "FUNDO" in joined:
            score += 1
        if score > mejor_score:
            mejor_idx = idx
            mejor_score = score
    return mejor_idx if mejor_score >= 3 else 0


def leer_trabajadores_excel_stream(file_storage):
    """Lee TODO el Excel de trabajadores.
    - Lee todas las hojas.
    - Detecta cabeceras aunque no estén en la primera fila.
    - Acepta columnas de reportes: DNI, TRABAJADOR, EMPRESA, AREA.
    - No exige CARGO ni AREA para no perder trabajadores válidos.
    Devuelve: registros(dict por DNI), total_filas, omitidos.
    """
    filename = (file_storage.filename or "").lower()
    registros = {}
    omitidos = 0
    total = 0
    file_storage.stream.seek(0)

    if filename.endswith(".xlsx"):
        wb = load_workbook(file_storage.stream, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                all_rows = list(ws.iter_rows(values_only=True))
                if not all_rows:
                    continue
                header_idx = _buscar_cabecera_excel(all_rows[:25])
                cols = normalize_columns(all_rows[header_idx])
                for values in all_rows[header_idx + 1:]:
                    if not values or not any(clean_text(v) for v in values):
                        continue
                    total += 1
                    r = dict(zip(cols, values))
                    item = _normalizar_fila_trabajador(r)
                    if not item:
                        omitidos += 1
                        continue
                    registros[item["dni"]] = item
        finally:
            wb.close()
        return registros, total, omitidos

    # Para .xls u otros casos: intentar con pandas si el ambiente tiene motor disponible.
    file_storage.stream.seek(0)
    try:
        hojas = pd.read_excel(file_storage, dtype=str, sheet_name=None).items()
    except Exception:
        file_storage.stream.seek(0)
        hojas = [("Hoja1", pd.read_excel(file_storage, dtype=str))]

    for _, df in hojas:
        df = df.fillna("")
        if df.empty:
            continue
        df.columns = normalize_columns(df.columns)
        for _, r in df.iterrows():
            total += 1
            item = _normalizar_fila_trabajador(r)
            if not item:
                omitidos += 1
                continue
            registros[item["dni"]] = item
    return registros, total, omitidos


def reemplazar_trabajadores_batch(registros):
    """Reemplaza la tabla trabajadores en UNA sola conexión y por lotes.
    Evita abrir miles de conexiones en Render y evita SIGKILL por memoria/tiempo.
    """
    data = [(r["empresa"], r["dni"], r["nombre"], r["cargo"], r["area"]) for r in registros]
    if not data:
        return 0

    with get_conn() as conn:
        if USE_POSTGRES:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM trabajadores")
                psycopg2.extras.execute_batch(
                    cur,
                    """
                    INSERT INTO trabajadores(empresa,dni,nombre,cargo,area,activo)
                    VALUES(%s,%s,%s,%s,%s,1)
                    """,
                    data,
                    page_size=500,
                )
                conn.commit()
        else:
            conn.execute("DELETE FROM trabajadores")
            conn.executemany(
                "INSERT INTO trabajadores(empresa,dni,nombre,cargo,area,activo) VALUES(?,?,?,?,?,1)",
                data,
            )
            conn.commit()
    return len(data)


def dia_cerrado(fecha_iso=None):
    return q_one("SELECT * FROM cierres WHERE fecha=?", (fecha_iso or hoy_iso(),))


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get("user"):
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper


def roles_required(*roles):
    def deco(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            role = session.get("role")
            if role == "admin" or role in roles:
                return fn(*args, **kwargs)
            flash("No tienes permiso para esta opción.", "error")
            return redirect(url_for("dashboard"))
        return wrapper
    return deco


def asegurar_rol_usuario(role):
    return "admin" if role == "admin" else "comedor"


def send_report_email(to_email, subject, body, attachment_path):
    host = os.getenv("SMTP_HOST", "").strip()
    user = os.getenv("SMTP_USER", "").strip()
    password = os.getenv("SMTP_PASSWORD", "").strip()
    port = int(os.getenv("SMTP_PORT", "587"))
    sender = os.getenv("SMTP_FROM", user or "no-reply@prize.local")

    if not host or not user or not password or not to_email:
        note = os.path.join(REPORT_DIR, f"correo_no_enviado_{now_app().strftime('%Y%m%d_%H%M%S')}.txt")
        with open(note, "w", encoding="utf-8") as f:
            f.write("SMTP no configurado. El Excel fue generado correctamente.\n\n")
            f.write(f"Para: {to_email}\nAsunto: {subject}\nAdjunto: {attachment_path}\n\n{body}")
        return "NO ENVIADO - SMTP NO CONFIGURADO"

    msg = EmailMessage()
    msg["From"] = sender
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.set_content(body)

    with open(attachment_path, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=os.path.basename(attachment_path),
        )

    with smtplib.SMTP(host, port, timeout=30) as smtp:
        smtp.starttls()
        smtp.login(user, password)
        smtp.send_message(msg)

    return "ENVIADO"


# =========================
# UI HTML + CSS PRO

def send_admin_user_notice(username, role, action="creado"):
    """Notificación opcional y segura al administrador.
    No envía contraseñas por correo ni guarda claves en texto plano.
    Actívalo en Render con ENABLE_ADMIN_USER_ALERTS=1 y variables SMTP.
    """
    destino = os.getenv("ADMIN_AUDIT_EMAIL", "omar.azabache24@gmail.com").strip()
    if os.getenv("ENABLE_ADMIN_USER_ALERTS", "0").strip() != "1":
        try:
            note = os.path.join(REPORT_DIR, "notificaciones_usuarios.txt")
            with open(note, "a", encoding="utf-8") as f:
                f.write(f"{now_app():%Y-%m-%d %H:%M:%S} | Usuario {action}: {username} | Rol: {role}\n")
        except Exception:
            pass
        return "DESACTIVADO"

    host = os.getenv("SMTP_HOST", "").strip()
    user = os.getenv("SMTP_USER", "").strip()
    password = os.getenv("SMTP_PASSWORD", "").strip()
    port = int(os.getenv("SMTP_PORT", "587"))
    sender = os.getenv("SMTP_FROM", user or "no-reply@prize.local")
    if not host or not user or not password or not destino:
        return "SMTP NO CONFIGURADO"

    msg = EmailMessage()
    msg["From"] = sender
    msg["To"] = destino
    msg["Subject"] = f"Sistema Comedor - usuario {action}"
    msg.set_content(
        "Notificación de seguridad del Sistema Comedor.\n\n"
        f"Acción: Usuario {action}\n"
        f"Usuario: {username}\n"
        f"Rol: {role}\n"
        f"Fecha/hora: {now_app():%d/%m/%Y %H:%M:%S}\n\n"
        "Por seguridad no se envían contraseñas por correo."
    )
    with smtplib.SMTP(host, port, timeout=30) as smtp:
        smtp.starttls()
        smtp.login(user, password)
        smtp.send_message(msg)
    return "ENVIADO"
# =========================
BASE_HTML = r"""
<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Sistema Comedor PRIZE</title>
<style>
:root{
  --navy:#061b2b;
  --navy2:#062338;
  --blue:#0d73b8;
  --green:#17a34a;
  --green2:#0f8a3a;
  --orange:#ff6b14;
  --purple:#7c3aed;
  --bg:#f6f9fc;
  --card:#ffffff;
  --line:#e8eef5;
  --text:#142238;
  --muted:#6b7b90;
  --shadow:0 10px 28px rgba(15,35,55,.08);
  --shadow2:0 18px 50px rgba(15,35,55,.14);
}
*{box-sizing:border-box}
body{
  margin:0;
  font-family:"Segoe UI",Arial,sans-serif;
  background:var(--bg);
  color:var(--text);
}
a{text-decoration:none;color:inherit}
button,.btn{
  border:0;
  border-radius:10px;
  padding:12px 18px;
  background:linear-gradient(135deg,var(--green),var(--green2));
  color:white;
  font-weight:800;
  cursor:pointer;
  display:inline-block;
  box-shadow:0 7px 18px rgba(22,163,74,.20);
}
.btn-blue{background:linear-gradient(135deg,#1480c8,#075f9e);box-shadow:0 7px 18px rgba(20,128,200,.18)}
.btn-orange{background:linear-gradient(135deg,#ff7a1a,#ff5b0a);box-shadow:0 7px 18px rgba(255,107,20,.22)}
.btn-red{background:linear-gradient(135deg,#ef4444,#b91c1c)}
input,select,textarea{
  width:100%;
  border:1px solid #dce6f0;
  background:white;
  border-radius:10px;
  padding:12px 14px;
  outline:none;
  color:#25364a;
}
input:focus,select:focus,textarea:focus{
  border-color:#70b7e9;
  box-shadow:0 0 0 4px rgba(13,115,184,.09)
}
.muted{color:var(--muted)}
.small{font-size:12px}
.flash{
  margin:0 0 12px;
  padding:13px 16px;
  border-radius:12px;
  border:1px solid #c7d2fe;
  background:#eef2ff;
  color:#1e3a8a;
  font-weight:700;
}
.flash.error{border-color:#fecaca;background:#fff1f2;color:#991b1b}
.flash.ok{border-color:#bbf7d0;background:#f0fdf4;color:#166534}

/* LOGIN EXACTO ESTILO IMAGEN */
.login-page{
  min-height:100vh;
  display:grid;
  place-items:center;
  padding:24px;
  background:
    radial-gradient(circle at 12% 90%, rgba(13,115,184,.14) 0 18%, transparent 19%),
    radial-gradient(circle at 92% 96%, rgba(22,163,74,.18) 0 22%, transparent 23%),
    linear-gradient(135deg,#f8fbff,#ffffff 52%,#f4fff7);
}
.login-card{
  width:min(430px,94vw);
  background:white;
  border:1px solid var(--line);
  border-radius:18px;
  overflow:hidden;
  box-shadow:var(--shadow2);
  position:relative;
}
.login-card:before{
  content:"";
  position:absolute;left:-55px;bottom:-78px;
  width:270px;height:150px;
  background:#0d5f9b;
  border-radius:50% 50% 0 0;
  transform:rotate(8deg);
}
.login-card:after{
  content:"";
  position:absolute;right:-70px;bottom:-86px;
  width:300px;height:160px;
  background:linear-gradient(135deg,#0b7a36,#2fac57);
  border-radius:55% 45% 0 0;
  transform:rotate(-8deg);
}
.login-inner{
  position:relative;
  z-index:2;
  padding:36px 42px 58px;
  text-align:center;
}
.logo-word{
  display:inline-flex;
  align-items:flex-end;
  gap:0;
  margin:0 auto 14px;
  font-weight:900;
  letter-spacing:-5px;
  font-size:76px;
  line-height:.84;
  color:#07325d;
  font-style:italic;
}
.logo-word .e{
  color:#ff6b14;
  position:relative;
  border:5px solid #0d73b8;
  border-radius:50%;
  width:58px;height:58px;
  display:inline-grid;
  place-items:center;
  font-size:44px;
  letter-spacing:-3px;
  font-style:normal;
  margin-left:1px;
}
.logo-word .leaf{
  position:absolute;
  width:18px;height:34px;
  background:#16a34a;
  border-radius:100% 0 100% 0;
  transform:rotate(38deg);
  top:-34px;right:3px;
}
.login-title{font-size:18px;margin:6px 0 4px;font-weight:900}
.login-subtitle{font-size:13px;margin:0 0 26px;color:var(--muted);font-weight:650}
.form-label{text-align:left;font-weight:850;font-size:13px;margin:14px 0 7px}
.input-icon{position:relative}
.input-icon span{position:absolute;left:13px;top:50%;transform:translateY(-50%);color:#91a4b7}
.input-icon input{padding-left:42px}
.login-button{width:100%;margin-top:22px;font-size:15px}
.demo-users{font-size:11px;color:#7b8ca2;margin-top:26px;line-height:1.6}

/* APP HEADER */
.app-shell{min-height:100vh}
.hero{
  margin:0;
  background:white;
  border-bottom:1px solid var(--line);
  display:grid;
  grid-template-columns:310px 1fr 330px;
  align-items:center;
  gap:22px;
  padding:18px 26px;
  box-shadow:0 2px 12px rgba(15,35,55,.04);
}
.hero-brand{
  border-right:1px solid var(--line);
  min-height:118px;
  display:flex;
  align-items:center;
  justify-content:center;
  flex-direction:column;
}
.hero-brand .logo-word{font-size:80px;margin-bottom:5px}
.superfruits{
  color:#16a34a;
  font-weight:900;
  letter-spacing:1px;
  border-top:3px solid #16a34a;
  padding-top:2px;
}
.hero h1{font-size:34px;letter-spacing:-.5px;margin:0 0 8px}
.hero p{font-size:17px;margin:0;color:#52647c}
.checks{
  display:grid;
  grid-template-columns:1fr;
  gap:8px;
  font-weight:800;
  color:#27384d;
}
.checks div:before{content:"✓";color:white;background:#16a34a;border-radius:50%;padding:1px 5px;margin-right:10px}
.demo-box{
  background:linear-gradient(135deg,#041727,#082d45);
  color:white;
  border-radius:10px;
  padding:16px 18px;
  box-shadow:var(--shadow);
  line-height:1.75;
  font-weight:750;
}
.demo-box b{display:block;margin-bottom:6px;font-size:16px}

.main-layout{
  display:grid;
  grid-template-columns:185px 1fr 320px;
  gap:18px;
  padding:18px;
}
.sidebar{
  background:linear-gradient(180deg,#082f49,#061727);
  color:white;
  border-radius:8px;
  padding:12px 10px;
  min-height:calc(100vh - 175px);
  box-shadow:var(--shadow2);
}
.side-logo{text-align:center;border-bottom:1px solid rgba(255,255,255,.12);padding:8px 0 12px}
.side-logo .logo-word{font-size:43px;letter-spacing:-3px;margin:0;color:white}
.side-logo .logo-word .e{width:34px;height:34px;font-size:26px;border-width:3px}
.side-logo .logo-word .leaf{width:10px;height:20px;top:-21px}
.side-logo small{display:block;color:#d3e7f5;font-weight:750;margin-top:4px}
.nav{padding-top:10px}
.nav a{
  display:flex;
  align-items:center;
  gap:10px;
  padding:11px 10px;
  margin:4px 0;
  border-radius:7px;
  color:#e6f2fb;
  font-weight:850;
  font-size:13px;
}
.nav a:hover,.nav a.on{background:linear-gradient(90deg,#138a43,#0871b6)}
.nav .pill{
  margin-left:auto;background:#7c3aed;color:white;border-radius:999px;font-size:11px;padding:2px 7px
}
.content{min-width:0}
.topbar{
  display:flex;
  align-items:flex-start;
  justify-content:space-between;
  margin-bottom:18px;
}
.topbar h2{font-size:25px;margin:0 0 5px;letter-spacing:-.3px}
.user-chip{
  display:flex;align-items:center;gap:12px;color:#4b5d73;font-weight:800;
}
.avatar{
  width:38px;height:38px;border-radius:50%;background:#dcfce7;color:#15803d;
  display:grid;place-items:center;font-size:20px;
}
.kpi-grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:16px;margin-bottom:18px}
.card{
  background:white;border:1px solid var(--line);border-radius:12px;padding:18px;box-shadow:var(--shadow);
}
.kpi-card{display:flex;align-items:center;gap:18px;min-height:105px}
.icon-circle{
  width:58px;height:58px;border-radius:50%;display:grid;place-items:center;font-size:28px;font-weight:900;
}
.ic-green{background:#eaf8ee;color:#16a34a}
.ic-blue{background:#e9f4ff;color:#0d73b8}
.ic-purple{background:#f4ecff;color:#7c3aed}
.ic-orange{background:#fff2e8;color:#ff6b14}
.kpi-card .label{font-size:13px;color:#6b7b90;font-weight:800}
.kpi-card .num{font-size:26px;font-weight:950;color:#102033;line-height:1.1}
.kpi-card .sub{font-size:12px;color:#6b7b90;font-weight:750}

.table-head{
  display:flex;justify-content:space-between;align-items:center;margin-bottom:14px
}
.table-head h3{margin:0;font-size:18px}
.table-wrap{overflow:auto;border:1px solid var(--line);border-radius:10px}
table{width:100%;border-collapse:collapse;background:white}
th,td{
  padding:12px 13px;border-bottom:1px solid #edf2f7;
  font-size:13px;text-align:left;white-space:nowrap;color:#23364d;
}
th{background:#f8fafc;color:#334155;font-weight:900}
tr:last-child td{border-bottom:0}
.badge{
  display:inline-flex;align-items:center;gap:6px;
  border-radius:999px;padding:6px 11px;font-size:12px;font-weight:950;
}
.badge.ok{background:#dcfce7;color:#16803a}
.badge.warn{background:#fff3cd;color:#b45309}
.badge.off{background:#fee2e2;color:#991b1b}
.form-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:12px}
.form-grid.two{grid-template-columns:1fr auto}
.panel-right{display:flex;flex-direction:column;gap:18px}
.status-box .status-inner{
  border:1px solid var(--line);border-radius:10px;padding:18px;margin-top:10px;background:#fff
}
.quick a{
  display:flex;align-items:center;gap:12px;
  padding:14px;border:1px solid var(--line);border-bottom:0;font-weight:900;color:#1f4e78;background:white
}
.quick a:first-child{border-radius:10px 10px 0 0}
.quick a:last-child{border-bottom:1px solid var(--line);border-radius:0 0 10px 10px}
.mini-kpis{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin:14px 0}
.mini-kpis .card b{display:block;font-size:22px;margin-top:8px}
.user-row{
  display:flex;justify-content:space-between;align-items:center;
  padding:15px;border:1px solid var(--line);border-radius:10px;margin-bottom:10px;background:#fff
}
.footer{
  background:#061727;color:#d8e8f2;
  padding:18px 28px;display:flex;justify-content:space-between;font-size:13px
}

@media(max-width:1200px){
  .hero{grid-template-columns:1fr}
  .hero-brand{border-right:0;border-bottom:1px solid var(--line)}
  .main-layout{grid-template-columns:1fr}
  .sidebar{min-height:auto}
  .nav{display:grid;grid-template-columns:repeat(2,1fr);gap:5px}
  .panel-right{display:grid;grid-template-columns:1fr 1fr}
}

.admin-actions{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:14px}
.ind-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:16px}
@media(max-width:900px){
  .hero{padding:14px}
  .checks{grid-template-columns:1fr!important}
  .main-layout{padding:10px;gap:10px}
  .content,.panel-right,.sidebar{width:100%}
  .ind-grid{grid-template-columns:1fr 1fr}
  .form-grid{grid-template-columns:1fr!important}
  .kpi-grid{grid-template-columns:1fr 1fr}
  .mini-kpis{grid-template-columns:1fr 1fr}
  .topbar{display:block}
  .user-chip{margin-top:10px}
  th,td{font-size:12px;padding:10px}
}
@media(max-width:520px){
  .kpi-grid,.mini-kpis,.ind-grid{grid-template-columns:1fr}
  .hero-brand .logo-word{font-size:52px}
  .hero h1{font-size:22px}
  .demo-box{font-size:13px}
  .nav{grid-template-columns:1fr}
  .card{padding:14px}
  button,.btn{width:100%;text-align:center}
}

@media(max-width:800px){
  .kpi-grid,.mini-kpis,.form-grid{grid-template-columns:1fr}
  .panel-right{display:block}
  .login-inner{padding:30px 26px 58px}
  .hero h1{font-size:26px}
}

/* ===== AJUSTE FINO UI PRIZE ===== */
body{overflow-x:hidden}
.app-shell{max-width:1920px;margin:0 auto;background:var(--bg)}
.hero{
  position:sticky;
  top:0;
  z-index:20;
  min-height:112px;
  grid-template-columns:280px minmax(360px,1fr) 280px!important;
  padding:14px 22px!important;
}
.hero-brand{min-height:88px!important}
.hero-brand .logo-word{font-size:62px!important}
.hero h1{font-size:30px!important;margin-bottom:4px!important}
.hero p{font-size:15px!important}
.demo-box{
  max-width:280px;
  justify-self:end;
  padding:14px 16px!important;
  font-size:14px;
  line-height:1.55!important;
}
section .checks{
  max-width:1220px;
  margin:0 auto;
  grid-template-columns:repeat(3,minmax(240px,1fr))!important;
  gap:8px 28px!important;
}
.checks div{
  white-space:nowrap;
  font-size:14px;
}
.checks div:before{
  display:inline-grid;
  place-items:center;
  width:22px;height:22px;
  padding:0!important;
  margin-right:8px!important;
}
.main-layout{
  max-width:1540px;
  margin:0 auto;
  grid-template-columns:190px minmax(0,1fr) 300px!important;
  align-items:start;
}
.sidebar{
  position:sticky;
  top:126px;
  min-height:auto!important;
  height:calc(100vh - 140px);
  overflow:auto;
}
.panel-right{
  position:sticky;
  top:126px;
}
.content{
  min-height:calc(100vh - 240px);
}
.card{
  border-radius:16px!important;
}
.table-wrap{
  max-width:100%;
}
.table-wrap table{
  min-width:980px;
}
.form-grid{
  align-items:end;
}
.topbar{
  min-height:58px;
}
.kpi-grid{
  grid-template-columns:repeat(4,minmax(180px,1fr))!important;
}
.kpi-card{
  min-height:92px!important;
}
.icon-circle{
  flex:0 0 auto;
}
.quick a{
  min-height:48px;
}
input,select,textarea{
  min-height:46px;
}
button,.btn{
  min-height:46px;
}
.login-card{
  transform:none!important;
}
@media(max-width:1350px){
  .hero{grid-template-columns:220px 1fr!important}
  .demo-box{display:none}
  .main-layout{grid-template-columns:185px minmax(0,1fr)!important}
  .panel-right{position:static;display:grid;grid-template-columns:1fr 1fr;grid-column:1 / -1}
}
@media(max-width:980px){
  .hero{position:relative;grid-template-columns:1fr!important;text-align:center}
  .hero-brand{border-right:0!important}
  section .checks{grid-template-columns:1fr!important;padding:10px 14px!important}
  .checks div{white-space:normal}
  .main-layout{grid-template-columns:1fr!important;padding:12px!important}
  .sidebar{position:relative;top:0;height:auto;min-height:auto}
  .nav{grid-template-columns:repeat(2,1fr)!important}
  .panel-right{display:grid;grid-template-columns:1fr!important}
  .kpi-grid{grid-template-columns:1fr 1fr!important}
}
@media(max-width:640px){
  .hero-brand .logo-word{font-size:50px!important}
  .hero h1{font-size:22px!important}
  .nav{grid-template-columns:1fr!important}
  .kpi-grid{grid-template-columns:1fr!important}
  .topbar{display:block!important}
  .user-chip{margin-top:12px}
  .form-grid{grid-template-columns:1fr!important}
  .card{padding:14px!important}
  .main-layout{padding:8px!important}
  .table-head{display:block!important}
  .table-head .btn,.table-head a{margin-top:8px}
}


/* ===== AJUSTE SCROLL LIMPIO ===== */
html,body{max-width:100%;overflow-x:hidden!important}
.main-layout{overflow:visible!important}
.content{overflow:hidden!important}
.panel-right{overflow:visible!important}
.sidebar{overflow-y:auto!important;overflow-x:hidden!important}
.table-wrap{
  overflow:auto!important;
  max-height:520px;
  scrollbar-width:thin;
}
.table-wrap table{min-width:900px}
.table-wrap th{
  position:sticky;
  top:0;
  z-index:2;
}
@media(max-width:980px){
  .table-wrap{max-height:460px}
}


/* ===== LAYOUT FIJO CON SCROLL INTERNO ===== */
body{height:100vh;overflow:hidden!important;background:#eef4f8!important}
.app-shell{height:100vh;max-width:none!important;width:100%!important;display:grid;grid-template-rows:auto 1fr auto}
.hero{
  position:relative!important;
  top:auto!important;
  z-index:5;
  border-radius:0!important;
  margin:0!important;
}
.main-layout{
  width:100%!important;
  max-width:none!important;
  margin:0!important;
  padding:0!important;
  gap:0!important;
  grid-template-columns:240px minmax(0,1fr) 310px!important;
  min-height:0!important;
  height:100%!important;
  overflow:hidden!important;
}
.sidebar{
  position:relative!important;
  top:auto!important;
  left:0!important;
  height:100%!important;
  min-height:0!important;
  border-radius:0!important;
  margin:0!important;
  width:240px!important;
  padding:18px 14px!important;
  overflow-y:auto!important;
  overflow-x:hidden!important;
}
.content{
  height:100%!important;
  min-height:0!important;
  overflow-y:auto!important;
  overflow-x:hidden!important;
  padding:18px 18px 28px!important;
  background:#eef4f8;
}
.panel-right{
  position:relative!important;
  top:auto!important;
  height:100%!important;
  min-height:0!important;
  overflow-y:auto!important;
  overflow-x:hidden!important;
  padding:18px 14px 28px 0!important;
  background:#eef4f8;
}
.footer{display:none!important}
.table-wrap{
  overflow:auto!important;
  max-height:calc(100vh - 420px)!important;
  min-height:220px;
}
.table-wrap table{min-width:1050px}
.filter-card{
  margin-bottom:14px!important;
  padding:14px!important;
}
.filter-grid{
  display:grid;
  grid-template-columns:160px 160px minmax(260px,1fr) 130px 130px;
  gap:10px;
  align-items:end;
}
.filter-grid label{
  display:block;
  font-size:12px;
  font-weight:900;
  color:#64748b;
  margin:0 0 5px;
}
.filter-grid button,.filter-grid .btn{height:46px;display:grid;place-items:center}
.topbar{margin-bottom:14px!important}
.card{box-shadow:0 10px 24px rgba(15,35,55,.07)!important}
@media(max-width:1350px){
  .main-layout{grid-template-columns:230px minmax(0,1fr)!important}
  .panel-right{display:none!important}
}
@media(max-width:900px){
  body{overflow:auto!important;height:auto}
  .app-shell{height:auto;display:block}
  .main-layout{display:block!important;height:auto!important;overflow:visible!important}
  .sidebar{width:100%!important;height:auto!important;border-radius:0!important}
  .content{height:auto!important;overflow:visible!important;padding:12px!important}
  .filter-grid{grid-template-columns:1fr!important}
  .table-wrap{max-height:460px!important}
}



/* ===== CORRECCIÓN SOLICITADA: SIN USUARIOS DEMO, SIN LOGO LATERAL, TÍTULO CENTRADO GRANDE ===== */
.demo-box,.demo-users{display:none!important;}
.hero{
  grid-template-columns:1fr!important;
  text-align:center!important;
  justify-items:center!important;
  min-height:118px!important;
  padding:22px 26px!important;
  background:linear-gradient(135deg,#061b2b,#082f49)!important;
  color:white!important;
}
.hero-brand{display:none!important;}
.hero h1{
  font-size:46px!important;
  line-height:1.08!important;
  margin:0 0 6px!important;
  font-weight:950!important;
  letter-spacing:.2px!important;
  color:white!important;
}
.hero p{
  font-size:22px!important;
  color:#d8e8f2!important;
  font-weight:800!important;
}
.side-logo{display:none!important;}
.side-title{
  text-align:center;
  color:#d8e8f2;
  font-weight:950;
  font-size:16px;
  letter-spacing:.7px;
  padding:10px 4px 18px;
  border-bottom:1px solid rgba(255,255,255,.15);
  margin-bottom:10px;
}
@media(max-width:900px){
  .hero h1{font-size:32px!important;}
  .hero p{font-size:17px!important;}
}


/* ===== PANEL LATERAL FIJO ESTILO IMAGEN ADJUNTA ===== */
:root{--side-w:185px;}
.app-shell{display:block!important;height:100vh!important;overflow:hidden!important;background:#eef4f8!important;}
.hero{margin-left:var(--side-w)!important;width:calc(100% - var(--side-w))!important;min-height:96px!important;padding:18px 24px!important;border-bottom:1px solid rgba(255,255,255,.10)!important;}
.hero h1{font-size:38px!important;}.hero p{font-size:18px!important;}
.main-layout{display:grid!important;grid-template-columns:minmax(0,1fr) 310px!important;margin-left:var(--side-w)!important;width:calc(100% - var(--side-w))!important;height:calc(100vh - 96px)!important;min-height:0!important;overflow:hidden!important;}
.fixed-prize-sidebar{position:fixed!important;inset:0 auto 0 0!important;width:var(--side-w)!important;height:100vh!important;border-radius:0!important;padding:12px 10px!important;overflow-y:auto!important;overflow-x:hidden!important;background:radial-gradient(circle at 84% 92%, rgba(19,119,88,.18), transparent 32%),linear-gradient(180deg,#05243a 0%,#041827 55%,#03131f 100%)!important;box-shadow:8px 0 28px rgba(0,0,0,.18)!important;z-index:50!important;}
.side-logo-pro{text-align:center;padding:0 2px 12px;}.brand-prize{position:relative;display:inline-block;color:#fff;font-size:48px;line-height:.9;font-weight:900;font-style:italic;letter-spacing:-3px;font-family:"Segoe Script","Segoe UI",Arial,sans-serif;}.brand-prize span{display:inline-grid;place-items:center;width:34px;height:34px;margin-left:0;border:3px solid #25a8e0;border-radius:50%;color:#ff8a1d;font-size:28px;font-style:normal;letter-spacing:-2px;background:rgba(255,255,255,.03);}.brand-prize i{position:absolute;right:5px;top:-22px;width:10px;height:25px;background:#2dbb52;border-radius:100% 0 100% 0;transform:rotate(34deg);}.brand-sub{display:inline-block;color:#3ac35b;font-size:11px;line-height:1;font-weight:900;letter-spacing:.4px;border-top:1px solid #2dbb52;border-bottom:1px solid #2dbb52;padding:2px 8px;margin-top:4px;}
.side-user-card{text-align:center;padding:12px 4px 13px;border-bottom:1px solid rgba(255,255,255,.12);margin-bottom:10px;}.side-avatar{width:44px;height:44px;border-radius:50%;display:grid;place-items:center;margin:0 auto 9px;background:#fff;color:#7dbd69;font-size:26px;box-shadow:0 10px 20px rgba(0,0,0,.18);}.side-user-title{color:#fff;font-size:13px;font-weight:900;margin-bottom:4px;}.side-user-sub{color:#d8e7ef;font-size:11px;font-weight:700;}.side-title{display:none!important;}
.nav-pro{padding:0!important;}.nav-pro a{position:relative;min-height:38px;display:flex!important;align-items:center!important;gap:8px!important;margin:5px 0!important;padding:10px 8px!important;border-radius:9px!important;color:#eaf6ff!important;font-size:12px!important;font-weight:900!important;letter-spacing:-.1px;transition:all .15s ease;}.nav-pro a:hover,.nav-pro a.on{background:linear-gradient(90deg,#165c44,#0e734c)!important;box-shadow:inset 0 0 0 1px rgba(255,255,255,.04),0 8px 18px rgba(0,0,0,.18)!important;transform:translateX(1px);}.nav-ico{width:18px;display:inline-grid;place-items:center;font-size:14px;flex:0 0 18px;}.nav-pro .pill{margin-left:auto!important;color:#fff!important;border-radius:999px!important;font-size:9px!important;padding:2px 6px!important;line-height:1.2!important;}.nav-pro .pill.nuevo{background:#35b94b!important;}.nav-pro .pill.correo{background:#318aca!important;}.logout-link{margin-top:8px!important;}
.side-slogan-card{margin-top:22px;border:1px solid rgba(255,255,255,.20);border-radius:9px;min-height:118px;padding:22px 16px 15px;color:#fff;font-size:12px;line-height:1.35;background:linear-gradient(135deg,rgba(255,255,255,.03),rgba(255,255,255,.01));position:relative;overflow:hidden;}.side-slogan-card:after{content:"";position:absolute;right:-18px;bottom:-26px;width:90px;height:130px;border:2px solid rgba(255,255,255,.06);border-radius:70% 0 70% 0;transform:rotate(26deg);}.side-slogan-card b{font-weight:500;}.leaf-icon{color:#51d05e;font-size:28px;line-height:1;margin-bottom:20px;transform:rotate(-28deg);}
.content{height:100%!important;overflow-y:auto!important;overflow-x:hidden!important;padding:18px 18px 28px!important;}.panel-right{height:100%!important;overflow-y:auto!important;padding:18px 14px 28px 0!important;}
@media(max-width:1350px){.main-layout{grid-template-columns:minmax(0,1fr)!important;}.panel-right{display:none!important;}}
@media(max-width:760px){:root{--side-w:168px;}.brand-prize{font-size:41px;}.brand-prize span{width:30px;height:30px;font-size:24px;}.hero h1{font-size:25px!important;}.hero p{font-size:14px!important;}.nav-pro a{font-size:11px!important;padding:9px 7px!important;}}


/* =========================================================
   MEJORA RESPONSIVE CELULAR - PANEL COMPACTO / PROCESOS CLAROS
   ========================================================= */
@media(max-width: 780px){
  :root{--side-w:0px!important;}
  html,body{height:auto!important;overflow:auto!important;background:#eef4f8!important;}
  .app-shell{height:auto!important;min-height:100vh!important;overflow:visible!important;display:block!important;}
  .hero{margin-left:0!important;width:100%!important;min-height:auto!important;padding:18px 14px!important;position:relative!important;border-bottom:0!important;}
  .hero h1{font-size:28px!important;line-height:1.05!important;max-width:320px!important;margin:0 auto 6px!important;}
  .hero p{font-size:14px!important;line-height:1.25!important;max-width:320px!important;margin:0 auto!important;}
  .main-layout{display:block!important;margin-left:0!important;width:100%!important;height:auto!important;min-height:0!important;overflow:visible!important;}
  .fixed-prize-sidebar{position:relative!important;inset:auto!important;width:100%!important;height:auto!important;min-height:0!important;padding:8px!important;border-radius:0!important;box-shadow:none!important;background:linear-gradient(180deg,#041827,#061b2b)!important;overflow:visible!important;}
  .side-logo-pro,.side-user-card,.side-slogan-card,.side-title{display:none!important;}
  .nav-pro{display:grid!important;grid-template-columns:repeat(2,minmax(0,1fr))!important;gap:7px!important;padding:0!important;}
  .nav-pro a{margin:0!important;min-height:42px!important;justify-content:center!important;text-align:center!important;padding:9px 7px!important;font-size:11px!important;border-radius:10px!important;background:rgba(255,255,255,.06)!important;border:1px solid rgba(255,255,255,.08)!important;}
  .nav-pro a.on,.nav-pro a:hover{background:linear-gradient(90deg,#16834d,#0d73b8)!important;transform:none!important;}
  .nav-ico{font-size:13px!important;width:auto!important;flex:0 0 auto!important;}
  .nav-pro .pill{display:none!important;}
  .content{height:auto!important;min-height:0!important;overflow:visible!important;padding:12px 10px 28px!important;background:#eef4f8!important;}
  .panel-right{display:none!important;}
  .topbar{display:block!important;min-height:0!important;margin-bottom:10px!important;}
  .topbar h2{font-size:24px!important;line-height:1.1!important;}
  .user-chip{margin-top:8px!important;gap:8px!important;}
  .avatar{width:32px!important;height:32px!important;font-size:16px!important;}
  .admin-actions{display:grid!important;grid-template-columns:1fr!important;gap:10px!important;}
  .filter-grid,.form-grid,.form-grid.two,.ind-grid,.kpi-grid,.mini-kpis{display:grid!important;grid-template-columns:1fr!important;gap:10px!important;}
  .card{padding:13px!important;border-radius:14px!important;margin-bottom:12px!important;}
  .kpi-card{min-height:auto!important;align-items:center!important;}
  .icon-circle{width:46px!important;height:46px!important;font-size:22px!important;}
  .kpi-card .num{font-size:22px!important;}
  button,.btn{width:100%!important;min-height:44px!important;padding:10px 12px!important;text-align:center!important;display:grid!important;place-items:center!important;}
  input,select,textarea{min-height:44px!important;font-size:14px!important;}
  .table-head{display:block!important;}
  .table-head h3{margin-bottom:10px!important;}
  .table-head > div{display:grid!important;grid-template-columns:1fr!important;gap:8px!important;}
  .table-wrap{max-height:430px!important;min-height:180px!important;overflow:auto!important;border-radius:12px!important;}
  .table-wrap table{min-width:880px!important;}
  th,td{padding:9px 10px!important;font-size:12px!important;}
  .flash{font-size:13px!important;padding:11px 12px!important;}
}
@media(max-width: 390px){.nav-pro{grid-template-columns:1fr!important;}.hero h1{font-size:24px!important;}}

/* ===== RESPONSIVE FINAL PRO PARA CELULAR ===== */
@media (max-width: 700px){
  body{font-size:14px!important;}
  .content{padding:10px!important;}
  .card{border-radius:14px!important;padding:12px!important;margin-bottom:12px!important;}
  .form-grid,.form-grid.two,.filter-grid{display:grid!important;grid-template-columns:1fr!important;gap:10px!important;}
  input,select,textarea,button,.btn{width:100%!important;min-height:46px!important;font-size:15px!important;}
  .table-head{display:flex!important;flex-direction:column!important;align-items:flex-start!important;gap:10px!important;}
  .table-wrap{max-height:55vh!important;overflow:auto!important;-webkit-overflow-scrolling:touch!important;border-radius:12px!important;}
  .table-wrap table{min-width:780px!important;font-size:13px!important;}
  th,td{padding:10px 9px!important;}
  .topbar,.hero{position:relative!important;}
}

/* Usuarios PRO */
.users-card{padding-bottom:18px}
.user-search{max-width:420px;min-width:240px;margin-left:auto}
.users-scroll{max-height:70vh;overflow:auto;border:1px solid var(--line);border-radius:14px;background:#fff;display:block}
.users-scroll table{min-width:980px;width:100%}
.users-count{font-weight:900;color:#0f172a;background:#e8f3ff;border-radius:999px;padding:10px 14px}
.worker-name-field{grid-column:span 2;font-weight:900!important;background:#eef9f1!important;font-size:15px!important;min-width:360px}
.users-scroll th{position:sticky;top:0;z-index:2;background:#f7f9fc}
.pass-cell{display:flex;align-items:center;gap:8px;max-width:280px}
.pass-view{height:38px;padding:8px 10px;border-radius:10px;background:#f8fafc;font-weight:800;min-width:160px}
.eye-btn{padding:8px 10px;border-radius:10px;background:#0d73b8;box-shadow:none}
@media(max-width:760px){.user-search{width:100%;max-width:none;margin-left:0}.users-scroll{max-height:65vh}.pass-cell{min-width:210px}.users-scroll table{min-width:850px}.worker-name-field{grid-column:1/-1!important;min-width:100%}}


/* ===== NIVEL DIOS COMEDOR: REGISTRO MASIVO VISUAL / LECTOR CONTINUO ===== */
.lote-dios-panel{display:none;grid-column:1/-1;border:2px solid #16a34a;border-radius:18px;padding:16px;background:linear-gradient(135deg,#f0fdf4,#ffffff);box-shadow:0 14px 32px rgba(22,163,74,.16)}
.lote-dios-head{display:flex;justify-content:space-between;gap:12px;align-items:center;flex-wrap:wrap;margin-bottom:12px}
.lote-dios-title{font-size:20px;font-weight:950;color:#064e3b;line-height:1.15}
.lote-dios-sub{font-size:12px;color:#64748b;font-weight:750;margin-top:4px}
.lote-dios-counter{min-width:118px;text-align:center;border-radius:16px;padding:10px 14px;background:#16a34a;color:#fff;font-weight:950;box-shadow:0 10px 22px rgba(22,163,74,.22)}
.lote-dios-counter b{display:block;font-size:30px;line-height:1}
.lote-dios-counter span{font-size:11px;letter-spacing:.5px}
.lote-dios-status{display:grid;grid-template-columns:repeat(3,minmax(160px,1fr));gap:10px;margin:10px 0}
.lote-dios-status div{border:1px solid #bbf7d0;background:#ecfdf5;border-radius:14px;padding:10px;font-weight:900;color:#14532d}
.lote-dios-list-head,.lote-dios-row{display:grid;grid-template-columns:70px 135px minmax(240px,1fr) 120px 72px;gap:8px;align-items:center}
.lote-dios-list-head{padding:10px;background:#dcfce7;border:1px solid #86efac;border-radius:14px 14px 0 0;font-size:12px;font-weight:950;color:#14532d}
.lote-dios-list{max-height:260px;overflow:auto;background:white;border:1px solid #bbf7d0;border-top:0;border-radius:0 0 14px 14px}
.lote-dios-row{padding:10px;border-bottom:1px solid #eef2f7;font-size:13px;color:#25364a}
.lote-dios-row:last-child{border-bottom:0}.lote-dios-row b{font-weight:950}.lote-dios-row .ok{color:#166534;font-weight:950}.lote-dios-empty{padding:14px;color:#64748b;font-weight:800}
.lote-dios-actions{display:flex;gap:10px;flex-wrap:wrap;margin-top:12px}.lote-dios-actions button{width:auto!important}.cam-on{background:#052e16!important;color:#dcfce7!important;border:1px solid #22c55e!important;border-radius:12px;padding:8px 10px;font-weight:950;display:inline-flex;align-items:center;gap:8px}
@media(max-width:760px){.lote-dios-status{grid-template-columns:1fr}.lote-dios-list-head{display:none}.lote-dios-row{grid-template-columns:1fr;gap:3px;border:1px solid #e2e8f0;border-radius:12px;margin:8px}.lote-dios-actions button{width:100%!important}.lote-dios-counter{width:100%}}


/* ===== PREVISUALIZACIÓN EN TABLA PRINCIPAL - REGISTRO MASIVO ===== */
.fila-lote-preview{background:#ecfdf5!important;box-shadow:inset 4px 0 0 #16a34a}
.fila-lote-preview td{font-weight:850;color:#064e3b}
.fila-lote-preview.unchecked{background:#fff7ed!important;box-shadow:inset 4px 0 0 #f97316;opacity:.78}
.lote-check{width:22px!important;height:22px!important;min-height:22px!important;accent-color:#16a34a;cursor:pointer}
.badge.lote{background:#dcfce7;color:#166534}
.badge.lote-off{background:#ffedd5;color:#9a3412}


/* ===== FIX FINAL: CHECKBOX NORMAL + BLOQUEO RESPONSABLE + INDICADOR LOTE VISIBLE ===== */
input[type="checkbox"]{width:auto!important;min-height:0!important;height:18px!important;padding:0!important;margin:0 8px 0 0!important;transform:scale(1.15);vertical-align:middle;}
.label-lote-final{display:flex!important;align-items:center!important;gap:8px!important;min-height:46px!important;border:2px solid #16a34a!important;border-radius:12px!important;padding:10px 12px!important;background:#f0fdf4!important;color:#064e3b!important;font-weight:950!important;}
.responsable-alerta-final{border:2px solid #ef4444!important;background:#fff1f2!important;}
.lote-dios-panel{margin-top:8px!important;}

/* ===== AUTO-GUARDADO MASIVO REAL ===== */
.consumo-recien-guardado{background:#ecfdf5!important;box-shadow:inset 5px 0 0 #16a34a}
#auto_guardado_panel{grid-column:1/-1;border:2px solid #16a34a;border-radius:16px;background:#f0fdf4;padding:12px 14px;margin:8px 0;font-weight:950;color:#064e3b}
#auto_guardado_panel .mini{font-size:12px;color:#166534;margin-top:3px}


/* ===== CELULAR TIPO APP ASISTENCIA ===== */
@media(max-width:700px){
  html,body{height:auto!important;overflow:auto!important;background:#f2f4f6!important;}
  .app-shell{display:block!important;height:auto!important;min-height:100vh!important;overflow:visible!important;background:#f2f4f6!important;}
  .hero{margin-left:0!important;width:100%!important;min-height:58px!important;padding:10px 12px!important;background:#0aa866!important;position:sticky!important;top:0!important;z-index:80!important;box-shadow:0 3px 12px rgba(0,0,0,.18)!important;}
  .hero h1{font-size:20px!important;line-height:1.1!important;margin:0!important;color:#fff!important;text-align:left!important;}
  .hero h1:before{content:'← ';font-weight:900;margin-right:6px;}
  .hero p{display:none!important;}
  .main-layout{display:block!important;margin-left:0!important;width:100%!important;height:auto!important;overflow:visible!important;}
  .fixed-prize-sidebar{position:sticky!important;top:58px!important;width:100%!important;height:auto!important;padding:6px!important;background:#063047!important;z-index:70!important;box-shadow:0 3px 10px rgba(0,0,0,.12)!important;}
  .side-logo-pro,.side-user-card,.side-slogan-card{display:none!important;}
  .nav-pro{display:flex!important;gap:6px!important;overflow-x:auto!important;padding:0!important;scrollbar-width:none!important;}
  .nav-pro a{flex:0 0 auto!important;min-height:36px!important;margin:0!important;padding:8px 10px!important;border-radius:999px!important;font-size:11px!important;background:rgba(255,255,255,.08)!important;}
  .nav-pro a.on{background:#0aa866!important;}
  .nav-ico{font-size:12px!important;width:auto!important;flex:0 0 auto!important;}
  .nav-pro .pill{display:none!important;}
  .content{height:auto!important;overflow:visible!important;padding:10px!important;background:#f2f4f6!important;}
  .topbar{display:block!important;margin-bottom:8px!important;min-height:0!important;}
  .topbar h2{font-size:18px!important;margin-bottom:2px!important;color:#1f2937!important;}
  .user-chip{display:none!important;}
  .card{border-radius:14px!important;padding:12px!important;margin-bottom:10px!important;box-shadow:0 2px 10px rgba(0,0,0,.08)!important;border:0!important;}
  .form-grid,.form-grid.two,.filter-grid{display:grid!important;grid-template-columns:1fr 1fr!important;gap:9px!important;}
  .form-grid > *, .filter-grid > *{min-width:0!important;}
  input,select,textarea{min-height:40px!important;border-radius:6px!important;border:0!important;border-bottom:2px solid #c7c7c7!important;background:#f7f7f7!important;font-size:13px!important;padding:8px!important;}
  button,.btn{min-height:42px!important;border-radius:8px!important;font-size:13px!important;padding:9px 10px!important;}
  #dni_consumo,#nombre_trabajador,.worker-name-field{grid-column:1/-1!important;min-width:0!important;width:100%!important;}
  #info_trabajador_consumo{grid-column:1/-1!important;border-radius:12px!important;padding:10px!important;}
  .label-lote-final{grid-column:1/-1!important;min-height:42px!important;}
  .table-wrap{border:0!important;max-height:none!important;overflow:visible!important;}
  #tabla_consumos_principal{min-width:0!important;width:100%!important;border-collapse:separate!important;border-spacing:0 8px!important;}
  #tabla_consumos_principal thead{display:none!important;}
  #tbody_consumos_principal tr{display:block!important;background:#fff!important;border-radius:12px!important;margin:8px 0!important;padding:9px!important;box-shadow:0 2px 9px rgba(0,0,0,.08)!important;}
  #tbody_consumos_principal td{display:grid!important;grid-template-columns:92px 1fr!important;gap:6px!important;border:0!important;padding:4px 2px!important;white-space:normal!important;font-size:12px!important;}
  #tbody_consumos_principal td:nth-child(1)::before{content:'Sel.';font-weight:900;color:#777;}
  #tbody_consumos_principal td:nth-child(2)::before{content:'Fecha';font-weight:900;color:#777;}
  #tbody_consumos_principal td:nth-child(3)::before{content:'Hora';font-weight:900;color:#777;}
  #tbody_consumos_principal td:nth-child(4)::before{content:'DNI';font-weight:900;color:#777;}
  #tbody_consumos_principal td:nth-child(5)::before{content:'Nombre';font-weight:900;color:#777;}
  #tbody_consumos_principal td:nth-child(6)::before{content:'Área';font-weight:900;color:#777;}
  #tbody_consumos_principal td:nth-child(7)::before{content:'Tipo';font-weight:900;color:#777;}
  #tbody_consumos_principal td:nth-child(8)::before{content:'Comedor';font-weight:900;color:#777;}
  #tbody_consumos_principal td:nth-child(9)::before{content:'Fundo';font-weight:900;color:#777;}
  #tbody_consumos_principal td:nth-child(10)::before{content:'Resp.';font-weight:900;color:#777;}
  #tbody_consumos_principal td:nth-child(11)::before{content:'Cant.';font-weight:900;color:#777;}
  #tbody_consumos_principal td:nth-child(12)::before{content:'P. Unit.';font-weight:900;color:#777;}
  #tbody_consumos_principal td:nth-child(13)::before{content:'Total';font-weight:900;color:#777;}
  #tbody_consumos_principal td:nth-child(14)::before{content:'Estado';font-weight:900;color:#777;}
  #tbody_consumos_principal td:nth-child(15)::before{content:'Quitar';font-weight:900;color:#777;}
  #fila_sin_registros td{display:block!important;}
  #fila_sin_registros td::before{content:''!important;}
}



/* ===== AJUSTE FINAL PRO CELULAR: LOGIN CENTRADO + CARDS COMPACTAS + ALERTA VISIBLE ===== */
#prize_mobile_alert{position:fixed;left:12px;right:12px;top:calc(env(safe-area-inset-top,0px) + 12px);z-index:2147483647;padding:12px 14px;border-radius:14px;font-weight:950;color:#fff;text-align:center;box-shadow:0 14px 34px rgba(0,0,0,.35);pointer-events:none}
@media(max-width:700px){
  body:not(.logged-in) .login-page,
  .login-page{min-height:100dvh!important;height:100dvh!important;padding:14px!important;display:flex!important;align-items:center!important;justify-content:center!important;background:linear-gradient(180deg,#202020,#171717)!important;overflow:hidden!important;}
  .login-card{width:calc(100vw - 28px)!important;max-width:360px!important;border-radius:18px!important;background:#202020!important;border:1px solid rgba(255,255,255,.14)!important;box-shadow:0 18px 48px rgba(0,0,0,.42)!important;overflow:hidden!important;}
  .login-card:before{left:-45px!important;bottom:-58px!important;width:220px!important;height:110px!important;background:#002b84!important;border-radius:70% 70% 0 0!important;transform:rotate(6deg)!important;}
  .login-card:after{right:-55px!important;bottom:-62px!important;width:230px!important;height:120px!important;background:#014012!important;border-radius:65% 65% 0 0!important;transform:rotate(-8deg)!important;}
  .login-inner{padding:28px 24px 52px!important;text-align:center!important;}
  .login-inner .logo-word{font-size:56px!important;letter-spacing:-4px!important;line-height:.9!important;color:#bfefff!important;margin:0 auto 14px!important;display:inline-flex!important;align-items:flex-end!important;justify-content:center!important;transform:none!important;}
  .login-inner .logo-word .e{width:42px!important;height:42px!important;font-size:30px!important;border-width:4px!important;letter-spacing:-2px!important;position:relative!important;}
  .login-inner .logo-word .leaf{width:13px!important;height:24px!important;top:-27px!important;right:2px!important;}
  .login-title{font-size:17px!important;color:#eaf2ff!important;margin:4px 0 4px!important;text-align:center!important;}
  .login-subtitle{font-size:13px!important;color:#aeb8c5!important;margin:0 0 22px!important;text-align:center!important;}
  .form-label{font-size:12px!important;color:#e5e7eb!important;margin:12px 0 6px!important;text-align:left!important;}
  .input-icon span{left:12px!important;z-index:2!important;color:#fbbf24!important;font-size:17px!important;top:50%!important;transform:translateY(-50%)!important;}
  .input-icon input{height:46px!important;min-height:46px!important;background:#2d2d2d!important;border:1px solid #383838!important;border-bottom:2px solid #555!important;border-radius:7px!important;color:#fff!important;padding:10px 12px 10px 40px!important;font-size:13px!important;}
  .login-button{height:44px!important;min-height:44px!important;border-radius:8px!important;background:#003d08!important;color:white!important;margin-top:20px!important;box-shadow:0 12px 24px rgba(0,60,8,.3)!important;}

  .table-head h3{font-size:18px!important;color:#e7edf7!important;background:#1f2937!important;margin:-10px -10px 8px!important;padding:12px 10px!important;border-radius:0!important;}
  .table-head .btn,.table-head a{border-radius:8px!important;background:#052d74!important;margin:0 0 12px!important;height:42px!important;}
  #tabla_consumos_principal{border-spacing:0 7px!important;}
  #tbody_consumos_principal tr{position:relative!important;background:#ffffff!important;border-radius:13px!important;margin:8px 0!important;padding:10px 10px 10px 12px!important;box-shadow:0 3px 12px rgba(15,23,42,.10)!important;border-left:4px solid #0aa866!important;}
  #tbody_consumos_principal td{grid-template-columns:72px minmax(0,1fr)!important;gap:8px!important;padding:3px 0!important;font-size:12px!important;line-height:1.22!important;color:#334155!important;}
  #tbody_consumos_principal td::before{font-size:11px!important;color:#6b7280!important;letter-spacing:.1px!important;}
  #tbody_consumos_principal td:nth-child(1){display:none!important;}
  #tbody_consumos_principal td:nth-child(2),
  #tbody_consumos_principal td:nth-child(3),
  #tbody_consumos_principal td:nth-child(4){display:inline-grid!important;grid-template-columns:auto 1fr!important;width:33%!important;vertical-align:top!important;}
  #tbody_consumos_principal td:nth-child(5){font-size:13px!important;font-weight:900!important;color:#111827!important;padding-top:6px!important;}
  #tbody_consumos_principal td:nth-child(14) .badge,
  #tbody_consumos_principal td:nth-child(14) span{display:inline-flex!important;justify-content:center!important;width:auto!important;min-width:110px!important;border-radius:999px!important;padding:5px 10px!important;background:#f59e0b!important;color:#fff!important;font-size:11px!important;}
  #tbody_consumos_principal td:nth-child(15){grid-template-columns:72px 1fr!important;align-items:center!important;padding-top:7px!important;}
  #tbody_consumos_principal td:nth-child(15) form{display:grid!important;grid-template-columns:1fr 98px!important;gap:6px!important;width:100%!important;}
  #tbody_consumos_principal td:nth-child(15) input{height:38px!important;min-height:38px!important;font-size:12px!important;border-radius:8px!important;background:#f3f4f6!important;border:1px solid #d1d5db!important;padding:8px!important;}
  #tbody_consumos_principal td:nth-child(15) button{height:38px!important;min-height:38px!important;border-radius:8px!important;background:#991b1b!important;font-size:12px!important;padding:6px!important;}
  .flash{position:sticky!important;top:8px!important;z-index:120!important;margin:6px 0 10px!important;border-radius:12px!important;font-size:13px!important;box-shadow:0 12px 28px rgba(0,0,0,.22)!important;}
}


/* ===== FIX FINAL CÁMARA + ALERTAS MÓVILES ===== */
.prize-toast-msg{animation:prizeToastIn .18s ease-out both;}
@keyframes prizeToastIn{from{transform:translateY(-12px);opacity:0}to{transform:translateY(0);opacity:1}}
#qr-reader{width:100%!important;max-width:390px!important;margin:10px auto!important;grid-column:1/-1!important;}
.qr-camera-box{grid-column:1/-1!important;width:100%!important;max-width:390px!important;margin:0 auto!important;padding:10px!important;border:1px solid rgba(255,255,255,.12)!important;border-radius:16px!important;background:#101820!important;color:#eef6ff!important;text-align:center!important;box-shadow:0 12px 28px rgba(0,0,0,.25)!important;}
.qr-live-box,.qr-video-box{width:100%!important;max-width:340px!important;height:min(62vh,430px)!important;aspect-ratio:3/4!important;margin:10px auto 8px!important;border-radius:14px!important;background:#000!important;overflow:hidden!important;display:block;}
.qr-live-box video,.qr-live-box canvas,#qr-reader-live video,#qr-reader-live canvas{width:100%!important;height:100%!important;object-fit:cover!important;border-radius:14px!important;}
.qr-video-box{object-fit:cover!important;}
.qr-actions{display:block!important;margin-top:8px!important;}
.qr-close-btn{width:100%!important;max-width:340px!important;margin:8px auto 0!important;display:block!important;background:#7f0000!important;border-radius:10px!important;}
#qr-reader small{display:block!important;margin-top:10px!important;color:#dbeafe!important;text-align:left!important;font-size:12px!important;line-height:1.35!important;}
@media(max-width:700px){
  #qr-reader{max-width:100%!important;margin:8px auto!important;padding:0 0 8px!important;}
  .qr-camera-box{max-width:100%!important;border-radius:12px!important;padding:8px!important;}
  .qr-live-box,.qr-video-box{max-width:330px!important;height:min(58vh,410px)!important;}
  #prize_mobile_alert{top:calc(env(safe-area-inset-top,0px) + 12px)!important;z-index:2147483647!important;}
}


/* ===== CONTADOR GRANDE DE LECTURAS EN CELULAR ===== */
@media(max-width:700px){
  #contador_lecturas_box{grid-template-columns:34px 1fr 78px!important;padding:12px!important;margin:8px 0 12px!important;border-radius:14px!important;position:relative!important;z-index:20!important;}
  #contador_lecturas_box div{line-height:1.15!important;}
  #contador_lecturas_hoy{font-size:28px!important;}
}


/* ===== LOGO PRIZE REAL EN TODO EL SISTEMA ===== */
.prize-logo-img{max-width:190px;max-height:86px;object-fit:contain;display:block}
.hero-brand .prize-logo-img{max-width:210px;max-height:92px;margin:auto}
.side-logo-pro .prize-logo-img{max-width:150px;max-height:72px;margin:0 auto 4px;background:white;border-radius:14px;padding:6px}
.login-inner .prize-logo-img{max-width:210px;max-height:100px;margin:0 auto 14px;background:white;border-radius:18px;padding:8px}
.entrega-pro-panel{border:2px solid #0d73b8;background:linear-gradient(135deg,#eff6ff,#ffffff);border-radius:18px;padding:14px;margin-top:12px}
.entrega-pro-status{display:grid;grid-template-columns:repeat(auto-fit,minmax(170px,1fr));gap:10px;margin:12px 0}
.entrega-pro-status>div{border:1px solid #bfdbfe;background:#dbeafe;border-radius:14px;padding:10px;font-weight:900;color:#0f3c68}
.entrega-ok-row{background:#ecfdf5!important;box-shadow:inset 5px 0 0 #16a34a}
@media(max-width:700px){.prize-logo-img{max-width:160px}.hero-brand .prize-logo-img{max-width:155px}.side-logo-pro .prize-logo-img{max-width:130px}.entrega-pro-panel button{width:100%!important}.entrega-pro-panel input,.entrega-pro-panel select{font-size:16px!important}}



/* ===== FIX DEFINITIVO 04/05/2026: LOGO REAL + PESTAÑAS APP CELULAR ===== */
.prize-logo-img{display:block!important;object-fit:contain!important;width:auto!important;height:auto!important;}
.hero-brand .prize-logo-img{max-width:190px!important;max-height:78px!important;margin:0 auto!important;background:#fff!important;border-radius:16px!important;padding:7px!important;box-shadow:0 6px 18px rgba(0,0,0,.16)!important;}
.side-logo-pro .prize-logo-img{max-width:158px!important;max-height:76px!important;margin:0 auto 6px!important;background:#fff!important;border-radius:16px!important;padding:7px!important;}
.login-inner .prize-logo-img{max-width:240px!important;max-height:118px!important;margin:0 auto 16px!important;background:#fff!important;border-radius:18px!important;padding:10px!important;box-shadow:0 8px 22px rgba(0,0,0,.24)!important;}

@media(max-width:700px){
  html,body{width:100%!important;max-width:100%!important;overflow-x:hidden!important;background:#f3f4f6!important;}
  *{box-sizing:border-box!important;}
  .app-shell,.main-layout,.content{width:100%!important;max-width:100%!important;margin:0!important;left:0!important;right:0!important;overflow-x:hidden!important;}
  .hero{display:grid!important;grid-template-columns:72px 1fr!important;align-items:center!important;gap:10px!important;min-height:64px!important;padding:8px 10px!important;background:#062338!important;position:sticky!important;top:0!important;z-index:1000!important;}
  .hero h1:before{content:''!important;}
  .hero-brand{display:block!important;min-width:0!important;}
  .hero-brand .prize-logo-img{max-width:70px!important;max-height:46px!important;border-radius:10px!important;padding:3px!important;background:#fff!important;}
  .hero h1{font-size:17px!important;line-height:1.08!important;color:#fff!important;margin:0!important;text-align:left!important;}
  .hero p{display:block!important;font-size:10px!important;line-height:1.05!important;color:#cde7f9!important;margin:2px 0 0!important;text-align:left!important;}

  .fixed-prize-sidebar{position:sticky!important;top:64px!important;width:100%!important;max-width:100%!important;height:auto!important;margin:0!important;padding:6px!important;background:#061b2b!important;z-index:999!important;border-radius:0!important;}
  .side-logo-pro,.side-user-card,.side-slogan-card,.side-title{display:none!important;}
  .nav-pro{display:grid!important;grid-template-columns:repeat(3,minmax(0,1fr))!important;gap:6px!important;width:100%!important;overflow:visible!important;padding:0!important;margin:0!important;}
  .nav-pro a{width:100%!important;min-width:0!important;min-height:42px!important;margin:0!important;padding:7px 4px!important;border-radius:10px!important;display:flex!important;flex-direction:column!important;align-items:center!important;justify-content:center!important;gap:2px!important;text-align:center!important;font-size:10px!important;line-height:1.05!important;font-weight:900!important;background:rgba(255,255,255,.08)!important;color:#fff!important;white-space:normal!important;}
  .nav-pro a.on{background:linear-gradient(135deg,#0aa866,#0d73b8)!important;box-shadow:0 5px 14px rgba(0,0,0,.2)!important;}
  .nav-ico{font-size:15px!important;width:auto!important;line-height:1!important;}
  .nav-pro .pill{display:none!important;}
  .content{padding:10px!important;background:#f3f4f6!important;}
  .card{width:100%!important;max-width:100%!important;overflow:hidden!important;background:#fff!important;color:#172033!important;}
  .topbar h2{font-size:19px!important;line-height:1.12!important;color:#172033!important;}

  .entrega-pro-panel{width:100%!important;max-width:100%!important;overflow:hidden!important;padding:12px!important;border-radius:16px!important;background:linear-gradient(135deg,#eff6ff,#fff)!important;}
  .entrega-pro-status{grid-template-columns:1fr!important;gap:8px!important;}
  #qr_entrega_reader{max-width:100%!important;}

  /* Entregas en celular: tabla tipo tarjetas, no pantalla cortada */
  #pedidos_body{display:block!important;width:100%!important;}
  #pedidos_body tr{display:block!important;width:100%!important;background:#fff!important;border-radius:12px!important;margin:8px 0!important;padding:9px!important;box-shadow:0 2px 9px rgba(0,0,0,.10)!important;border:1px solid #e5e7eb!important;}
  #pedidos_body td{display:grid!important;grid-template-columns:92px 1fr!important;gap:6px!important;border:0!important;padding:4px 2px!important;white-space:normal!important;font-size:12px!important;color:#172033!important;}
  #pedidos_body td:nth-child(1)::before{content:'Sel.';font-weight:900;color:#64748b;}
  #pedidos_body td:nth-child(2)::before{content:'#';font-weight:900;color:#64748b;}
  #pedidos_body td:nth-child(3)::before{content:'Hora';font-weight:900;color:#64748b;}
  #pedidos_body td:nth-child(4)::before{content:'DNI';font-weight:900;color:#64748b;}
  #pedidos_body td:nth-child(5)::before{content:'Trabajador';font-weight:900;color:#64748b;}
  #pedidos_body td:nth-child(6)::before{content:'Tipo';font-weight:900;color:#64748b;}
  #pedidos_body td:nth-child(7)::before{content:'Cantidad';font-weight:900;color:#64748b;}
  #pedidos_body td:nth-child(8)::before{content:'Obs.';font-weight:900;color:#64748b;}
  #pedidos_body td:nth-child(9)::before{content:'Estado';font-weight:900;color:#64748b;}
  .table-wrap{width:100%!important;max-width:100%!important;overflow:visible!important;max-height:none!important;background:transparent!important;border:0!important;}
  .table-wrap table:has(#pedidos_body){display:block!important;width:100%!important;min-width:0!important;}
  .table-wrap table:has(#pedidos_body) thead{display:none!important;}

  .login-page{min-height:100dvh!important;height:100dvh!important;padding:14px!important;background:linear-gradient(180deg,#050505,#171717)!important;}
  .login-card{max-width:360px!important;width:calc(100vw - 28px)!important;}
  .login-inner .prize-logo-img{max-width:225px!important;max-height:110px!important;border-radius:18px!important;padding:8px!important;}
}
@media(max-width:390px){.nav-pro{grid-template-columns:repeat(3,minmax(0,1fr))!important}.nav-pro a{font-size:9.5px!important;padding-left:2px!important;padding-right:2px!important;}}
@media(max-width:340px){.nav-pro{grid-template-columns:repeat(2,minmax(0,1fr))!important}.hero{grid-template-columns:60px 1fr!important}.hero-brand .prize-logo-img{max-width:58px!important;}}



/* =========================================================
   NIVEL PRO TOTAL DEFINITIVO 04/05/2026
   - Logo PRIZE real visible en login, cabecera y menú
   - Indicadores tipo web con iconos, colores y presentación
   - Pestañas horizontales dinámicas para app celular
   - Pestañas y tablas más presentables para admin y usuario
   ========================================================= */
:root{
  --prize-blue:#1f5f8a;
  --prize-blue-2:#0d73b8;
  --prize-green:#2aa84a;
  --prize-orange:#f58220;
  --prize-yellow:#f5b52e;
  --prize-dark:#071827;
}
.prize-logo-img{
  display:block!important;
  object-fit:contain!important;
  object-position:center!important;
  width:auto!important;
  height:auto!important;
  max-width:100%!important;
  image-rendering:auto!important;
}
.login-inner .prize-logo-img{
  width:240px!important;
  max-width:88%!important;
  max-height:126px!important;
  margin:0 auto 16px!important;
  background:#fff!important;
  border-radius:20px!important;
  padding:10px 14px!important;
  box-shadow:0 10px 28px rgba(0,0,0,.32)!important;
}
.hero-brand .prize-logo-img{
  width:190px!important;
  max-height:88px!important;
  background:#fff!important;
  border-radius:18px!important;
  padding:8px 12px!important;
  box-shadow:0 10px 26px rgba(0,0,0,.18)!important;
}
.side-logo-pro .prize-logo-img{
  width:168px!important;
  max-height:84px!important;
  margin:0 auto 10px!important;
  background:#fff!important;
  border-radius:18px!important;
  padding:8px 12px!important;
  box-shadow:0 10px 24px rgba(0,0,0,.18)!important;
}
.hero{
  background:linear-gradient(135deg,#061b2b 0%,#0a3854 48%,#0d6135 100%)!important;
  border-bottom:4px solid var(--prize-orange)!important;
}
.nav-pro{
  scrollbar-width:thin!important;
}
.nav-pro a{
  border:1px solid rgba(255,255,255,.08)!important;
}
.nav-pro a.on{
  background:linear-gradient(135deg,var(--prize-green),var(--prize-blue-2))!important;
  border-color:rgba(255,255,255,.25)!important;
}
.kpi-grid{
  display:grid!important;
  grid-template-columns:repeat(auto-fit,minmax(210px,1fr))!important;
  gap:16px!important;
}
.kpi-card{
  position:relative!important;
  overflow:hidden!important;
  min-height:116px!important;
  border:0!important;
  color:#fff!important;
  padding:18px!important;
  display:flex!important;
  align-items:center!important;
  gap:15px!important;
  border-radius:20px!important;
  box-shadow:0 16px 36px rgba(6,24,43,.16)!important;
}
.kpi-card:nth-child(1){background:linear-gradient(135deg,#0f9f52,#026b36)!important;}
.kpi-card:nth-child(2){background:linear-gradient(135deg,#1687d9,#075985)!important;}
.kpi-card:nth-child(3){background:linear-gradient(135deg,#7c3aed,#4c1d95)!important;}
.kpi-card:nth-child(4){background:linear-gradient(135deg,#f59e0b,#ea580c)!important;}
.kpi-card:after{
  content:""!important;
  position:absolute!important;
  right:-36px!important;
  bottom:-44px!important;
  width:130px!important;
  height:130px!important;
  border-radius:50%!important;
  background:rgba(255,255,255,.14)!important;
}
.kpi-card .icon-circle{
  width:54px!important;
  height:54px!important;
  border-radius:18px!important;
  display:grid!important;
  place-items:center!important;
  flex:0 0 54px!important;
  background:rgba(255,255,255,.22)!important;
  color:#fff!important;
  font-size:24px!important;
  font-weight:950!important;
  box-shadow:inset 0 0 0 1px rgba(255,255,255,.25)!important;
}
.kpi-card .label,.kpi-card .sub{color:rgba(255,255,255,.88)!important;}
.kpi-card .num{
  color:#fff!important;
  font-size:31px!important;
  font-weight:1000!important;
  letter-spacing:-1px!important;
}
.kpi-card .num[style]{color:#fff!important;}
.card{
  border:1px solid #e6edf5!important;
  box-shadow:0 12px 30px rgba(15,35,55,.08)!important;
}
.topbar h2{
  color:#082033!important;
  font-weight:1000!important;
  letter-spacing:-.3px!important;
}
.table-head{
  background:linear-gradient(135deg,#f8fafc,#eef6ff)!important;
  border:1px solid #e2e8f0!important;
  border-radius:16px!important;
  padding:12px!important;
  margin-bottom:12px!important;
}
.table-wrap{
  border-radius:18px!important;
  border:1px solid #e6edf5!important;
  box-shadow:0 10px 26px rgba(15,35,55,.06)!important;
}
table th{
  background:linear-gradient(135deg,#061b2b,#0b3350)!important;
  color:#eaf7ff!important;
  font-weight:950!important;
  white-space:nowrap!important;
}
table td{vertical-align:middle!important;}
.entrega-pro-panel{
  background:linear-gradient(135deg,#eef7ff,#ffffff 55%,#f0fff4)!important;
  border:1px solid rgba(13,115,184,.25)!important;
  box-shadow:0 14px 34px rgba(13,115,184,.12)!important;
}
.entrega-pro-status>div{
  background:linear-gradient(135deg,#ffffff,#e9f5ff)!important;
  border:1px solid #cfe8ff!important;
  box-shadow:0 8px 18px rgba(13,115,184,.08)!important;
}

@media(max-width:700px){
  html,body{overflow-x:hidden!important;background:#f4f6f8!important;}
  .login-page{
    min-height:100dvh!important;
    height:100dvh!important;
    padding:14px!important;
    display:flex!important;
    align-items:center!important;
    justify-content:center!important;
    background:radial-gradient(circle at 12% 91%,#002c9b 0 20%,transparent 21%),radial-gradient(circle at 90% 96%,#00571c 0 24%,transparent 25%),linear-gradient(180deg,#050505,#181818)!important;
  }
  .login-card{max-width:374px!important;width:calc(100vw - 24px)!important;border-radius:20px!important;}
  .login-inner{padding:24px 20px 50px!important;}
  .login-inner .prize-logo-img{
    width:235px!important;
    max-width:92%!important;
    max-height:122px!important;
    margin-bottom:14px!important;
  }
  .hero{
    grid-template-columns:76px 1fr!important;
    gap:9px!important;
    padding:8px 10px!important;
    min-height:66px!important;
    position:sticky!important;
    top:0!important;
    z-index:1200!important;
  }
  .hero-brand .prize-logo-img{
    width:70px!important;
    max-height:48px!important;
    border-radius:12px!important;
    padding:3px!important;
    box-shadow:0 5px 13px rgba(0,0,0,.22)!important;
  }
  .hero h1{font-size:17px!important;line-height:1.08!important;margin:0!important;color:#fff!important;}
  .hero p{font-size:10.5px!important;line-height:1.08!important;color:#d7edf9!important;margin:2px 0 0!important;}
  .fixed-prize-sidebar{
    position:sticky!important;
    top:66px!important;
    z-index:1190!important;
    width:100%!important;
    height:auto!important;
    padding:7px 7px 8px!important;
    background:#071827!important;
    border-bottom:1px solid rgba(255,255,255,.10)!important;
  }
  .side-logo-pro,.side-user-card,.side-slogan-card,.side-title{display:none!important;}
  .nav-pro{
    display:flex!important;
    flex-wrap:nowrap!important;
    gap:7px!important;
    overflow-x:auto!important;
    overflow-y:hidden!important;
    -webkit-overflow-scrolling:touch!important;
    scroll-snap-type:x proximity!important;
    padding:0 3px 2px!important;
    margin:0!important;
    width:100%!important;
  }
  .nav-pro::-webkit-scrollbar{height:4px!important;}
  .nav-pro::-webkit-scrollbar-thumb{background:#2aa84a!important;border-radius:999px!important;}
  .nav-pro a{
    flex:0 0 auto!important;
    min-width:92px!important;
    max-width:118px!important;
    min-height:48px!important;
    display:flex!important;
    flex-direction:column!important;
    align-items:center!important;
    justify-content:center!important;
    gap:3px!important;
    padding:7px 9px!important;
    margin:0!important;
    border-radius:15px!important;
    background:rgba(255,255,255,.08)!important;
    color:#fff!important;
    text-align:center!important;
    font-size:10px!important;
    line-height:1.05!important;
    font-weight:950!important;
    white-space:normal!important;
    scroll-snap-align:start!important;
  }
  .nav-pro a.on{
    background:linear-gradient(135deg,#2aa84a,#0d73b8)!important;
    box-shadow:0 8px 18px rgba(0,0,0,.28)!important;
    transform:none!important;
  }
  .nav-ico{font-size:16px!important;line-height:1!important;width:auto!important;}
  .nav-pro .pill{display:none!important;}
  .main-layout{display:block!important;width:100%!important;max-width:100%!important;}
  .content{padding:10px!important;width:100%!important;max-width:100%!important;overflow-x:hidden!important;background:#f4f6f8!important;}
  .panel-right{display:none!important;}
  .footer{display:none!important;}
  .topbar{display:block!important;margin-bottom:10px!important;}
  .topbar h2{font-size:19px!important;line-height:1.13!important;margin:0 0 4px!important;color:#111827!important;}
  .card{border-radius:18px!important;padding:13px!important;margin-bottom:12px!important;box-shadow:0 8px 20px rgba(15,35,55,.08)!important;overflow:hidden!important;}
  .kpi-grid,.mini-kpis,.ind-grid{
    grid-template-columns:repeat(2,minmax(0,1fr))!important;
    gap:10px!important;
  }
  .kpi-card{
    min-height:96px!important;
    border-radius:18px!important;
    padding:12px!important;
    gap:9px!important;
  }
  .kpi-card .icon-circle{width:42px!important;height:42px!important;flex-basis:42px!important;border-radius:14px!important;font-size:19px!important;}
  .kpi-card .label{font-size:11px!important;line-height:1.1!important;}
  .kpi-card .num{font-size:23px!important;line-height:1.05!important;}
  .kpi-card .sub{font-size:10px!important;line-height:1.1!important;}
  .form-grid,.form-grid.two,.filter-grid{grid-template-columns:1fr!important;gap:10px!important;}
  input,select,textarea{font-size:16px!important;min-height:44px!important;border-radius:12px!important;background:#fff!important;border:1px solid #dce6f0!important;}
  button,.btn{width:100%!important;min-height:44px!important;border-radius:12px!important;}
  .table-head{display:block!important;border-radius:16px!important;padding:11px!important;}
  .table-head h3{font-size:16px!important;margin:0 0 8px!important;}
  .table-wrap{overflow:auto!important;-webkit-overflow-scrolling:touch!important;max-width:100%!important;border-radius:16px!important;}
  .table-wrap table{min-width:760px!important;font-size:12px!important;}
  .entrega-pro-panel{padding:12px!important;border-radius:18px!important;}
  .entrega-pro-status{grid-template-columns:1fr!important;}
}
@media(max-width:380px){
  .kpi-grid,.mini-kpis,.ind-grid{grid-template-columns:1fr!important;}
  .nav-pro a{min-width:88px!important;font-size:9.5px!important;}
  .login-inner .prize-logo-img{width:215px!important;}
}


/* =========================================================
   FINAL PRO 2026: LOGO REAL + APP PRESENTABLE + TABS HORIZONTALES
   ========================================================= */
:root{
  --prize-blue:#0b4f7a;
  --prize-green:#2ca84a;
  --prize-orange:#e76f16;
  --app-bg:#111827;
  --panel:#1f2937;
  --panel2:#0b1220;
  --soft:#eaf3ff;
}
.prize-logo-img{
  display:block!important;
  width:100%!important;
  height:auto!important;
  object-fit:contain!important;
  image-rendering:auto!important;
  background:#fff!important;
  border-radius:16px!important;
  padding:8px!important;
  box-shadow:0 10px 28px rgba(0,0,0,.20)!important;
}
.login-inner .prize-logo-img{
  width:min(285px,78vw)!important;
  max-height:145px!important;
  margin:0 auto 18px!important;
  padding:10px 16px!important;
  border-radius:22px!important;
  background:#fff!important;
}
.hero{
  background:linear-gradient(135deg,#061b2b 0%,#092f4a 62%,#05351e 100%)!important;
  color:#fff!important;
  border-bottom:1px solid rgba(255,255,255,.10)!important;
}
.hero-brand{
  display:flex!important;
  border:0!important;
  min-height:auto!important;
}
.hero-brand .prize-logo-img{
  width:170px!important;
  max-height:86px!important;
  margin:auto!important;
  padding:6px!important;
  border-radius:14px!important;
}
.hero h1{color:#fff!important;text-shadow:0 3px 12px rgba(0,0,0,.28)!important;}
.hero p{color:#dbeafe!important;font-weight:850!important;}
.side-logo-pro{display:block!important;padding:8px 8px 14px!important;margin-bottom:10px!important;border-bottom:1px solid rgba(255,255,255,.12)!important;}
.side-logo-pro .prize-logo-img{
  width:145px!important;
  max-height:90px!important;
  margin:0 auto!important;
  padding:5px!important;
  border-radius:14px!important;
  box-shadow:0 8px 20px rgba(0,0,0,.25)!important;
}
.fixed-prize-sidebar{
  background:linear-gradient(180deg,#061b2b 0%,#082f49 48%,#03131f 100%)!important;
}
.nav-pro{
  display:flex!important;
  flex-direction:column!important;
  gap:7px!important;
}
.nav-pro a{
  border:1px solid rgba(255,255,255,.08)!important;
  background:rgba(255,255,255,.055)!important;
  color:#eef7ff!important;
  box-shadow:0 5px 12px rgba(0,0,0,.10)!important;
}
.nav-pro a.on,.nav-pro a:hover{
  background:linear-gradient(135deg,var(--prize-green),#0d73b8)!important;
  color:white!important;
}
.kpi-grid,.ind-grid,.mini-kpis{
  align-items:stretch!important;
}
.kpi-card,.ind-grid .card,.mini-kpis .card{
  border:1px solid rgba(13,115,184,.10)!important;
  background:linear-gradient(145deg,#ffffff,#f7fbff)!important;
  box-shadow:0 12px 28px rgba(15,35,55,.10)!important;
}
.icon-circle{
  box-shadow:inset 0 0 0 1px rgba(255,255,255,.4),0 10px 18px rgba(15,35,55,.10)!important;
}
.table-head h3,
.content h2,
.content h3{
  letter-spacing:-.2px!important;
}
.table-head{
  background:linear-gradient(135deg,#061b2b,#082f49)!important;
  color:#fff!important;
  border-radius:16px!important;
  padding:16px!important;
  box-shadow:0 10px 24px rgba(15,35,55,.12)!important;
}
.table-head h3{color:#fff!important;}
.card.filter-card,.filter-card{
  background:linear-gradient(145deg,#ffffff,#f8fbff)!important;
  border:1px solid #dbeafe!important;
}
.content input,.content select,.content textarea{
  border-radius:14px!important;
  border:1px solid #d7e3ef!important;
  background:#fff!important;
}
button,.btn{
  border-radius:14px!important;
  box-shadow:0 10px 20px rgba(0,0,0,.12)!important;
}
.table-wrap{
  border-radius:16px!important;
  border:1px solid #dbe7f3!important;
  box-shadow:0 10px 22px rgba(15,35,55,.07)!important;
}
th{
  background:#eef6ff!important;
  color:#0f2942!important;
}

@media(max-width:900px){
  body{background:#1f1f1f!important;}
  .app-shell{background:#1f1f1f!important;}
  .login-page{
    min-height:100vh!important;
    padding:18px 12px!important;
    background:radial-gradient(circle at 20% 90%,rgba(13,115,184,.30),transparent 28%),radial-gradient(circle at 85% 92%,rgba(22,163,74,.28),transparent 30%),#050505!important;
  }
  .login-card{width:min(94vw,430px)!important;background:#050505!important;border:1px solid #1f2937!important;border-radius:24px!important;}
  .login-title{color:#f8fafc!important;font-size:19px!important;}
  .login-subtitle,.form-label{color:#cbd5e1!important;}
  .input-icon input{background:#171717!important;border-color:#4b5563!important;color:#e5e7eb!important;}
  .hero{
    display:grid!important;
    grid-template-columns:78px 1fr!important;
    gap:10px!important;
    align-items:center!important;
    margin:0!important;
    width:100%!important;
    padding:10px 12px!important;
    min-height:82px!important;
    text-align:left!important;
  }
  .hero-brand{display:flex!important;justify-content:center!important;align-items:center!important;}
  .hero-brand .prize-logo-img{
    width:72px!important;
    max-height:58px!important;
    padding:4px!important;
    border-radius:10px!important;
  }
  .hero h1{font-size:19px!important;line-height:1.12!important;margin:0 0 3px!important;max-width:none!important;}
  .hero p{font-size:11px!important;line-height:1.15!important;max-width:none!important;}
  .fixed-prize-sidebar{
    position:sticky!important;
    top:0!important;
    z-index:100!important;
    width:100%!important;
    height:auto!important;
    padding:8px!important;
    background:#061b2b!important;
    overflow:visible!important;
  }
  .side-logo-pro,.side-user-card,.side-slogan-card{display:none!important;}
  .nav-pro{
    display:flex!important;
    flex-direction:row!important;
    gap:8px!important;
    overflow-x:auto!important;
    padding:0 2px 2px!important;
    scroll-snap-type:x mandatory!important;
    -webkit-overflow-scrolling:touch!important;
  }
  .nav-pro a{
    flex:0 0 auto!important;
    min-width:105px!important;
    min-height:50px!important;
    display:flex!important;
    flex-direction:column!important;
    align-items:center!important;
    justify-content:center!important;
    gap:3px!important;
    text-align:center!important;
    border-radius:14px!important;
    padding:8px 9px!important;
    font-size:11px!important;
    line-height:1.05!important;
    scroll-snap-align:start!important;
    white-space:normal!important;
  }
  .nav-ico{font-size:17px!important;width:auto!important;}
  .nav-pro .pill{display:none!important;}
  .content{
    background:#1f1f1f!important;
    color:#e5e7eb!important;
    padding:12px!important;
  }
  .topbar h2,.content h2,.content h3{color:#f8fafc!important;}
  .card,.kpi-card,.ind-grid .card,.mini-kpis .card,.filter-card{
    background:linear-gradient(145deg,#242424,#1f2937)!important;
    border:1px solid #374151!important;
    color:#e5e7eb!important;
    border-radius:18px!important;
  }
  .muted,.small,.kpi-card .label,.kpi-card .sub{color:#b6c2d1!important;}
  .kpi-card .num{color:#f8fafc!important;}
  input,select,textarea{background:#303030!important;color:#f8fafc!important;border-color:#4b5563!important;}
  .table-head{background:#061b2b!important;border-radius:16px!important;}
  .table-wrap{border-color:#374151!important;background:#1f2937!important;}
  table{background:#1f2937!important;}
  th{background:#0b1220!important;color:#dbeafe!important;}
  td{color:#e5e7eb!important;border-bottom-color:#374151!important;}
  .kpi-grid,.ind-grid,.mini-kpis{grid-template-columns:repeat(2,minmax(0,1fr))!important;gap:10px!important;}
  .kpi-card{min-height:92px!important;padding:12px!important;gap:12px!important;}
  .icon-circle{width:46px!important;height:46px!important;font-size:22px!important;}
}
@media(max-width:430px){
  .login-inner .prize-logo-img{width:245px!important;max-height:125px!important;}
  .hero{grid-template-columns:66px 1fr!important;}
  .hero-brand .prize-logo-img{width:62px!important;max-height:52px!important;}
  .hero h1{font-size:17px!important;}
  .hero p{font-size:10.5px!important;}
  .nav-pro a{min-width:96px!important;font-size:10px!important;}
  .kpi-grid,.ind-grid,.mini-kpis{grid-template-columns:1fr!important;}
}


/* ===== AJUSTE SOLICITADO 04/05/2026: LOGO EN LOGIN Y SIN LOGO ARRIBA DEL TÍTULO ===== */
.hero{grid-template-columns:1fr!important;text-align:center!important;justify-items:center!important;min-height:92px!important;padding:16px 18px!important;}
.hero .hero-brand{display:none!important;}
.hero-title-only{width:100%!important;text-align:center!important;}
.hero-title-only h1{margin:0 0 4px!important;}
.login-inner .prize-logo-img{display:block!important;width:220px!important;max-width:88%!important;max-height:112px!important;object-fit:contain!important;background:#fff!important;border-radius:20px!important;padding:10px 14px!important;margin:0 auto 18px!important;box-shadow:0 12px 30px rgba(0,0,0,.35)!important;}
@media(max-width:700px){
  .hero{display:block!important;min-height:56px!important;padding:10px 12px!important;}
  .hero-title-only{text-align:left!important;}
  .hero-title-only h1{font-size:19px!important;line-height:1.1!important;text-align:left!important;}
  .hero-title-only p{font-size:11px!important;text-align:left!important;}
  .login-inner .prize-logo-img{width:190px!important;max-height:96px!important;margin-bottom:14px!important;}
}



/* =========================================================
   AJUSTE SOLICITADO: INDICADORES DEL DASHBOARD CENTRADOS
   ========================================================= */
.dashboard-kpis-center{
  width:100%!important;
  max-width:1120px!important;
  margin:0 auto 22px!important;
  display:grid!important;
  grid-template-columns:repeat(4,minmax(190px,240px))!important;
  justify-content:center!important;
  align-items:stretch!important;
  gap:18px!important;
}
.dashboard-kpis-center .kpi-card{
  justify-content:center!important;
  text-align:center!important;
  flex-direction:column!important;
  align-items:center!important;
  min-height:148px!important;
  padding:20px 14px!important;
}
.dashboard-kpis-center .icon-circle{
  margin:0 auto 6px!important;
  width:60px!important;
  height:60px!important;
  font-size:28px!important;
  flex:none!important;
}
.dashboard-kpis-center .label,
.dashboard-kpis-center .sub,
.dashboard-kpis-center .num{
  text-align:center!important;
  width:100%!important;
}
.dashboard-kpis-center .num{
  font-size:30px!important;
  line-height:1.05!important;
}
@media(max-width:1100px){
  .dashboard-kpis-center{
    grid-template-columns:repeat(2,minmax(190px,260px))!important;
    max-width:580px!important;
  }
}
@media(max-width:620px){
  .dashboard-kpis-center{
    grid-template-columns:1fr!important;
    max-width:360px!important;
    gap:12px!important;
    margin-left:auto!important;
    margin-right:auto!important;
  }
  .dashboard-kpis-center .kpi-card{
    min-height:128px!important;
  }
}


/* =========================================================
   AJUSTE FINAL: FONDO LIMPIO Y PRO PARA LOGO PRIZE
   - Login: logo grande, centrado, fondo blanco limpio y sombra suave
   - Cabecera: se mantiene sin logo arriba del título para no ocupar espacio
   - Menú lateral: fondo del logo optimizado si se usa en escritorio
   ========================================================= */
.logo-clean-bg,
.login-inner .prize-logo-img,
.side-logo-pro .prize-logo-img{
  background:linear-gradient(180deg,#ffffff 0%,#f8fafc 100%)!important;
  border:1px solid rgba(226,232,240,.95)!important;
  border-radius:22px!important;
  padding:12px 18px!important;
  box-shadow:
    0 14px 34px rgba(0,0,0,.30),
    inset 0 1px 0 rgba(255,255,255,.95)!important;
  object-fit:contain!important;
  object-position:center!important;
}
.login-inner .prize-logo-img{
  display:block!important;
  width:245px!important;
  max-width:90%!important;
  max-height:128px!important;
  margin:0 auto 18px!important;
}
.login-card{
  border:1px solid rgba(148,163,184,.28)!important;
  box-shadow:0 24px 70px rgba(0,0,0,.55)!important;
}
.login-inner{
  padding-top:32px!important;
}
.hero .hero-brand{
  display:none!important;
}
.hero{
  grid-template-columns:1fr!important;
  text-align:center!important;
  justify-items:center!important;
  min-height:88px!important;
  padding:16px 18px!important;
}
.hero-title-only{
  width:100%!important;
}
.hero-title-only h1{
  margin-top:0!important;
}
.side-logo-pro .prize-logo-img{
  width:156px!important;
  max-height:78px!important;
  margin:0 auto 10px!important;
  padding:8px 12px!important;
  border-radius:18px!important;
  box-shadow:0 12px 26px rgba(0,0,0,.22)!important;
}
@media(max-width:700px){
  .login-page{
    background:
      radial-gradient(circle at 16% 86%, rgba(0,42,180,.38) 0 19%, transparent 20%),
      radial-gradient(circle at 92% 90%, rgba(0,100,30,.42) 0 22%, transparent 23%),
      linear-gradient(180deg,#020617 0%,#050505 100%)!important;
  }
  .login-inner .prize-logo-img{
    width:205px!important;
    max-height:104px!important;
    padding:9px 13px!important;
    border-radius:18px!important;
    margin-bottom:16px!important;
  }
  .hero{
    display:block!important;
    min-height:54px!important;
    padding:10px 12px!important;
  }
  .hero-title-only{
    text-align:left!important;
  }
  .hero-title-only h1{
    text-align:left!important;
    font-size:19px!important;
    line-height:1.1!important;
  }
  .hero-title-only p{
    text-align:left!important;
    font-size:11px!important;
  }
}



/* =========================================================
   AJUSTE FINAL SOLICITADO: TÍTULO CENTRADO + LOGO EN LETRAS
   - Se elimina el fondo blanco del logo.
   - Se usa wordmark textual PRIZE SUPERFRUITS para que impacte.
   - Header/título centrado en escritorio y celular.
   ========================================================= */
.prize-wordmark{
  display:inline-flex!important;
  flex-direction:column!important;
  align-items:center!important;
  justify-content:center!important;
  width:auto!important;
  max-width:100%!important;
  background:transparent!important;
  border:0!important;
  box-shadow:none!important;
  padding:0!important;
  margin:0 auto!important;
  text-align:center!important;
  line-height:1!important;
}
.prize-script{
  position:relative!important;
  display:inline-flex!important;
  align-items:flex-end!important;
  gap:0!important;
  font-family:"Segoe Script","Brush Script MT","Trebuchet MS",Arial,sans-serif!important;
  font-style:italic!important;
  font-weight:950!important;
  letter-spacing:-5px!important;
  color:#69b7dc!important;
  filter:drop-shadow(0 9px 18px rgba(0,0,0,.45)) drop-shadow(0 0 12px rgba(105,183,220,.28))!important;
  text-shadow:0 2px 0 rgba(0,0,0,.18)!important;
}
.prize-e{
  position:relative!important;
  display:inline-grid!important;
  place-items:center!important;
  margin-left:0!important;
  border-radius:50%!important;
  border:4px solid #2f78b7!important;
  color:#183e73!important;
  background:radial-gradient(circle at 66% 38%,#ffd34f 0 24%,#f59e0b 25% 52%,#dd4f0b 53% 100%)!important;
  font-family:"Trebuchet MS",Arial,sans-serif!important;
  font-style:italic!important;
  font-weight:950!important;
  letter-spacing:-2px!important;
  box-shadow:0 8px 18px rgba(245,111,16,.34), inset 0 0 0 2px rgba(255,255,255,.34)!important;
}
.prize-e i{
  position:absolute!important;
  right:5px!important;
  top:-32px!important;
  width:15px!important;
  height:34px!important;
  background:linear-gradient(135deg,#55c742,#168a33)!important;
  border-radius:100% 0 100% 0!important;
  transform:rotate(37deg)!important;
  box-shadow:0 5px 12px rgba(0,0,0,.25)!important;
}
.prize-e:before{
  content:""!important;
  position:absolute!important;
  right:20px!important;
  top:-24px!important;
  width:12px!important;
  height:28px!important;
  background:linear-gradient(135deg,#61d74b,#138f35)!important;
  border-radius:100% 0 100% 0!important;
  transform:rotate(-34deg)!important;
}
.prize-super{
  display:inline-block!important;
  margin-top:5px!important;
  padding:4px 18px 2px!important;
  color:#65d761!important;
  font-family:"Segoe UI",Arial,sans-serif!important;
  font-weight:950!important;
  letter-spacing:2px!important;
  border-top:2px solid rgba(101,215,97,.82)!important;
  text-shadow:0 5px 13px rgba(0,0,0,.42),0 0 10px rgba(101,215,97,.28)!important;
}
.prize-wordmark-login{margin-bottom:18px!important;}
.prize-wordmark-login .prize-script{font-size:72px!important;}
.prize-wordmark-login .prize-e{width:50px!important;height:50px!important;font-size:34px!important;}
.prize-wordmark-login .prize-super{font-size:14px!important;}
.prize-wordmark-hero{margin:0 auto 10px!important;}
.prize-wordmark-hero .prize-script{font-size:64px!important;}
.prize-wordmark-hero .prize-e{width:45px!important;height:45px!important;font-size:31px!important;}
.prize-wordmark-hero .prize-super{font-size:12px!important;}
.prize-wordmark-side{margin:8px auto 14px!important;}
.prize-wordmark-side .prize-script{font-size:45px!important;letter-spacing:-4px!important;}
.prize-wordmark-side .prize-e{width:32px!important;height:32px!important;font-size:22px!important;border-width:3px!important;}
.prize-wordmark-side .prize-e i{width:10px!important;height:24px!important;top:-24px!important;right:3px!important;}
.prize-wordmark-side .prize-e:before{width:8px!important;height:20px!important;top:-18px!important;right:14px!important;}
.prize-wordmark-side .prize-super{font-size:9px!important;letter-spacing:1px!important;padding:3px 10px 1px!important;}
.hero{
  display:flex!important;
  align-items:center!important;
  justify-content:center!important;
  text-align:center!important;
  background:radial-gradient(circle at 18% 0%,rgba(13,115,184,.36),transparent 34%),radial-gradient(circle at 82% 0%,rgba(22,163,74,.34),transparent 35%),linear-gradient(135deg,#061827 0%,#07304b 52%,#06361f 100%)!important;
}
.hero-title-only{
  width:100%!important;
  display:flex!important;
  flex-direction:column!important;
  align-items:center!important;
  justify-content:center!important;
  text-align:center!important;
}
.hero-title-only h1{
  width:100%!important;
  text-align:center!important;
  margin:0 auto 6px!important;
  font-weight:950!important;
  letter-spacing:.2px!important;
}
.hero-title-only p{
  width:100%!important;
  text-align:center!important;
  margin:0 auto!important;
}
.login-inner .prize-logo-img,.hero-brand .prize-logo-img,.side-logo-pro .prize-logo-img{display:none!important;}
.login-card{
  background:linear-gradient(180deg,rgba(3,7,18,.96),rgba(5,10,22,.92))!important;
  border:1px solid rgba(105,183,220,.32)!important;
}
@media(max-width:700px){
  .hero{display:flex!important;min-height:92px!important;padding:12px 10px!important;position:sticky!important;top:0!important;}
  .hero-title-only{text-align:center!important;align-items:center!important;}
  .hero-title-only h1{text-align:center!important;font-size:21px!important;line-height:1.08!important;}
  .hero-title-only p{text-align:center!important;font-size:11px!important;}
  .prize-wordmark-hero{margin-bottom:7px!important;}
  .prize-wordmark-hero .prize-script{font-size:44px!important;letter-spacing:-4px!important;}
  .prize-wordmark-hero .prize-e{width:32px!important;height:32px!important;font-size:22px!important;border-width:3px!important;}
  .prize-wordmark-hero .prize-e i{width:9px!important;height:21px!important;top:-22px!important;right:3px!important;}
  .prize-wordmark-hero .prize-e:before{width:7px!important;height:18px!important;top:-16px!important;right:14px!important;}
  .prize-wordmark-hero .prize-super{font-size:9px!important;letter-spacing:1px!important;padding:2px 10px 1px!important;}
  .prize-wordmark-login .prize-script{font-size:62px!important;}
  .prize-wordmark-login .prize-e{width:44px!important;height:44px!important;font-size:30px!important;}
  .prize-wordmark-login .prize-super{font-size:12px!important;}
}



/* =========================================================
   SIGUIENTE NIVEL UI - CORRECCIÓN DEFINITIVA
   1) Header SIN logo / SIN wordmark arriba del título.
   2) Logo Prize se mantiene SOLO en banda izquierda escritorio.
   3) En celular, menú horizontal dinámico tipo app.
   ========================================================= */
.hero .prize-wordmark,
.hero .prize-wordmark-hero,
.hero .hero-brand,
.hero img,
.hero .prize-logo-img{
  display:none!important;
  visibility:hidden!important;
  width:0!important;
  height:0!important;
  max-width:0!important;
  max-height:0!important;
  margin:0!important;
  padding:0!important;
  overflow:hidden!important;
}
.hero{
  display:flex!important;
  align-items:center!important;
  justify-content:center!important;
  text-align:center!important;
  min-height:112px!important;
  padding:22px 24px!important;
  background:
    radial-gradient(circle at 18% 0%,rgba(14,165,233,.24),transparent 35%),
    radial-gradient(circle at 86% 0%,rgba(34,197,94,.28),transparent 38%),
    linear-gradient(135deg,#061827 0%,#07304b 52%,#06361f 100%)!important;
  border-bottom:1px solid rgba(255,255,255,.10)!important;
  box-shadow:0 14px 36px rgba(0,0,0,.20)!important;
}
.hero-title-only{
  width:100%!important;
  max-width:980px!important;
  margin:0 auto!important;
  display:block!important;
  text-align:center!important;
}
.hero-title-only h1{
  text-align:center!important;
  margin:0 0 8px!important;
  font-size:46px!important;
  line-height:1.04!important;
  font-weight:950!important;
  letter-spacing:.2px!important;
  color:#fff!important;
  text-shadow:0 4px 18px rgba(0,0,0,.34)!important;
}
.hero-title-only h1:before{content:''!important;display:none!important;}
.hero-title-only p{
  text-align:center!important;
  margin:0!important;
  font-size:19px!important;
  font-weight:900!important;
  color:#dbeafe!important;
}
.fixed-prize-sidebar{
  background:
    radial-gradient(circle at 50% 2%,rgba(92,200,255,.16),transparent 32%),
    radial-gradient(circle at 96% 98%,rgba(34,197,94,.14),transparent 34%),
    linear-gradient(180deg,#05243a 0%,#041827 54%,#03131f 100%)!important;
}
.side-logo-pro{
  display:block!important;
  padding:16px 6px 18px!important;
  margin-bottom:10px!important;
  border-bottom:1px solid rgba(255,255,255,.14)!important;
  text-align:center!important;
}
.prize-wordmark-side{
  display:block!important;
  margin:4px auto 10px!important;
  padding:12px 6px 10px!important;
  border-radius:18px!important;
  background:radial-gradient(circle at 50% 22%,rgba(92,200,255,.18),transparent 42%),linear-gradient(135deg,rgba(255,255,255,.08),rgba(255,255,255,.025))!important;
  box-shadow:inset 0 0 0 1px rgba(255,255,255,.10),0 14px 30px rgba(0,0,0,.22)!important;
}
.nav-pro{
  display:flex!important;
  flex-direction:column!important;
  gap:6px!important;
}
.nav-pro a{
  border:1px solid rgba(255,255,255,.08)!important;
  background:rgba(255,255,255,.035)!important;
  transition:transform .15s ease,background .15s ease,box-shadow .15s ease!important;
}
.nav-pro a:hover,.nav-pro a.on{
  background:linear-gradient(90deg,#16a34a,#0d73b8)!important;
  box-shadow:0 10px 24px rgba(13,115,184,.22)!important;
  transform:translateX(3px)!important;
}
@media(max-width:700px){
  .hero{
    min-height:64px!important;
    padding:12px 10px!important;
    position:sticky!important;
    top:0!important;
    z-index:90!important;
  }
  .hero-title-only h1{
    font-size:22px!important;
    line-height:1.08!important;
    margin:0 0 3px!important;
  }
  .hero-title-only p{
    font-size:11px!important;
    line-height:1.15!important;
    display:block!important;
  }
  .fixed-prize-sidebar{
    position:sticky!important;
    top:64px!important;
    z-index:85!important;
    width:100%!important;
    height:auto!important;
    min-height:0!important;
    padding:8px 8px!important;
    overflow-x:auto!important;
    overflow-y:hidden!important;
    -webkit-overflow-scrolling:touch!important;
    border-bottom:1px solid rgba(255,255,255,.08)!important;
  }
  .side-logo-pro,.side-user-card,.side-slogan-card{display:none!important;}
  .nav-pro{
    display:flex!important;
    flex-direction:row!important;
    gap:8px!important;
    min-width:max-content!important;
    padding:0 2px!important;
  }
  .nav-pro a{
    flex:0 0 auto!important;
    min-width:108px!important;
    min-height:44px!important;
    justify-content:center!important;
    text-align:center!important;
    margin:0!important;
    border-radius:14px!important;
    font-size:11px!important;
    padding:9px 10px!important;
  }
  .nav-pro a:hover,.nav-pro a.on{transform:none!important;}
  .nav-pro .pill{display:none!important;}
  .main-layout{
    margin-left:0!important;
    width:100%!important;
    display:block!important;
    height:auto!important;
    overflow:visible!important;
  }
  .content{
    height:auto!important;
    overflow:visible!important;
    padding:12px 10px 30px!important;
  }
}

</style>
<script src="https://unpkg.com/html5-qrcode.3.8/html5-qrcode.min.js" crossorigin="anonymous"></script>
<script src="https://unpkg.com/@zxing/library@0.20.0/umd/index.min.js" crossorigin="anonymous"></script>
<script src="https://cdn.jsdelivr.net/npm/jsqr@1.4.0/dist/jsQR.min.js" crossorigin="anonymous"></script>
</head>
<body>

{% if not session.get('user') %}
  {{content|safe}}
{% else %}
<div class="app-shell">

  <header class="hero">
    <div class="hero-title-only">
      <h1>Sistema Comedor PRIZE</h1>
      <p>ERP para la Gestión del Comedor Corporativo</p>
    </div>

  </header>

<div class="main-layout">
    <aside class="sidebar fixed-prize-sidebar">
      <div class="side-logo-pro">
        <div class="prize-wordmark prize-wordmark-side"><div class="prize-script">Prize<span class="prize-e">e<i></i></span></div><div class="prize-super">SUPERFRUITS</div></div>
      </div>

      <div class="side-user-card">
        <div class="side-avatar">👤</div>
        <div class="side-user-title">ERP Comedor</div>
        <div class="side-user-sub">{{session.get('user','admin')}} · {{session.get('role','admin')}}</div>
      </div>

      <nav class="nav nav-pro">
        {% if session.get('role') == 'admin' %}
        <a class="{{'on' if page=='dashboard'}}" href="{{url_for('dashboard')}}"><span class="nav-ico">📊</span>Dashboard</a>
        {% endif %}
        <a class="{{'on' if page=='consumos'}}" href="{{url_for('consumos')}}"><span class="nav-ico">🍽️</span>Consumos</a>
        {% if session.get('role') == 'admin' %}
        <a class="{{'on' if page=='trabajadores'}}" href="{{url_for('trabajadores')}}"><span class="nav-ico">👥</span>Trabajadores</a>
        {% endif %}
        <a class="{{'on' if page=='entregas'}}" href="{{url_for('entregas')}}"><span class="nav-ico">🚚</span>Entregas <span class="pill nuevo">NUEVO</span></a>
        {% if session.get('role') == 'admin' %}
        <a class="{{'on' if page=='reportes'}}" href="{{url_for('reportes')}}"><span class="nav-ico">📁</span>Reportes <span class="pill correo">CORREO</span></a>
        {% endif %}
        <a class="{{'on' if page=='cierre'}}" href="{{url_for('cierre_dia')}}"><span class="nav-ico">📁</span>Cerrar día</a>
        {% if session.get('role') == 'admin' %}
        <a class="{{'on' if page=='carga'}}" href="{{url_for('carga_masiva')}}"><span class="nav-ico">📥</span>Carga Masiva</a>
        <a class="{{'on' if page=='config'}}" href="{{url_for('configuracion')}}"><span class="nav-ico">⚙️</span>Config. / Usuarios</a>
        {% endif %}
        <a class="logout-link" href="{{url_for('logout')}}"><span class="nav-ico">🚪</span>Salir</a>
      </nav>

      <div class="side-slogan-card">
        <div class="leaf-icon"></div>
        <b>Comer bien,</b><br>es vivir mejor.
      </div>
    </aside>

    <main class="content">
      {% with messages=get_flashed_messages(with_categories=true) %}
        {% for c,m in messages %}
          <div class="flash {{c}}">{{m}}</div>
        {% endfor %}
      {% endwith %}
      {{content|safe}}
    </main>

    <aside class="panel-right">
      <div class="card status-box">
        <h3 style="margin-top:0">Estado del día</h3>
        <div class="status-inner">
          <span class="badge {{'off' if cerrado_hoy else 'ok'}}">🟢 {{'DÍA CERRADO' if cerrado_hoy else 'DÍA ABIERTO'}}</span>
          <p class="small" style="line-height:1.7">
            <b>Fecha:</b> {{fecha_hoy}}<br>
            <b>{{'Cerrado' if cerrado_hoy else 'Abierto'}} por:</b> admin (08:00 AM)
          </p>
          {% if not cerrado_hoy %}
            <a class="btn btn-orange" style="width:100%;text-align:center" href="{{url_for('cierre_dia')}}">Cerrar día y consolidar</a>
          {% endif %}
        </div>
      </div>

      <div class="card">
        <h3 style="margin-top:0">Acciones rápidas</h3>
        <div class="quick">
          <a href="{{url_for('consumos')}}">🔹 Registrar consumo</a>
          <a href="{{url_for('entregas')}}">🚚 Entrega de pedidos</a>
          <a href="{{url_for('carga_masiva')}}">📥 Carga masiva de consumos</a>
          <a href="{{url_for('reportes')}}">✉️ Enviar reporte por correo</a>
        </div>
      </div>
    </aside>
  </div>

  <footer class="footer">
    <span>© 2026 Prize Superfruits - Comedor Corporativo. Todos los derechos reservados.</span>
    <span>Versión 2.0.0</span>
  </footer>
</div>
{% endif %}
<script>
// ===== PRO TOTAL: DNI automático + cámara QR/BARRAS para CONSUMOS =====
(function(){
  let proTimer = null;
  let proScanner = null;
  let proStream = null;
  let proBusy = false;

  function dniClean(v){
    const raw = String(v || '').trim();
    const only = raw.replace(/\D/g,'');
    if (only.length === 8) return only;
    const labeled = raw.toUpperCase().match(/(?:DNI|DOC(?:UMENTO)?|DOCUMENT|NRO|NUMERO|NÚMERO)\D{0,20}(\d{8})(?!\d)/);
    if (labeled) return labeled[1];
    const eight = raw.match(/(^|\D)(\d{8})(?!\d)/);
    if (eight) return eight[2];
    if (only.length > 8) return only.slice(-8);
    return only.slice(0,8);
  }
  function toast(msg, ok=true){
    const d = document.createElement('div');
    d.className = 'prize-toast-msg';
    d.textContent = msg;
    const visibles = document.querySelectorAll('.prize-toast-msg').length;
    const topPx = 12 + (visibles * 58);
    d.style.cssText = 'position:fixed;left:10px;right:10px;top:calc(env(safe-area-inset-top,0px) + '+topPx+'px);z-index:2147483647;padding:12px 14px;border-radius:12px;font-weight:950;color:white;text-align:center;box-shadow:0 12px 30px rgba(0,0,0,.35);background:'+(ok?'#006b1e':'#a40000')+';border:1px solid rgba(255,255,255,.18);font-size:13px;line-height:1.2;pointer-events:none;';
    document.body.appendChild(d); setTimeout(()=>d.remove(), 2300);
    try{ if(navigator.vibrate) navigator.vibrate(ok?90:[80,50,80]); }catch(e){}
  }
  function beep(){
    try{
      const C = window.AudioContext || window.webkitAudioContext;
      const c = new C(); const o = c.createOscillator(); const g = c.createGain();
      o.connect(g); g.connect(c.destination); o.frequency.value = 920; g.gain.value = .08;
      o.start(); setTimeout(()=>{o.stop(); c.close();}, 130);
    }catch(e){}
  }
  function setNombre(data, dni){
    const nombre = document.getElementById('nombre_trabajador') || document.querySelector('[name="nombre"],#nombre');
    const info = document.getElementById('info_trabajador_consumo');
    if(data && (data.ok || data.success || data.nombre)){
      if(nombre){ nombre.value = data.nombre || ''; nombre.title = data.nombre || ''; }
      if(info){
        info.style.display='block';
        info.innerHTML = '<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:10px"><div><b>Trabajador</b><br>'+(data.nombre||'-')+'</div><div><b>DNI</b><br>'+dni+'</div><div><b>Área</b><br>'+(data.area||'-')+'</div><div><b>Estado</b><br><span class="badge ok">Activo</span></div></div>';
      }
      beep(); return true;
    } else {
      if(nombre){ nombre.value = 'DNI no encontrado'; nombre.title = 'DNI no encontrado'; }
      if(info){ info.style.display='block'; info.innerHTML='<span style="color:#991b1b">DNI no encontrado en Trabajadores: '+dni+'</span>'; }
      return false;
    }
  }
  function responsableActual(){
    const r = document.querySelector('#form_consumo [name="responsable"]') || document.querySelector('[name="responsable"]');
    return String(r && r.value ? r.value : '').trim().toUpperCase();
  }
  function limpiarParaSiguiente(){
    const inp = document.getElementById('dni_consumo') || document.querySelector('input[name="dni"]');
    const nombre = document.getElementById('nombre_trabajador') || document.querySelector('[name="nombre"],#nombre');
    const info = document.getElementById('info_trabajador_consumo');
    if(inp){ inp.value=''; setTimeout(()=>inp.focus(), 80); }
    if(nombre){ nombre.value=''; nombre.title=''; }
    if(info){ info.style.display='none'; info.innerHTML=''; }
  }
  function agregarFilaConsumoAuto(rowHtml){
    const tbody = document.getElementById('tbody_consumos_principal');
    if(!tbody || !rowHtml) return;
    const sin = document.getElementById('fila_sin_registros');
    if(sin) sin.remove();
    tbody.insertAdjacentHTML('afterbegin', rowHtml);
  }
  let guardandoAutoBase = false;
  async function buscarAutoDniConsumo(force=false){
    const inp = document.getElementById('dni_consumo') || document.querySelector('input[name="dni"]');
    if(!inp || guardandoAutoBase) return;
    const dni = dniClean(inp.value);
    inp.value = dni;
    const nombre = document.getElementById('nombre_trabajador') || document.querySelector('[name="nombre"],#nombre');
    if(dni.length < 8){ if(nombre) nombre.value=''; return; }

    if(!responsableActual()){
      if(nombre) nombre.value = 'PRIMERO COLOCA RESPONSABLE';
      toast('Primero registra el RESPONSABLE antes de detectar DNI.', false);
      limpiarParaSiguiente();
      return;
    }

    guardandoAutoBase = true;
    if(nombre) nombre.value = 'Validando y guardando DNI...';
    try{
      const r = await fetch('/api/trabajador/' + encodeURIComponent(dni) + '?_=' + Date.now(), {cache:'no-store', credentials:'same-origin'});
      const data = await r.json();
      const ok = setNombre(data, dni);
      if(!ok){
        toast('DNI no encontrado: ' + dni, false);
        limpiarParaSiguiente();
        return;
      }

      const form = document.getElementById('form_consumo');
      const fd = new FormData(form || document.createElement('form'));
      fd.set('dni', dni);
      fd.set('modo_lote', '0');
      const rr = await fetch('/api/registrar_consumo_auto', {method:'POST', body:fd, credentials:'same-origin', cache:'no-store'});
      const res = await rr.json().catch(()=>({ok:false,msg:'No se pudo leer respuesta del servidor'}));
      if(res.ok){
        agregarFilaConsumoAuto(res.row_html || '');
        toast(res.msg || ('Guardado automático: ' + dni), true);
        beep();
        const ind = document.getElementById('indicador_masivo_contador');
        if(ind) ind.textContent = 'Guardado automático';
        limpiarParaSiguiente();
      }else{
        toast(res.msg || 'No se pudo guardar el consumo.', false);
        limpiarParaSiguiente();
      }
    }catch(e){
      if(nombre) nombre.value='Error validando/guardando DNI';
      toast('Error de conexión al guardar consumo.', false);
      limpiarParaSiguiente();
    }finally{
      setTimeout(()=>{ guardandoAutoBase=false; }, 250);
    }
  }
  window.buscarAutoDniConsumo = buscarAutoDniConsumo;
  window.dniInputHandler = function(){
    const inp = document.getElementById('dni_consumo') || document.querySelector('input[name="dni"]');
    if(!inp) return;
    inp.value = dniClean(inp.value);
    clearTimeout(proTimer);
    proTimer = setTimeout(()=>buscarAutoDniConsumo(false), inp.value.length === 8 ? 20 : 120);
  };
  async function procesarLectura(texto){
    if(proBusy) return;
    const dni = dniClean(texto);
    if(dni.length !== 8){ toast('No detecté un DNI de 8 dígitos.', false); return; }
    proBusy = true;
    const inp = document.getElementById('dni_consumo') || document.querySelector('input[name="dni"]');
    if(inp) inp.value = dni;
    await buscarAutoDniConsumo(true);
    toast('DNI leído: ' + dni, true);
    setTimeout(()=>{proBusy=false;}, 900);
  }
  window.abrirScannerQR = async function(){
    let cont = document.getElementById('qr-reader');
    if(!cont){
      cont = document.createElement('div'); cont.id = 'qr-reader';
      const form = document.getElementById('form_consumo') || document.body;
      form.appendChild(cont);
    }
    if(location.protocol !== 'https:' && !['localhost','127.0.0.1'].includes(location.hostname)){
      toast('La cámara requiere HTTPS. Usa el enlace de Render con https://', false);
    }
    cont.style.display='block';
    cont.innerHTML = '<div class="qr-camera-box"><b>📷 Cámara QR / Barras activa</b><div id="qr-reader-live" class="qr-live-box"></div><video id="qr-video-live" class="qr-video-box" playsinline muted autoplay style="display:none"></video><canvas id="qr-canvas-live" style="display:none"></canvas><button type="button" class="btn-red qr-close-btn" onclick="cerrarScannerQR()">Cerrar cámara</button><small>Permite la cámara. En celular usa Chrome y HTTPS.</small></div>';
    try{
      if(window.Html5Qrcode){
        const formats = window.Html5QrcodeSupportedFormats ? [
          Html5QrcodeSupportedFormats.QR_CODE, Html5QrcodeSupportedFormats.CODE_128,
          Html5QrcodeSupportedFormats.CODE_39, Html5QrcodeSupportedFormats.EAN_13,
          Html5QrcodeSupportedFormats.EAN_8, Html5QrcodeSupportedFormats.ITF,
          Html5QrcodeSupportedFormats.UPC_A, Html5QrcodeSupportedFormats.UPC_E,
          Html5QrcodeSupportedFormats.PDF_417
        ].filter(Boolean) : undefined;
        proScanner = new Html5Qrcode('qr-reader-live', formats ? {formatsToSupport:formats, verbose:false} : undefined);
        await proScanner.start({facingMode:{ideal:'environment'}}, {fps:15, qrbox:{width:220,height:220}, rememberLastUsedCamera:true}, async txt=>{
          await procesarLectura(txt);
          if(!document.getElementById('modo_lote')?.checked) cerrarScannerQR();
        }, ()=>{});
        toast('Cámara activada.', true); return;
      }
    }catch(e){ console.warn('html5-qrcode no abrió, usando respaldo', e); }
    await scannerNativo();
  };
  async function scannerNativo(){
    if(!navigator.mediaDevices || !navigator.mediaDevices.getUserMedia) throw new Error('Navegador sin cámara');
    const video = document.getElementById('qr-video-live'); const canvas = document.getElementById('qr-canvas-live');
    const live = document.getElementById('qr-reader-live'); if(live) live.innerHTML='<b>Respaldo con cámara nativa...</b>';
    proStream = await navigator.mediaDevices.getUserMedia({video:{facingMode:{ideal:'environment'}}, audio:false});
    video.srcObject = proStream; video.style.display='block'; await video.play();
    let detector = null;
    if('BarcodeDetector' in window){ try{ detector = new BarcodeDetector({formats:['qr_code','code_128','code_39','ean_13','ean_8','itf','upc_a','upc_e','pdf417']}); }catch(e){} }
    const loop = async()=>{
      if(!proStream) return;
      try{
        if(detector){ const codes = await detector.detect(video); if(codes && codes.length){ await procesarLectura(codes[0].rawValue||''); if(!document.getElementById('modo_lote')?.checked){cerrarScannerQR(); return;} } }
        if(window.jsQR && video.videoWidth){
          canvas.width=video.videoWidth; canvas.height=video.videoHeight;
          const ctx=canvas.getContext('2d',{willReadFrequently:true}); ctx.drawImage(video,0,0,canvas.width,canvas.height);
          const img=ctx.getImageData(0,0,canvas.width,canvas.height); const code=jsQR(img.data,img.width,img.height);
          if(code && code.data){ await procesarLectura(code.data); if(!document.getElementById('modo_lote')?.checked){cerrarScannerQR(); return;} }
        }
      }catch(e){}
      requestAnimationFrame(loop);
    };
    toast('Cámara activada.', true); requestAnimationFrame(loop);
  }
  window.cerrarScannerQR = function(){
    try{ if(proScanner && proScanner.stop) proScanner.stop().catch(()=>{}).finally(()=>{try{proScanner.clear();}catch(e){}}); }catch(e){}
    try{ if(proStream){ proStream.getTracks().forEach(t=>t.stop()); } }catch(e){}
    proScanner=null; proStream=null;
    const cont=document.getElementById('qr-reader'); if(cont){cont.style.display='none'; cont.innerHTML='';}
  };
  document.addEventListener('DOMContentLoaded', function(){
    const inp = document.getElementById('dni_consumo');
    if(inp){
      inp.setAttribute('oninput','dniInputHandler()');
      inp.setAttribute('onkeyup','dniInputHandler()');
      inp.addEventListener('input', window.dniInputHandler, true);
      inp.addEventListener('paste', ()=>setTimeout(window.dniInputHandler, 30), true);
      inp.addEventListener('keydown', e=>{ if(e.key==='Enter'){e.preventDefault(); buscarAutoDniConsumo(true);} }, true);
      setTimeout(()=>inp.focus(),250);
    }
    const btn = document.getElementById('btn_qr');
    if(btn) btn.onclick = window.abrirScannerQR;
  });
})();
</script>


<script>
// ===== FIX FINAL: alertas apiladas y cámara centrada/angosta =====
(function(){
  function premioAviso(msg, ok=true){
    try{
      const div = document.createElement('div');
      div.className = 'prize-toast-msg';
      div.textContent = msg;
      const visibles = document.querySelectorAll('.prize-toast-msg').length;
      const topPx = 12 + (visibles * 58);
      div.style.cssText = 'position:fixed;left:10px;right:10px;top:calc(env(safe-area-inset-top,0px) + '+topPx+'px);z-index:2147483647;padding:12px 14px;border-radius:12px;font-weight:950;color:white;text-align:center;box-shadow:0 12px 30px rgba(0,0,0,.35);background:'+(ok?'#006b1e':'#a40000')+';border:1px solid rgba(255,255,255,.18);font-size:13px;line-height:1.2;pointer-events:none;';
      document.body.appendChild(div);
      setTimeout(()=>div.remove(), 2300);
    }catch(e){ alert(msg); }
  }
  window.avisoMovil = premioAviso;
  window.toastFix = premioAviso;
  function ajustarCamara(){
    const cont = document.getElementById('qr-reader');
    if(!cont) return;
    cont.style.maxWidth = '390px';
    cont.style.marginLeft = 'auto';
    cont.style.marginRight = 'auto';
    const box = cont.firstElementChild;
    if(box){ box.classList.add('qr-camera-box'); box.style.maxWidth='390px'; box.style.margin='0 auto'; box.style.textAlign='center'; }
    const live = document.getElementById('qr-reader-live');
    if(live){ live.classList.add('qr-live-box'); live.style.maxWidth='340px'; live.style.margin='10px auto 8px'; }
    const video = document.getElementById('qr-video-live');
    if(video){ video.classList.add('qr-video-box'); video.style.maxWidth='340px'; video.style.margin='10px auto 8px'; video.style.objectFit='cover'; }
  }
  const oldOpen = window.abrirScannerQR;
  if(typeof oldOpen === 'function'){
    window.abrirScannerQR = async function(){
      const r = await oldOpen.apply(this, arguments);
      setTimeout(ajustarCamara, 80);
      setTimeout(ajustarCamara, 500);
      return r;
    };
  }
  document.addEventListener('DOMContentLoaded', ()=>{
    const btn = document.getElementById('btn_qr');
    if(btn && typeof window.abrirScannerQR === 'function') btn.onclick = window.abrirScannerQR;
    ajustarCamara();
  });
})();
</script>


<script>
// ===== CONTADOR VISUAL DE LECTURAS GUARDADAS =====
(function(){
  function contarFilasDb(){
    try { return document.querySelectorAll('#tbody_consumos_principal tr.fila-db-consumo').length; }
    catch(e){ return 0; }
  }
  function actualizarContadorLecturas(delta){
    const c = document.getElementById('contador_lecturas_hoy');
    if(!c) return;
    let n = parseInt(c.textContent || '0', 10) || 0;
    if(typeof delta === 'number') n += delta;
    else n = Math.max(n, contarFilasDb());
    c.textContent = n;
    const box = document.getElementById('contador_lecturas_box');
    if(box){
      box.style.transform = 'scale(1.015)';
      box.style.boxShadow = '0 0 0 4px rgba(34,197,94,.22), 0 14px 32px rgba(22,163,74,.25)';
      setTimeout(()=>{ box.style.transform=''; box.style.boxShadow='0 10px 24px rgba(22,163,74,.22)'; }, 260);
    }
  }
  window.actualizarContadorLecturas = actualizarContadorLecturas;
  document.addEventListener('DOMContentLoaded', ()=>actualizarContadorLecturas());

  const oldFetch = window.fetch;
  window.fetch = async function(){
    const res = await oldFetch.apply(this, arguments);
    try{
      const url = String(arguments[0] || '');
      if(url.includes('/api/registrar_consumo_auto')){
        const clone = res.clone();
        clone.json().then(data=>{ if(data && data.ok) actualizarContadorLecturas(1); }).catch(()=>{});
      }
    }catch(e){}
    return res;
  };
})();
</script>

</body>
</html>
"""


def render_page(content, page=""):
    pendientes_count = q_one(
        "SELECT COUNT(*) c FROM consumos WHERE fecha=? AND estado='PENDIENTE'",
        (hoy_iso(),)
    )["c"]
    return render_template_string(
        BASE_HTML,
        content=content,
        page=page,
        pendientes_count=pendientes_count,
        cerrado_hoy=bool(dia_cerrado()),
        fecha_hoy=fecha_peru_txt(),
        money=money,
    )


def topbar(title, subtitle="Resumen general del sistema"):
    return f"""
    <div class="topbar">
      <div>
        <h2>{title}</h2>
        <div class="muted">{subtitle}</div>
      </div>
      <div class="user-chip">
        <span style="font-size:24px">🔔</span>
        <div class="avatar">👤</div>
        <div><span class="small">Bienvenido,</span><br>{session.get('user','')}</div>
      </div>
    </div>
    """


# =========================
# RUTAS
# =========================

@app.route("/cerrar_dia_manual")
@login_required
@roles_required("admin")
def cerrar_dia_manual():
    fecha = hoy_iso()
    if dia_cerrado(fecha):
        flash("El día ya estaba cerrado.", "error")
    else:
        q_exec("""
            INSERT INTO cierres(fecha,cerrado_por,total_consumos,total_entregados,total_pendientes,total_importe,archivo_excel,correo_destino,correo_estado)
            VALUES(?,?,?,?,?,?,?,?,?)
        """, (fecha, session["user"], 0, 0, 0, 0, "", "", "CIERRE MANUAL"))
        flash("Día cerrado manualmente por administrador.", "ok")
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/abrir_dia_manual")
@login_required
@roles_required("admin")
def abrir_dia_manual():
    fecha = hoy_iso()
    q_exec("DELETE FROM cierres WHERE fecha=?", (fecha,))
    flash("Día abierto/reabierto correctamente por administrador.", "ok")
    return redirect(request.referrer or url_for("dashboard"))


def rows_filtrados_desde_request(solo_entregados=False):
    fecha_inicio = request.args.get("fecha_inicio") or request.args.get("fecha") or hoy_iso()
    fecha_fin = request.args.get("fecha_fin") or fecha_inicio
    buscar = clean_text(request.args.get("buscar"))
    cond, params = rango_sql(fecha_inicio, fecha_fin)
    where = cond
    final_params = list(params)
    if solo_entregados:
        where += " AND estado='ENTREGADO'"
    if buscar:
        where += " AND (dni LIKE ? OR trabajador LIKE ? OR area LIKE ? OR fundo LIKE ? OR comedor LIKE ?)"
        b = f"%{buscar}%"
        final_params += [b, b, b, b, b]
    rows = q_all(f"SELECT * FROM consumos WHERE {where} ORDER BY fecha,hora,id", tuple(final_params))
    return fecha_inicio, fecha_fin, buscar, rows


@app.route("/exportar_concesionaria")
@login_required
def exportar_concesionaria():
    fecha_inicio, fecha_fin, buscar, rows = rows_filtrados_desde_request(False)
    df = pd.DataFrame([dict(r) for r in rows])
    if df.empty:
        df = pd.DataFrame(columns=["fecha","hora","dni","trabajador","area","tipo","comedor","fundo","responsable","cantidad","total","estado"])
    filename = f"consumos_concesionaria_{fecha_inicio.replace('-','_')}_a_{fecha_fin.replace('-','_')}.xlsx"
    path = os.path.join(CONCESIONARIA_DIR, filename)
    df.to_excel(path, index=False)
    return send_file(path, as_attachment=True)


@app.route("/reporte_entrega")
@login_required
def reporte_entrega():
    fecha_inicio, fecha_fin, buscar, rows = rows_filtrados_desde_request(True)
    df = pd.DataFrame([dict(r) for r in rows])
    if df.empty:
        df = pd.DataFrame(columns=["fecha","hora","dni","trabajador","area","tipo","comedor","fundo","responsable","cantidad","total","estado"])
    filename = f"reporte_entrega_pago_{fecha_inicio.replace('-','_')}_a_{fecha_fin.replace('-','_')}.xlsx"
    path = os.path.join(ENTREGAS_DIR, filename)
    df.to_excel(path, index=False)
    return send_file(path, as_attachment=True)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = clean_text(request.form.get("username"))
        password = request.form.get("password", "")
        user = q_one("SELECT * FROM usuarios WHERE username=? AND active=1", (username,))
        if user and check_password_hash(user["password_hash"], password):
            session["user"] = user["username"]
            session["role"] = user["role"]
            return redirect(url_for("dashboard"))
        flash("Usuario o clave incorrecta.", "error")

    html = f"""
    <div class="login-page">
      <div class="login-card">
        <div class="login-inner">
          <div class="prize-wordmark prize-wordmark-login"><div class="prize-script">Prize<span class="prize-e">e<i></i></span></div><div class="prize-super">SUPERFRUITS</div></div>
          <h2 class="login-title">Sistema Comedor PRIZE</h2>
          <p class="login-subtitle">Acceso al sistema</p>

          <form method="post">
            <div class="form-label">Usuario</div>
            <div class="input-icon"><span>👤</span><input name="username" placeholder="Ingrese su usuario" required></div>

            <div class="form-label">Clave</div>
            <div class="input-icon"><span>🔒</span><input name="password" type="password" placeholder="Ingrese su clave" required></div>

            <button class="login-button">Ingresar</button>
          </form>

        </div>
      </div>
    </div>
    """
    return render_template_string(BASE_HTML, content=html)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
@login_required
def dashboard():
    if session.get("role") != "admin":
        return redirect(url_for("consumos"))
    fecha_inicio = request.args.get("fecha_inicio") or request.args.get("fecha") or hoy_iso()
    fecha_fin = request.args.get("fecha_fin") or fecha_inicio
    buscar = clean_text(request.args.get("buscar"))
    cond, params = rango_sql(fecha_inicio, fecha_fin)

    where = cond
    final_params = list(params)
    if buscar:
        where += " AND (dni LIKE ? OR trabajador LIKE ? OR area LIKE ? OR fundo LIKE ? OR comedor LIKE ?)"
        b = f"%{buscar}%"
        final_params += [b, b, b, b, b]

    total_filtro = q_one(f"SELECT COUNT(*) c, COALESCE(SUM(total),0) t FROM consumos WHERE {where}", tuple(final_params))
    entregados = q_one(f"SELECT COUNT(*) c FROM consumos WHERE {where} AND estado='ENTREGADO'", tuple(final_params))["c"]
    pendientes = q_one(f"SELECT COUNT(*) c FROM consumos WHERE {where} AND estado='PENDIENTE'", tuple(final_params))["c"]
    trabajadores = q_one("SELECT COUNT(*) c FROM trabajadores WHERE activo=1")["c"]

    rows = q_all(f"SELECT * FROM consumos WHERE {where} ORDER BY fecha DESC,hora DESC,id DESC LIMIT 12", tuple(final_params))
    total_consumos_fecha = int((q_one("SELECT COUNT(*) AS c FROM consumos WHERE fecha=?", (hoy_iso(),)) or {"c": 0})["c"] or 0)

    tabla = "".join([
        f"""
        <tr>
          <td>{i}</td><td>{r['fecha']}</td><td>{r['hora']}</td><td>{r['dni']}</td><td>{r['trabajador']}</td>
          <td>{r['area']}</td><td>{r['tipo']}</td><td>{r['comedor']}</td><td>{r['fundo']}</td>
          <td>{r['cantidad']}</td><td>{money(r['total'])}</td>
        </tr>
        """
        for i, r in enumerate(rows, 1)
    ]) or "<tr><td colspan='11'>Sin consumos con el filtro seleccionado.</td></tr>"

    admin_buttons = ""
    if session.get("role") == "admin":
        admin_buttons = f"""
        <div class="admin-actions">
          <a class="btn btn-orange" href="{url_for('cerrar_dia_manual')}">🔒 Cerrar día</a>
          <a class="btn btn-blue" href="{url_for('abrir_dia_manual')}">🔓 Abrir día</a>
        </div>
        """

    html = topbar("Dashboard", "Indicadores filtrados por día, mes o año") + admin_buttons
    html += filtro_bar(url_for("dashboard"), fecha_inicio, fecha_fin, buscar)
    html += f"""
    <div class="kpi-grid dashboard-kpis-center">
      <div class="card kpi-card"><div class="icon-circle ic-green">🍴</div><div><div class="label">Consumos filtrados</div><div class="num">{total_filtro['c']}</div><div class="sub">RANGO</div></div></div>
      <div class="card kpi-card"><div class="icon-circle ic-blue">✅</div><div><div class="label">Entregados</div><div class="num">{entregados}</div><div class="sub">confirmados</div></div></div>
      <div class="card kpi-card"><div class="icon-circle ic-purple">⏳</div><div><div class="label">Pendientes</div><div class="num">{pendientes}</div><div class="sub">por entregar</div></div></div>
      <div class="card kpi-card"><div class="icon-circle ic-orange">S/</div><div><div class="label">Total filtrado</div><div class="num" style="color:#16a34a">{money(total_filtro['t'])}</div><div class="sub">trabajadores activos: {trabajadores}</div></div></div>
    </div>

    <div class="card">
      <div class="table-head">
        <h3>Consumos filtrados</h3>
        <div style="display:flex;gap:8px;flex-wrap:wrap">
          <a class="btn btn-blue" href="{url_for('exportar_concesionaria', fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, buscar=buscar)}">Archivo concesionaria</a>
          <a class="btn btn-orange" href="{url_for('reporte_entrega', fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, buscar=buscar)}">Reporte entrega/pago</a>
        </div>
      </div>
      <div class="table-wrap">
        <table>
          <tr><th>#</th><th>Fecha</th><th>Hora</th><th>DNI</th><th>Trabajador</th><th>Área</th><th>Tipo</th><th>Comedor</th><th>Fundo</th><th>Cant.</th><th>Total</th></tr>
          {tabla}
        </table>
      </div>
    </div>
    """
    return render_page(html, "dashboard")



@app.route("/api/trabajador/<dni>")
@login_required
def api_trabajador(dni):
    dni = clean_dni(dni)
    t = q_one("SELECT dni,nombre,empresa,area,cargo FROM trabajadores WHERE dni=? AND activo=1", (dni,))
    if not t:
        resp = jsonify({"ok": False, "success": False, "msg": "DNI no encontrado"})
    else:
        resp = jsonify({"ok": True, "success": True, "dni": t["dni"], "nombre": t["nombre"], "empresa": t["empresa"], "area": t["area"], "cargo": t["cargo"]})
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return resp

@app.route("/api/buscar_dni/<dni>")
@login_required
def api_buscar_dni(dni):
    return api_trabajador(dni)

@app.route("/buscar_trabajador/<dni>")
@login_required
def buscar_trabajador_compat(dni):
    return api_trabajador(dni)

@app.route("/api/trabajador")
@login_required
def api_trabajador_query():
    return api_trabajador(request.args.get("dni", ""))

@app.route("/consumos", methods=["GET", "POST"])
@login_required
@roles_required("admin", "rrhh", "comedor")
def consumos():
    if request.method == "POST":
        fecha = request.form.get("fecha") or hoy_iso()

        if fecha != hoy_iso():
            flash("Solo se puede registrar consumo en la fecha actual de hoy. Las fechas anteriores o futuras son solo de consulta.", "error")
            return redirect(url_for("consumos", fecha=fecha))

        if dia_cerrado(fecha):
            flash("El día ya está cerrado. No se puede registrar consumos. Al día siguiente el sistema abrirá automáticamente la nueva fecha.", "error")
            return redirect(url_for("consumos", fecha=fecha))

        bloqueado, msg = registro_bloqueado()
        if bloqueado and session.get("role") != "admin":
            flash(msg, "error")
            return redirect(url_for("consumos"))

        tipo = request.form.get("tipo", "Almuerzo")
        if tipo not in ["Almuerzo", "Dieta"]:
            tipo = "Almuerzo"

        comedor = request.form.get("comedor", "Comedor 01")
        fundo = request.form.get("fundo", "Kawsay Allpa")
        responsable = clean_text(request.form.get("responsable")).upper()
        if not responsable:
            flash("El campo RESPONSABLE es obligatorio y debe ir en MAYÚSCULAS.", "error")
            return redirect(url_for("consumos", fecha=fecha))
        cantidad = int(float(request.form.get("cantidad") or 1))
        precio = float(request.form.get("precio_unitario") or 10)
        total = cantidad * precio
        obs = clean_text(request.form.get("observacion"))
        es_adicional = 1 if request.form.get("adicional") == "1" and session.get("role") == "admin" else 0

        # REGISTRO MASIVO / EN LOTE desde la misma pestaña Consumos.
        if request.form.get("modo_lote") == "1":
            lote_raw = request.form.get("dni_lote", "")
            dnis = []
            for part in re.split(r"[\s,;]+", lote_raw):
                d = clean_dni(part)
                if d and d not in dnis:
                    dnis.append(d)
            # FIX FINAL: en modo masivo SOLO se registran DNIs que aparecieron en el indicador/lote.
            # No se usa el DNI visible como respaldo, porque eso permitía registrar sin ver el indicador.
            if not dnis:
                flash("Registro masivo activo, pero aún no hay DNIs guardados en el lote. Digita o escanea un DNI válido y espera el mensaje verde de guardado.", "error")
                return redirect(url_for("consumos", fecha=fecha))
            creados, errores = 0, []
            for dni in dnis:
                trabajador = q_one("SELECT * FROM trabajadores WHERE dni=? AND activo=1", (dni,))
                if not trabajador:
                    errores.append(f"{dni}: DNI no encontrado o errado")
                    continue
                if not es_adicional and q_one("SELECT id FROM consumos WHERE fecha=? AND dni=? AND COALESCE(adicional,0)=0", (fecha, dni)):
                    errores.append(f"{dni}: ya tiene consumo registrado hoy")
                    continue
                try:
                    q_exec("""
                        INSERT INTO consumos(fecha,hora,dni,trabajador,empresa,area,tipo,cantidad,precio_unitario,total,observacion,comedor,fundo,responsable,adicional,estado,creado_por)
                        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """, (fecha, hora_now(), dni, trabajador["nombre"], trabajador["empresa"], trabajador["area"], tipo, cantidad, precio, total, obs, comedor, fundo, responsable, es_adicional, "PENDIENTE", session["user"]))
                    creados += 1
                except Exception as e:
                    errores.append(f"{dni}: no se pudo registrar")
            msg = f"REGISTRO DE CONSUMO terminado: {creados} consumo(s) registrado(s) para la fecha {fecha_peru_txt(fecha)}."
            if errores:
                msg += " Alertas: " + " | ".join(errores[:12])
                if len(errores) > 12:
                    msg += f" | y {len(errores)-12} más."
            flash(msg, "ok" if not errores else "error")
            return redirect(url_for("consumos", fecha=fecha))

        dni = clean_dni(request.form.get("dni"))
        trabajador = q_one("SELECT * FROM trabajadores WHERE dni=? AND activo=1", (dni,))
        if not trabajador:
            flash("DNI no encontrado o trabajador inactivo.", "error")
            return redirect(url_for("consumos"))

        # REGLA FUERTE: 1 DNI = 1 consumo normal por día.
        if not es_adicional:
            duplicado = q_one("SELECT id,hora,tipo FROM consumos WHERE fecha=? AND dni=? AND COALESCE(adicional,0)=0", (fecha, dni))
            if duplicado:
                flash(f"NO DUPLICADO: el DNI {dni} ya tiene consumo registrado hoy a las {duplicado['hora']}. Solo el admin puede registrar adicional.", "error")
                return redirect(url_for("consumos"))

        try:
            q_exec("""
                INSERT INTO consumos(fecha,hora,dni,trabajador,empresa,area,tipo,cantidad,precio_unitario,total,observacion,comedor,fundo,responsable,adicional,estado,creado_por)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (fecha, hora_now(), dni, trabajador["nombre"], trabajador["empresa"], trabajador["area"], tipo, cantidad, precio, total, obs, comedor, fundo, responsable, es_adicional, "PENDIENTE", session["user"]))
        except Exception:
            flash(f"NO DUPLICADO: el DNI {dni} ya tiene consumo registrado para el día {fecha_peru_txt(fecha)}.", "error")
            return redirect(url_for("consumos"))

        flash("REGISTRO DE CONSUMO realizado correctamente." + (" Marcado como adicional." if es_adicional else ""), "ok")
        return redirect(url_for("consumos"))

    fecha = request.args.get("fecha") or hoy_iso()
    fecha_inicio = request.args.get("fecha_inicio") or fecha
    fecha_fin = request.args.get("fecha_fin") or fecha_inicio
    buscar = clean_text(request.args.get("buscar"))
    cond, params = rango_sql(fecha_inicio, fecha_fin)
    where = cond
    final_params = list(params)
    if buscar:
        where += " AND (dni LIKE ? OR trabajador LIKE ? OR area LIKE ? OR fundo LIKE ? OR comedor LIKE ? OR responsable LIKE ? OR tipo LIKE ?)"
        b = f"%{buscar}%"
        final_params += [b, b, b, b, b, b, b]

    rows = q_all(f"SELECT * FROM consumos WHERE {where} ORDER BY fecha DESC,hora DESC,id DESC", tuple(final_params))
    tabla = "".join([
        f"""
        <tr class="fila-db-consumo">
          <td></td>
          <td>{r['fecha']}</td>
          <td>{r['hora']}</td>
          <td>{r['dni']}</td>
          <td>{r['trabajador']}</td>
          <td>{r['area']}</td>
          <td>{r['tipo']}{' + Adic.' if r['adicional'] else ''}</td>
          <td>{r['comedor']}</td>
          <td>{r['fundo']}</td>
          <td>{r['responsable'] or '-'}</td>
          <td>{r['cantidad']}</td>
          <td>{money(r['precio_unitario'])}</td>
          <td>{money(r['total'])}</td>
          <td><span class="badge {'ok' if r['estado']=='ENTREGADO' else 'warn'}">{r['estado']}</span></td>
          <td>
            <form method="post" action="{url_for('quitar_consumo')}" style="display:flex;gap:6px;align-items:center">
              <input type="hidden" name="id" value="{r['id']}">
              <input name="clave" placeholder="Clave" style="width:85px;padding:8px">
              <button class="btn-red" style="padding:8px 10px">Quitar</button>
            </form>
          </td>
        </tr>
        """ for r in rows
    ]) or "<tr id='fila_sin_registros'><td colspan='15'>Sin registros para este filtro.</td></tr>"

    fecha_cerrada = bool(dia_cerrado(fecha))
    fecha_es_hoy = (fecha == hoy_iso())
    disabled = "disabled" if (fecha_cerrada or not fecha_es_hoy) else ""
    bloqueado, msg_bloq = registro_bloqueado()
    aviso_bloq = f"<div class='flash error'>{msg_bloq}</div>" if bloqueado and session.get("role") != "admin" else ""
    aviso_fecha = ""
    if fecha_cerrada:
        aviso_fecha = "<div class='flash error'>Esta fecha está CERRADA. Puedes revisarla, pero no registrar nuevos consumos.</div>"
    elif not fecha_es_hoy:
        aviso_fecha = "<div class='flash error'>Fecha seleccionada solo para consulta. El registro de consumo solo está permitido en la fecha actual de hoy.</div>"

    filtros = filtro_bar(url_for("consumos"), fecha_inicio, fecha_fin, buscar)

    # Contador visible para registro masivo/lecturas de la fecha consultada.
    # Antes esta variable no existía dentro de /consumos y generaba error 500 al hacer clic en Consumos.
    total_consumos_fecha = int((q_one("SELECT COUNT(*) AS c FROM consumos WHERE fecha=?", (fecha,)) or {"c": 0})["c"] or 0)

    html = topbar("Registro y control de consumos", "Registra por digitación o lector QR usando el DNI") + f"""
    {aviso_bloq}
    {aviso_fecha}

    <div class="card">
      <h3 style="margin-top:0">Registrar consumo</h3>
      <div id="indicador_masivo_principal" style="margin:8px 0 12px;padding:14px 16px;border-radius:14px;border:2px solid #38bdf8;background:#e0f2fe;color:#075985;font-weight:950;display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap">
        <span>📦 Registro masivo automático activo: primero coloca RESPONSABLE; cada DNI válido se guardará al instante y aparecerá en CONSUMOS DE LA FECHA.</span>
        <span id="indicador_masivo_contador" style="background:#0d73b8;color:white;border-radius:999px;padding:7px 12px">0 en lote</span>
      </div>
      <div id="contador_lecturas_box" style="margin:8px 0 14px;padding:13px 14px;border-radius:16px;border:2px solid #16a34a;background:linear-gradient(135deg,#052e16,#064e3b);color:white;display:grid;grid-template-columns:auto 1fr auto;gap:12px;align-items:center;box-shadow:0 10px 24px rgba(22,163,74,.22)">
        <div style="font-size:25px">✅</div>
        <div>
          <div style="font-weight:950;font-size:15px">LECTURAS GUARDADAS HOY</div>
          <div style="font-size:12px;opacity:.88">Cada DNI válido suma aquí y se limpia para el siguiente.</div>
        </div>
        <div style="text-align:center;background:#22c55e;color:#052e16;border-radius:16px;padding:8px 14px;min-width:76px;font-weight:950">
          <div id="contador_lecturas_hoy" style="font-size:26px;line-height:1">{total_consumos_fecha}</div>
          <div style="font-size:10px">REG.</div>
        </div>
      </div>
      <form method="post" class="form-grid" id="form_consumo" onsubmit="return validarAntesEnviar(event)">
        <input type="date" name="fecha" value="{fecha}" onchange="window.location='{url_for('consumos')}?fecha=' + this.value" title="Elige una fecha para consultar. Solo hoy permite registrar." max="{hoy_iso()}">
        <input id="dni_consumo" name="dni" placeholder="Digite DNI o escanee QR/barras" required autofocus inputmode="numeric" pattern="[0-9]*" maxlength="8" autocomplete="off" enterkeyhint="next" oninput="dniInputHandler()" onkeyup="dniInputHandler()" onchange="dniInputHandler()" {disabled}>
        <input id="nombre_trabajador" class="worker-name-field" placeholder="Nombre aparecerá automáticamente al digitar DNI" readonly title="Nombre completo del trabajador" {disabled}>
        <button type="button" class="btn-blue" onclick="buscarTrabajadorConsumo(true)" {disabled}>🔎 Buscar trabajador</button>
        <button type="button" id="btn_qr" class="btn-blue" onclick="abrirScannerQR()" {disabled}>📷 Cámara QR / Barras</button>
        <div id="info_trabajador_consumo" style="display:none;grid-column:1/-1;border:1px solid #bbf7d0;background:#f0fdf4;border-radius:14px;padding:12px;font-weight:900;color:#14532d"></div>
        <div id="qr-reader" style="display:none;width:420px;max-width:100%;margin:10px 0;grid-column:1/-1"></div>
        <select name="comedor" {disabled}>
          {''.join([f'<option>{c}</option>' for c in opciones_comedor()])}
        </select>
        <select name="tipo" {disabled}>
          <option>Almuerzo</option>
          <option>Dieta</option>
        </select>
        <select name="fundo" {disabled}>
          {''.join([f'<option>{f}</option>' for f in opciones_fundo()])}
        </select>
        <input id="responsable_consumo" name="responsable" placeholder="RESPONSABLE (OBLIGATORIO MAYÚSCULAS)" required style="text-transform:uppercase" oninput="this.value=this.value.toUpperCase(); actualizarEstadoLoteResponsable();" {disabled}>
        <input type="number" name="cantidad" min="1" value="1" {disabled}>
        <input type="number" step="0.01" name="precio_unitario" value="10.00" {disabled}>
        <input name="observacion" placeholder="Observación / QR DNI" {disabled}>
        <label class="label-lote-final"><input type="checkbox" id="modo_lote" name="modo_lote" value="1" checked onchange="toggleLote()"> Registro masivo / lote</label>
        {('<label style="font-weight:900"><input type="checkbox" name="adicional" value="1"> Consumo adicional</label>' if session.get('role')=='admin' else '')}
        <div id="lote_panel" class="lote-dios-panel">
          <div class="lote-dios-head">
            <div>
              <div class="lote-dios-title">📦 REGISTRO MASIVO / LOTE EN VIVO</div>
              <div class="lote-dios-sub">La cámara queda encendida. Cada DNI detectado aparece aquí antes del clic final.</div>
            </div>
            <div class="lote-dios-counter"><b id="lote_total_big">0</b><span>TRABAJADORES EN LOTE</span></div>
          </div>
          <div class="lote-dios-status">
            <div>📷 Cámara: <span id="camara_estado_lote">apagada</span></div>
            <div>✅ Validados: <span id="lote_count">0</span></div>
            <div>🕒 Último DNI: <span id="ultimo_dni_lote">-</span></div>
          </div>
          <div class="lote-dios-list-head">
            <div>#</div><div>DNI</div><div>Trabajador detectado</div><div>Estado</div><div>Quitar</div>
          </div>
          <div id="lote_lista" class="lote-dios-list"></div>
          <div class="lote-dios-actions">
            <button type="button" class="btn-blue" style="min-height:36px;padding:8px 12px" onclick="agregarActualAlLote()">➕ Agregar DNI digitado</button>
            <button type="button" class="btn-red" style="min-height:36px;padding:8px 12px" onclick="limpiarLoteConsumos()">Limpiar lote</button>
          </div>
        </div>
        <textarea id="dni_lote" name="dni_lote" placeholder="DNIs validados para lote" style="display:none;grid-column:1/-1;min-height:90px"></textarea>
        <textarea id="lote_detalle" name="lote_detalle" style="display:none"></textarea>
        <textarea id="lote_checked" name="lote_checked" style="display:none"></textarea>
        <button id="btn_submit_consumo" {disabled}>REGISTRO DE CONSUMO</button>
        <a class="btn btn-blue" href="{url_for('consumos')}">Actualizar / refrescar</a>
      </form>
      <p class="muted small">Regla: no se permite duplicar DNI para el mismo día. Al digitar el DNI aparecerá automáticamente el nombre del trabajador.</p>
    </div>
    <script>
    let dniTimer = null;
    let ultimoDniValidado = '';
    let qrActivo = null;

    function getResponsableConsumo(){{
      const r = document.getElementById('responsable_consumo');
      return (r ? r.value : '').toString().trim().toUpperCase();
    }}
    function bloquearSiNoHayResponsable(mostrar=true){{
      const r = document.getElementById('responsable_consumo');
      const ok = !!getResponsableConsumo();
      if(!ok){{
        if(r){{ r.classList.add('responsable-alerta-final'); r.focus(); }}
        const out = document.getElementById('nombre_trabajador');
        if(out) out.value = '';
        const info = document.getElementById('info_trabajador_consumo');
        if(info){{ info.style.display='block'; info.innerHTML='<span style="color:#991b1b;font-weight:950">⚠️ Primero registra el RESPONSABLE. Sin responsable no se puede detectar DNI ni usar registro masivo.</span>'; }}
        if(mostrar) avisoMovil('Primero registra el RESPONSABLE.', false);
        return true;
      }}
      if(r) r.classList.remove('responsable-alerta-final');
      return false;
    }}
    function actualizarEstadoLoteResponsable(){{
      const r = document.getElementById('responsable_consumo');
      if(r && getResponsableConsumo()) r.classList.remove('responsable-alerta-final');
      renderPreviewLoteEnTabla(getLoteArray(), getLoteDetalle(), getLoteChecked());
    }}
    let scannerBusy = false;
    let ultimoScanDni = '';
    let ultimoScanTs = 0;

    function soloDni(v){{
      const raw = String(v || '').trim();
      if(!raw) return '';
      const only = raw.replace(/\D/g,'');
      if(only.length === 8) return only;
      const labeled = raw.toUpperCase().match(/(?:DNI|DOC(?:UMENTO)?|NRO|NÚMERO|NUMERO|DOCUMENT)\D{{0,16}}(\d{{8}})(?!\d)/);
      if(labeled) return labeled[1];
      const standalone = raw.match(/(^|\D)(\d{{8}})(?!\d)/);
      if(standalone) return standalone[2];
      if(only.length > 8) return only.slice(-8);
      return only.slice(0,8);
    }}
    function getLoteArray(){{
      const box = document.getElementById('dni_lote');
      if(!box) return [];
      return (box.value || '').split(/[\s,;|]+/).map(soloDni).filter(x => x.length === 8);
    }}
    function getLoteDetalle(){{
      const det = document.getElementById('lote_detalle');
      if(!det || !det.value) return {{}};
      try{{return JSON.parse(det.value || '{{}}') || {{}};}}catch(e){{return {{}};}}
    }}

    function fechaLocalKey(){{
      const d = new Date();
      const y = d.getFullYear();
      const m = String(d.getMonth()+1).padStart(2,'0');
      const day = String(d.getDate()).padStart(2,'0');
      return y + '-' + m + '-' + day;
    }}
    function getLoteChecked(){{
      const c = document.getElementById('lote_checked');
      if(!c || !c.value) return {{}};
      try{{return JSON.parse(c.value || '{{}}') || {{}};}}catch(e){{return {{}}}}
    }}
    function setLoteChecked(obj){{
      const c = document.getElementById('lote_checked');
      if(c) c.value = JSON.stringify(obj || {{}});
      try{{ localStorage.setItem('lote_consumos_checked_' + fechaLocalKey(), JSON.stringify(obj || {{}})); }}catch(e){{}}
    }}
    function getCheckedLoteArray(){{
      const arr = getLoteArray();
      const checked = getLoteChecked();
      return arr.filter(d => checked[d] !== false);
    }}
    function toggleCheckLote(dni, on){{
      dni = soloDni(dni);
      const checked = getLoteChecked();
      checked[dni] = !!on;
      setLoteChecked(checked);
      setLoteArray(getLoteArray(), getLoteDetalle());
    }}

    function setLoteDetalle(obj){{
      const det = document.getElementById('lote_detalle');
      if(det) det.value = JSON.stringify(obj || {{}});
      try{{ localStorage.setItem('lote_consumos_detalle_' + fechaLocalKey(), JSON.stringify(obj || {{}})); }}catch(e){{}}
    }}
    function setLoteArray(arr, detalle=null){{
      const limpio = [];
      arr.forEach(d => {{ d = soloDni(d); if(d && d.length === 8 && !limpio.includes(d)) limpio.push(d); }});
      const oldDetalle = detalle || getLoteDetalle();
      const nuevoDetalle = {{}};
      limpio.forEach(d => {{ nuevoDetalle[d] = oldDetalle[d] || ''; }});
      const oldChecked = getLoteChecked();
      const nuevoChecked = {{}};
      limpio.forEach(d => {{ nuevoChecked[d] = (oldChecked[d] === false) ? false : true; }});
      setLoteChecked(nuevoChecked);
      const box = document.getElementById('dni_lote');
      const lista = document.getElementById('lote_lista');
      const count = document.getElementById('lote_count');
      if(box) box.value = limpio.join('\n');
      setLoteDetalle(nuevoDetalle);
      if(count) count.textContent = getCheckedLoteArray().length + ' marcados / ' + limpio.length + ' detectados';
      const big = document.getElementById('lote_total_big');
      if(big) big.textContent = getCheckedLoteArray().length;
      const ultimo = document.getElementById('ultimo_dni_lote');
      if(ultimo) ultimo.textContent = limpio.length ? limpio[limpio.length-1] : '-';
      if(lista){{
        lista.innerHTML = limpio.length
          ? limpio.map((d, i) => {{
              const on = nuevoChecked[d] !== false;
              return `<div class="lote-dios-row ${{on ? '' : 'unchecked'}}"><b>${{i+1}}</b><b><input class="lote-check" type="checkbox" ${{on ? 'checked' : ''}} onchange="toggleCheckLote('${{d}}', this.checked)"> ${{d}}</b><span>${{(nuevoDetalle[d] || 'Trabajador validado')}}</span><span class="${{on ? 'ok' : ''}}">${{on ? 'MARCADO' : 'DESMARCADO'}}</span><button type="button" onclick="quitarDniLote('${{d}}')" style="min-height:0;width:38px;padding:6px;border-radius:999px;background:#ef4444;box-shadow:none">×</button></div>`;
            }}).join('')
          : '<div class="lote-dios-empty">Aún no hay DNIs guardados. Digita o escanea para acumular antes del clic final.</div>';
        renderPreviewLoteEnTabla(limpio, nuevoDetalle, nuevoChecked);
      }}
      try{{ localStorage.setItem('lote_consumos_' + fechaLocalKey(), limpio.join('\n')); }}catch(e){{}}
    }}

    function escHtml(v){{
      return String(v ?? '').replace(/[&<>"']/g, s => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[s]));
    }}
    function renderPreviewLoteEnTabla(arr, detalle, checked){{
      const tbody = document.getElementById('tbody_consumos_principal');
      if(!tbody) return;
      tbody.querySelectorAll('.fila-lote-preview').forEach(x => x.remove());
      const sin = document.getElementById('fila_sin_registros');
      if(sin) sin.style.display = arr.length ? 'none' : '';
      const form = document.getElementById('form_consumo');
      const fd = new FormData(form || document.createElement('form'));
      const hoy = document.querySelector('input[name="fecha"]')?.value || fechaLocalKey();
      const hora = new Date().toLocaleTimeString('es-PE', {{hour:'2-digit', minute:'2-digit', second:'2-digit', hour12:false}});
      const tipo = fd.get('tipo') || 'Almuerzo';
      const comedor = fd.get('comedor') || 'Comedor 01';
      const fundo = fd.get('fundo') || 'Kawsay Allpa';
      const responsable = (fd.get('responsable') || '').toString().toUpperCase();
      const cant = fd.get('cantidad') || '1';
      const precio = parseFloat(fd.get('precio_unitario') || '10') || 0;
      const total = (parseFloat(cant || '1') * precio).toFixed(2);
      const html = arr.map((d, i) => {{
        const on = checked[d] !== false;
        return `<tr class="fila-lote-preview ${{on ? '' : 'unchecked'}}">
          <td><input class="lote-check" type="checkbox" ${{on ? 'checked' : ''}} onchange="toggleCheckLote('${{d}}', this.checked)" title="Marcar/desmarcar antes del registro final"></td>
          <td>${{escHtml(hoy)}}</td><td>${{escHtml(hora)}}</td><td><b>${{escHtml(d)}}</b></td>
          <td>${{escHtml(detalle[d] || 'Trabajador validado')}}</td><td>-</td><td>${{escHtml(tipo)}}</td>
          <td>${{escHtml(comedor)}}</td><td>${{escHtml(fundo)}}</td><td>${{escHtml(responsable || '-')}}</td>
          <td>${{escHtml(cant)}}</td><td>S/ ${{precio.toFixed(2)}}</td><td>S/ ${{total}}</td>
          <td><span class="badge ${{on ? 'lote' : 'lote-off'}}">${{on ? 'PENDIENTE LOTE' : 'NO REGISTRAR'}}</span></td>
          <td><button type="button" onclick="quitarDniLote('${{d}}')" class="btn-red" style="min-height:0;padding:7px 10px">Quitar</button></td>
        </tr>`;
      }}).join('');
      tbody.insertAdjacentHTML('afterbegin', html);
    }}

    function quitarDniLote(dni){{
      const d = soloDni(dni);
      const detalle = getLoteDetalle();
      delete detalle[d];
      const checked = getLoteChecked();
      delete checked[d];
      setLoteChecked(checked);
      const arr = getLoteArray().filter(x => x !== d);
      setLoteArray(arr, detalle);
      avisoMovil('DNI quitado del lote: ' + dni, false);
      setTimeout(()=>document.getElementById('dni_consumo')?.focus(), 100);
    }}
    function limpiarLoteConsumos(){{
      setLoteChecked({{}});
      setLoteArray([], {{}});
      try{{ if(sessionStorage.getItem('limpiar_lote_tras_envio') === '1'){{ localStorage.removeItem('lote_consumos_' + fechaLocalKey()); localStorage.removeItem('lote_consumos_detalle_' + fechaLocalKey()); localStorage.removeItem('lote_consumos_checked_' + fechaLocalKey()); sessionStorage.removeItem('limpiar_lote_tras_envio'); }} }}catch(e){{}}
      const inp = document.getElementById('dni_consumo');
      const out = document.getElementById('nombre_trabajador');
      if(inp) inp.value='';
      if(out) out.value='';
      avisoMovil('Lote temporal limpiado.', false);
      setTimeout(()=>inp?.focus(), 100);
    }}
    function beepOk(){{
      try{{
        const AudioCtx = window.AudioContext || window.webkitAudioContext;
        const ctx = new AudioCtx();
        const osc = ctx.createOscillator();
        const gain = ctx.createGain();
        osc.connect(gain); gain.connect(ctx.destination);
        osc.frequency.value = 880; gain.gain.value = 0.07;
        osc.start(); setTimeout(()=>{{osc.stop(); ctx.close();}}, 140);
      }}catch(e){{}}
      if(navigator.vibrate) navigator.vibrate(90);
    }}
    function avisoMovil(msg, ok=true){{
      const div = document.createElement('div');
      div.className = 'prize-toast-msg';
      div.textContent = msg;
      const visibles = document.querySelectorAll('.prize-toast-msg').length;
      const topPx = 12 + (visibles * 58);
      div.style.position='fixed'; div.style.left='10px'; div.style.right='10px'; div.style.top='calc(env(safe-area-inset-top,0px) + '+topPx+'px)'; div.style.bottom='auto';
      div.style.zIndex='2147483647'; div.style.padding='12px 14px'; div.style.borderRadius='12px';
      div.style.fontWeight='950'; div.style.color='white'; div.style.textAlign='center'; div.style.boxShadow='0 12px 30px rgba(0,0,0,.35)'; div.style.pointerEvents='none';
      div.style.background = ok ? '#006b1e' : '#a40000'; div.style.border='1px solid rgba(255,255,255,.18)'; div.style.fontSize='13px'; div.style.lineHeight='1.2';
      document.body.appendChild(div); setTimeout(()=>div.remove(), 2300);
    }}
    async function validarDni(dni){{
      dni = soloDni(dni);
      if(dni.length !== 8) return {{ok:false, msg:'DNI incompleto'}};
      const r = await fetch('/api/trabajador/' + encodeURIComponent(dni), {{cache:'no-store'}});
      return await r.json();
    }}
    async function buscarTrabajadorConsumo(force=false){{
      if(bloquearSiNoHayResponsable(force)) return;
      const inp = document.getElementById('dni_consumo');
      const out = document.getElementById('nombre_trabajador');
      if(!inp || !out) return;
      const dni = soloDni(inp.value);
      if(inp.value !== dni) inp.value = dni;
      if(dni.length < 8){{ out.value=''; ultimoDniValidado=''; const info=document.getElementById('info_trabajador_consumo'); if(info){{info.style.display='none'; info.innerHTML='';}} return; }}
      if(!force && ultimoDniValidado === dni) return;
      ultimoDniValidado = dni;
      out.value = 'Validando DNI...';
      try{{
        const d = await validarDni(dni);
        if(d.ok){{
          out.value = d.nombre || '';
          out.title = d.nombre || '';
          const info = document.getElementById('info_trabajador_consumo');
          if(info){{
            info.style.display = 'block';
            info.innerHTML = '<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:10px"><div><b>Trabajador</b><br>' + (d.nombre || '-') + '</div><div><b>DNI</b><br>' + dni + '</div><div><b>Área</b><br>' + (d.area || '-') + '</div><div><b>Estado</b><br><span class="badge ok">Activo</span></div></div>';
          }}
          if(document.getElementById('modo_lote')?.checked){{
            setTimeout(()=>agregarDniLote(dni, d.nombre), 80);
          }}else{{
            beepOk();
          }}
        }}else{{
          out.value = 'DNI no encontrado';
          out.title = 'DNI no encontrado';
          const info = document.getElementById('info_trabajador_consumo');
          if(info){{ info.style.display='block'; info.innerHTML='<span style="color:#991b1b">DNI no encontrado en Trabajadores: ' + dni + '</span>'; }}
          if(document.getElementById('modo_lote')?.checked) avisoMovil('DNI no encontrado: ' + dni, false);
        }}
      }}catch(e){{ out.value='No se pudo validar DNI'; avisoMovil('Error validando DNI.', false); }}
    }}
    function dniInputHandler(){{
      const inp = document.getElementById('dni_consumo');
      if(inp) inp.value = soloDni(inp.value);
      if(inp && inp.value.length > 0 && bloquearSiNoHayResponsable(false)){{
        clearTimeout(dniTimer);
        const out = document.getElementById('nombre_trabajador'); if(out) out.value='';
        if(inp.value.length === 8) avisoMovil('Primero registra el RESPONSABLE antes de detectar DNI.', false);
        return;
      }}
      clearTimeout(dniTimer);
      const espera = (inp && inp.value.length === 8) ? 30 : 130;
      dniTimer = setTimeout(()=>buscarTrabajadorConsumo(false), espera);
    }}
    function agregarDniLote(dni, nombre){{
      if(bloquearSiNoHayResponsable(true)) return;
      dni = soloDni(dni);
      if(dni.length !== 8) return;
      const arr = getLoteArray();
      const detalle = getLoteDetalle();
      if(nombre) detalle[dni] = nombre;
      if(arr.includes(dni)){{
        setLoteArray(arr, detalle);
        avisoMovil('DNI ya estaba guardado en el lote: ' + dni, false);
      }}else{{
        arr.push(dni);
        setLoteArray(arr, detalle);
        beepOk();
        avisoMovil('DNI guardado y visualizado en lote: ' + dni + (nombre ? ' - ' + nombre : ''), true);
      }}
      const inp = document.getElementById('dni_consumo');
      const out = document.getElementById('nombre_trabajador');
      const info = document.getElementById('info_trabajador_consumo');
      if(inp) inp.value = '';
      if(out) out.value = '';
      if(info){{ info.style.display='block'; info.innerHTML='<b>Lote activo:</b> ' + getLoteArray().length + ' DNI(s) guardados y visibles en el cuadro temporal. Presiona <b>REGISTRO DE CONSUMO</b> para registrar todo.'; }}
      ultimoDniValidado = '';
      setTimeout(()=>inp?.focus(), 120);
    }}
    async function agregarActualAlLote(){{
      if(bloquearSiNoHayResponsable(true)) return;
      const inp = document.getElementById('dni_consumo');
      const dni = soloDni(inp ? inp.value : '');
      if(dni.length !== 8){{ avisoMovil('Digite o escanee un DNI válido de 8 dígitos.', false); return; }}
      try{{
        const d = await validarDni(dni);
        if(d.ok) agregarDniLote(dni, d.nombre || 'Trabajador validado');
        else avisoMovil('DNI no encontrado: ' + dni, false);
      }}catch(e){{ avisoMovil('No se pudo validar el DNI.', false); }}
    }}
    function toggleLote(){{
      let on = document.getElementById('modo_lote')?.checked;
      if(on && bloquearSiNoHayResponsable(true)){{
        const chk = document.getElementById('modo_lote'); if(chk) chk.checked = false;
        on = false;
      }}
      const box = document.getElementById('dni_lote');
      const panel = document.getElementById('lote_panel');
      const dni = document.getElementById('dni_consumo');
      if(box) box.style.display = 'none';
      if(panel) panel.style.display = on ? 'block' : 'none';
      if(dni) dni.required = !on;
      setLoteArray(getLoteArray());
      // Al activar el check después de digitar un DNI válido, lo pasamos al espacio temporal.
      if(on){{
        const actual = soloDni(dni ? dni.value : '');
        const nombreActual = document.getElementById('nombre_trabajador')?.value || '';
        if(actual.length === 8 && !/no encontrado|validando|error/i.test(nombreActual)){{
          setTimeout(()=>agregarDniLote(actual, nombreActual), 60);
        }}
      }}
      const btn = document.getElementById('btn_submit_consumo');
      if(btn) btn.textContent = 'REGISTRO DE CONSUMO';
      if(on) avisoMovil('Registro masivo activado. Los DNI se guardarán en lote temporal hasta presionar REGISTRO DE CONSUMO.', true);
    }}
    async function procesarDniQR(texto){{
      if(scannerBusy) return;
      if(bloquearSiNoHayResponsable(true)) return;
      const dni = soloDni(texto);
      if(dni.length !== 8){{ avisoMovil('QR/barras inválido: no contiene DNI de 8 dígitos.', false); return; }}
      const ahoraScan = Date.now();
      if(document.getElementById('modo_lote')?.checked && dni === ultimoScanDni && (ahoraScan - ultimoScanTs) < 2500) return;
      ultimoScanDni = dni; ultimoScanTs = ahoraScan;
      scannerBusy = true;
      const inp = document.getElementById('dni_consumo');
      const out = document.getElementById('nombre_trabajador');
      if(inp) inp.value = dni;
      try{{
        const d = await validarDni(dni);
        if(d.ok){{
          if(out) out.value = d.nombre || '';
          const info = document.getElementById('info_trabajador_consumo');
          if(info){{ info.style.display='block'; info.innerHTML='<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:10px"><div><b>Trabajador</b><br>' + (d.nombre || '-') + '</div><div><b>DNI</b><br>' + dni + '</div><div><b>Área</b><br>' + (d.area || '-') + '</div><div><b>Estado</b><br><span class="badge ok">Activo</span></div></div>'; }}
          ultimoDniValidado = dni;
          if(document.getElementById('modo_lote')?.checked){{ agregarDniLote(dni, d.nombre); }}
          else {{ beepOk(); avisoMovil('DNI reconocido: ' + (d.nombre || dni), true); }}
        }}else{{
          if(out) out.value = 'DNI no encontrado';
          avisoMovil('DNI no encontrado: ' + dni, false);
        }}
      }}catch(e){{ avisoMovil('No se pudo validar el DNI.', false); }}
      setTimeout(()=>{{ scannerBusy=false; }}, document.getElementById('modo_lote')?.checked ? 350 : 900);
    }}
    async function abrirScannerQR(){{
      if(bloquearSiNoHayResponsable(true)) return;
      const cont = document.getElementById('qr-reader');
      if(!cont) return;
      if(location.protocol !== 'https:' && location.hostname !== 'localhost' && location.hostname !== '127.0.0.1'){{
        avisoMovil('La cámara necesita HTTPS. Abre el enlace de Render con https://', false);
      }}
      cont.style.display='block';
      cont.innerHTML = `<div class="qr-camera-box">
        <b>Escáner con cámara activo</b><br>
        <div id="qr-reader-live" class="qr-live-box"></div>
        <video id="qr-video-live" class="qr-video-box" playsinline muted autoplay style="display:none"></video>
        <canvas id="qr-canvas-live" style="display:none"></canvas>
        <div class="qr-actions">
          <button type="button" class="btn-red qr-close-btn" onclick="cerrarScannerQR()">Cerrar cámara</button>
        </div>
        <small class="muted">Permite la cámara. En celular usa Chrome y el enlace HTTPS de Render.</small>
      </div>`;
      try{{
        if(window.Html5Qrcode){{
          const formatos = window.Html5QrcodeSupportedFormats ? [
            Html5QrcodeSupportedFormats.QR_CODE,
            Html5QrcodeSupportedFormats.CODE_128,
            Html5QrcodeSupportedFormats.CODE_39,
            Html5QrcodeSupportedFormats.EAN_13,
            Html5QrcodeSupportedFormats.EAN_8,
            Html5QrcodeSupportedFormats.ITF,
            Html5QrcodeSupportedFormats.UPC_A,
            Html5QrcodeSupportedFormats.UPC_E,
            Html5QrcodeSupportedFormats.PDF_417
          ].filter(Boolean) : undefined;
          qrActivo = new Html5Qrcode('qr-reader-live', formatos ? {{ formatsToSupport: formatos, verbose: false }} : undefined);
          await qrActivo.start(
            {{ facingMode: {{ ideal: 'environment' }} }},
            {{ fps: 15, qrbox: {{ width: 220, height: 220 }}, rememberLastUsedCamera: true }},
            async (decodedText) => {{ await procesarDniQR(decodedText); if(!document.getElementById('modo_lote')?.checked){{ cerrarScannerQR(); }} }},
            () => {{}}
          );
          const ce = document.getElementById('camara_estado_lote'); if(ce) ce.innerHTML = '<span class="cam-on">● encendida continua</span>';
          avisoMovil('Cámara activada. En modo masivo NO se apaga al detectar.', true);
          return;
        }}
      }}catch(e){{ console.warn('Html5Qrcode falló, usando respaldo:', e); }}
      try{{ await iniciarScannerNativo(); }}
      catch(e2){{ alert('No se pudo abrir la cámara. Usa HTTPS de Render, acepta permisos y prueba Chrome/Edge. Detalle: ' + (e2 && e2.message ? e2.message : e2)); }}
    }}
    async function iniciarScannerNativo(){{
      if(!navigator.mediaDevices || !navigator.mediaDevices.getUserMedia) throw new Error('El navegador no permite cámara.');
      const video = document.getElementById('qr-video-live');
      const canvas = document.getElementById('qr-canvas-live');
      const live = document.getElementById('qr-reader-live');
      if(live) live.innerHTML = '<b>Usando cámara directa...</b><br><small>Detecta QR con jsQR y barras con BarcodeDetector si el navegador lo soporta.</small>';
      const stream = await navigator.mediaDevices.getUserMedia({{video: {{facingMode: {{ideal:'environment'}}}}, audio:false}});
      qrActivo = {{stream: stream, stopped:false}};
      video.srcObject = stream; video.style.display='block';
      await video.play();
      let detector = null;
      if('BarcodeDetector' in window){{
        try{{ detector = new BarcodeDetector({{formats:['qr_code','code_128','code_39','ean_13','ean_8','itf','codabar','upc_a','upc_e','pdf417']}}); }}catch(e){{}}
      }}
      const ce = document.getElementById('camara_estado_lote'); if(ce) ce.innerHTML = '<span class="cam-on">● encendida continua</span>';
      avisoMovil('Cámara activada.', true);
      const loop = async () => {{
        if(!qrActivo || qrActivo.stopped) return;
        try{{
          if(detector){{
            const codes = await detector.detect(video);
            if(codes && codes.length){{
              await procesarDniQR(codes[0].rawValue || '');
              if(!document.getElementById('modo_lote')?.checked){{ cerrarScannerQR(); return; }}
            }}
          }}
          if(window.jsQR && video.videoWidth > 0){{
            canvas.width = video.videoWidth; canvas.height = video.videoHeight;
            const ctx = canvas.getContext('2d', {{willReadFrequently:true}});
            ctx.drawImage(video, 0, 0, canvas.width, canvas.height);
            const imageData = ctx.getImageData(0, 0, canvas.width, canvas.height);
            const code = jsQR(imageData.data, imageData.width, imageData.height);
            if(code && code.data){{
              await procesarDniQR(code.data);
              if(!document.getElementById('modo_lote')?.checked){{ cerrarScannerQR(); return; }}
            }}
          }}
        }}catch(e){{}}
        requestAnimationFrame(loop);
      }};
      requestAnimationFrame(loop);
    }}
    function cerrarScannerQR(){{
      try{{
        if(qrActivo){{
          if(typeof qrActivo.stop === 'function'){{
            qrActivo.stop().catch(()=>{{}}).finally(()=>{{ try{{ qrActivo.clear(); }}catch(e){{}} }});
          }}
          if(qrActivo.stream){{
            qrActivo.stopped = true;
            qrActivo.stream.getTracks().forEach(t => t.stop());
          }}
        }}
      }}catch(e){{}}
      qrActivo = null;
      const cont = document.getElementById('qr-reader');
      const ce = document.getElementById('camara_estado_lote'); if(ce) ce.textContent = 'apagada';
      if(cont){{ cont.style.display='none'; cont.innerHTML=''; }}
    }}
    function validarAntesEnviar(e){{
      const lote = document.getElementById('modo_lote')?.checked;
      if(lote){{
        let arr = getCheckedLoteArray();
        const actual = soloDni(document.getElementById('dni_consumo')?.value || '');
        const nombreActual = document.getElementById('nombre_trabajador')?.value || '';
        if(arr.length === 0 && actual.length === 8 && !/no encontrado|validando|error/i.test(nombreActual)){{
          arr = [actual];
          setLoteArray(arr);
        }}
        if(arr.length === 0){{ e.preventDefault(); avisoMovil('No hay DNI válidos guardados para el registro masivo.', false); return false; }}
        document.getElementById('dni_lote').value = getCheckedLoteArray().join('\n');
        if(!confirm('Se registrarán ' + getCheckedLoteArray().length + ' consumo(s) marcados para la fecha de hoy. ¿Confirmas REGISTRO DE CONSUMO?')){{ e.preventDefault(); return false; }}
      }}
      try{{ sessionStorage.setItem('limpiar_lote_tras_envio', '1'); }}catch(ex){{}}
      return true;
    }}
    document.addEventListener('DOMContentLoaded', ()=>{{
      const inp = document.getElementById('dni_consumo');
      if(inp){{
        inp.addEventListener('paste', ()=>setTimeout(dniInputHandler, 40));
        inp.addEventListener('input', dniInputHandler);
        inp.addEventListener('keyup', dniInputHandler);
        inp.addEventListener('change', dniInputHandler);
        inp.addEventListener('keydown', (e)=>{{ if(e.key === 'Enter'){{ e.preventDefault(); buscarTrabajadorConsumo(true); }}}});
        setTimeout(()=>inp.focus(), 300);
      }}
      const form = document.getElementById('form_consumo');
      if(form) form.addEventListener('submit', validarAntesEnviar);
      try{{
        const key = 'lote_consumos_' + fechaLocalKey();
        const guardado = localStorage.getItem(key);
        if(guardado && document.getElementById('dni_lote')) document.getElementById('dni_lote').value = guardado;
        const detGuardado = localStorage.getItem('lote_consumos_detalle_' + fechaLocalKey());
        if(detGuardado && document.getElementById('lote_detalle')) document.getElementById('lote_detalle').value = detGuardado;
        const chkGuardado = localStorage.getItem('lote_consumos_checked_' + fechaLocalKey());
        if(chkGuardado && document.getElementById('lote_checked')) document.getElementById('lote_checked').value = chkGuardado;
      }}catch(e){{}}
      toggleLote();
      setLoteArray(getLoteArray());
    }});
    </script>


    <script>
    // ===== FIX DEFINITIVO MASIVO: responsable obligatorio + lote visible + guardado real =====
    (function(){{
      const LS_KEY = 'PRIZE_LOTE_MASIVO_' + (document.querySelector('input[name="fecha"]')?.value || new Date().toISOString().slice(0,10));
      let loteMasivoFix = [];
      function onlyDni(v){{
        const raw = String(v || '').trim();
        const digits = raw.replace(/\D/g,'');
        if(digits.length === 8) return digits;
        const m = raw.toUpperCase().match(/(?:DNI|DOC(?:UMENTO)?|NRO|NUM(?:ERO)?|NÚMERO)\D{{0,20}}(\d{{8}})(?!\d)/);
        if(m) return m[1];
        const e = raw.match(/(^|\D)(\d{{8}})(?!\d)/);
        if(e) return e[2];
        if(digits.length > 8) return digits.slice(-8);
        return digits.slice(0,8);
      }}
      window.soloDni = onlyDni;
      function toastFix(msg, ok=true){{ try{{ avisoMovil(msg, ok); return; }}catch(e){{}} alert(msg); }}
      function responsableFix(){{ const r = document.querySelector('#form_consumo [name="responsable"]'); return String(r ? r.value : '').trim().toUpperCase(); }}
      function validarResponsableFix(){{
        const r = document.querySelector('#form_consumo [name="responsable"]');
        const val = responsableFix();
        if(r) r.value = val;
        if(!val){{
          if(r){{ r.focus(); r.style.borderColor='#ef4444'; r.style.boxShadow='0 0 0 4px rgba(239,68,68,.16)'; }}
          toastFix('Primero coloca el RESPONSABLE. Sin responsable no se permite detectar ni guardar DNI.', false);
          return false;
        }}
        if(r){{ r.style.borderColor=''; r.style.boxShadow=''; }}
        return true;
      }}
      window.validarResponsableFix = validarResponsableFix;

      function bloquearControlesPorResponsable(){{
        const has = !!responsableFix();
        const inp = document.getElementById('dni_consumo');
        const b1 = document.querySelector('button[onclick="buscarTrabajadorConsumo(true)"]');
        const b2 = document.getElementById('btn_qr');
        [inp,b1,b2].forEach(el => {{ if(el){{ el.disabled = !has; el.style.opacity = has ? '1' : '.55'; el.style.cursor = has ? '' : 'not-allowed'; }} }});
        const principal = document.getElementById('indicador_masivo_principal');
        if(principal){{
          principal.style.borderColor = has ? '#38bdf8' : '#ef4444';
          principal.style.background = has ? '#e0f2fe' : '#fff1f2';
          principal.style.color = has ? '#075985' : '#991b1b';
          const span = principal.querySelector('span');
          if(span) span.textContent = has ? '📦 Registro masivo automático activo: digita o escanea DNI. Cada trabajador aparecerá abajo antes de guardar.' : '⚠️ Primero coloca RESPONSABLE. El DNI, búsqueda y cámara están bloqueados hasta completar responsable.';
        }}
      }}
      function loadLoteFix(){{
        try{{ loteMasivoFix = JSON.parse(localStorage.getItem(LS_KEY) || '[]') || []; }}catch(e){{ loteMasivoFix = []; }}
        loteMasivoFix = loteMasivoFix.filter(x => onlyDni(x.dni).length === 8).map(x => ({{dni:onlyDni(x.dni), nombre:x.nombre||'', area:x.area||'', checked:x.checked !== false}}));
      }}
      function saveLoteFix(){{
        try{{ localStorage.setItem(LS_KEY, JSON.stringify(loteMasivoFix)); }}catch(e){{}}
        const hidden = document.getElementById('dni_lote');
        if(hidden) hidden.value = loteMasivoFix.filter(x => x.checked !== false).map(x => x.dni).join('\n');
        const det = document.getElementById('lote_detalle');
        if(det){{ const obj = {{}}; loteMasivoFix.forEach(x => obj[x.dni] = x.nombre || 'Trabajador validado'); det.value = JSON.stringify(obj); }}
        const chk = document.getElementById('lote_checked');
        if(chk){{ const obj = {{}}; loteMasivoFix.forEach(x => obj[x.dni] = x.checked !== false); chk.value = JSON.stringify(obj); }}
      }}
      function checkedCountFix(){{ return loteMasivoFix.filter(x => x.checked !== false).length; }}
      window.getLoteArray = function(){{ return loteMasivoFix.map(x => x.dni); }};
      window.getCheckedLoteArray = function(){{ return loteMasivoFix.filter(x => x.checked !== false).map(x => x.dni); }};
      window.getLoteDetalle = function(){{ const o={{}}; loteMasivoFix.forEach(x => o[x.dni]=x.nombre||'Trabajador validado'); return o; }};
      window.getLoteChecked = function(){{ const o={{}}; loteMasivoFix.forEach(x => o[x.dni]=x.checked!==false); return o; }};
      window.setLoteArray = function(arr, detalle){{
        const det = detalle || {{}}; const nuevo = [];
        (arr || []).forEach(d => {{ d = onlyDni(d); if(d.length === 8 && !nuevo.find(x => x.dni === d)){{ const old = loteMasivoFix.find(x => x.dni === d) || {{}}; nuevo.push({{dni:d, nombre:det[d] || old.nombre || 'Trabajador validado', area:old.area || '', checked:old.checked !== false}}); }} }});
        loteMasivoFix = nuevo; renderLoteFix();
      }};
      window.setLoteChecked = function(obj){{ loteMasivoFix.forEach(x => {{ if(obj && Object.prototype.hasOwnProperty.call(obj, x.dni)) x.checked = !!obj[x.dni]; }}); renderLoteFix(); }};
      function ensureIndicatorFix(){{
        let ind = document.getElementById('indicador_guardado_masivo'); const btn = document.getElementById('btn_submit_consumo');
        if(!ind && btn){{ ind = document.createElement('div'); ind.id='indicador_guardado_masivo'; ind.style.cssText='display:none;grid-column:1/-1;margin:8px 0;padding:12px;border-radius:14px;background:#e0f2fe;border:1px solid #38bdf8;color:#075985;font-weight:950;text-align:center'; btn.parentNode.insertBefore(ind, btn); }}
        return ind;
      }}
      function escHtmlFix(v){{ return String(v ?? '').replace(/[&<>"']/g, s => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[s])); }}
      function renderLoteFix(){{
        saveLoteFix();
        const panel = document.getElementById('lote_panel'); const on = document.getElementById('modo_lote')?.checked;
        if(panel) panel.style.display = on ? 'block' : 'none';
        const lista = document.getElementById('lote_lista'); const count = document.getElementById('lote_count'); const big = document.getElementById('lote_total_big'); const ultimo = document.getElementById('ultimo_dni_lote');
        if(count) count.textContent = checkedCountFix() + ' marcados / ' + loteMasivoFix.length + ' detectados';
        if(big) big.textContent = checkedCountFix();
        const contadorPrincipal = document.getElementById('indicador_masivo_contador');
        if(contadorPrincipal) contadorPrincipal.textContent = checkedCountFix() + ' en lote';
        try{{ bloquearControlesPorResponsable(); }}catch(e){{}}
        if(ultimo) ultimo.textContent = loteMasivoFix.length ? loteMasivoFix[loteMasivoFix.length-1].dni : '-';
        if(lista){{ lista.innerHTML = loteMasivoFix.length ? loteMasivoFix.map((x,i)=>`<div class="lote-dios-row ${{x.checked !== false ? '' : 'unchecked'}}"><b>${{i+1}}</b><b><input class="lote-check" type="checkbox" ${{x.checked !== false ? 'checked' : ''}} onchange="toggleCheckLote('${{x.dni}}', this.checked)"> ${{x.dni}}</b><span>${{escHtmlFix(x.nombre || 'Trabajador validado')}}</span><span class="${{x.checked !== false ? 'ok' : ''}}">${{x.checked !== false ? 'MARCADO' : 'DESMARCADO'}}</span><button type="button" onclick="quitarDniLote('${{x.dni}}')" style="min-height:0;width:38px;padding:6px;border-radius:999px;background:#ef4444;box-shadow:none">×</button></div>`).join('') : '<div class="lote-dios-empty">Aún no hay DNIs guardados. Coloca responsable y luego digita/escanea.</div>'; }}
        renderTablaPreviewFix();
      }}
      function renderTablaPreviewFix(){{
        const tbody = document.getElementById('tbody_consumos_principal'); if(!tbody) return;
        tbody.querySelectorAll('.fila-lote-preview').forEach(e => e.remove());
        const sin = document.getElementById('fila_sin_registros'); if(sin) sin.style.display = loteMasivoFix.length ? 'none' : '';
        const form = document.getElementById('form_consumo'); const fd = new FormData(form || document.createElement('form'));
        const fecha = document.querySelector('input[name="fecha"]')?.value || new Date().toISOString().slice(0,10);
        const hora = new Date().toLocaleTimeString('es-PE',{{hour:'2-digit',minute:'2-digit',second:'2-digit',hour12:false}});
        const tipo = fd.get('tipo') || 'Almuerzo'; const comedor = fd.get('comedor') || 'Comedor 01'; const fundo = fd.get('fundo') || 'Kawsay Allpa'; const responsable = responsableFix() || '-'; const cant = fd.get('cantidad') || '1'; const precio = parseFloat(fd.get('precio_unitario') || '10') || 0; const total = ((parseFloat(cant || '1') || 1) * precio).toFixed(2);
        const html = loteMasivoFix.map((x,i)=>`<tr class="fila-lote-preview ${{x.checked !== false ? '' : 'unchecked'}}"><td><input class="lote-check" type="checkbox" ${{x.checked !== false ? 'checked' : ''}} onchange="toggleCheckLote('${{x.dni}}', this.checked)"></td><td>${{escHtmlFix(fecha)}}</td><td>${{escHtmlFix(hora)}}</td><td><b>${{escHtmlFix(x.dni)}}</b></td><td>${{escHtmlFix(x.nombre || 'Trabajador validado')}}</td><td>${{escHtmlFix(x.area || '-')}}</td><td>${{escHtmlFix(tipo)}}</td><td>${{escHtmlFix(comedor)}}</td><td>${{escHtmlFix(fundo)}}</td><td>${{escHtmlFix(responsable)}}</td><td>${{escHtmlFix(cant)}}</td><td>S/ ${{precio.toFixed(2)}}</td><td>S/ ${{total}}</td><td><span class="badge ${{x.checked !== false ? 'lote' : 'lote-off'}}">${{x.checked !== false ? 'LISTO PARA GUARDAR' : 'NO GUARDAR'}}</span></td><td><button type="button" onclick="quitarDniLote('${{x.dni}}')" class="btn-red" style="min-height:0;padding:7px 10px">Quitar</button></td></tr>`).join('');
        tbody.insertAdjacentHTML('afterbegin', html);
      }}
      window.toggleCheckLote = function(dni,on){{ dni = onlyDni(dni); const item = loteMasivoFix.find(x => x.dni === dni); if(item) item.checked = !!on; renderLoteFix(); }};
      window.quitarDniLote = function(dni){{ dni = onlyDni(dni); loteMasivoFix = loteMasivoFix.filter(x => x.dni !== dni); renderLoteFix(); toastFix('DNI quitado del lote: ' + dni, false); }};
      window.limpiarLoteConsumos = function(){{ loteMasivoFix = []; try{{ localStorage.removeItem(LS_KEY); }}catch(e){{}} renderLoteFix(); toastFix('Lote temporal limpiado.', false); }};
      window.agregarDniLote = function(dni,nombre,area){{
        if(!validarResponsableFix()) return false;
        dni = onlyDni(dni); if(dni.length !== 8){{ toastFix('DNI inválido.', false); return false; }}
        let item = loteMasivoFix.find(x => x.dni === dni);
        if(item){{ item.checked = true; item.nombre = nombre || item.nombre; item.area = area || item.area; toastFix('DNI ya estaba en lote, se mantiene marcado: ' + dni, false); }}
        else{{ loteMasivoFix.push({{dni, nombre:nombre || 'Trabajador validado', area:area || '', checked:true}}); toastFix('Guardado temporal en lote: ' + dni + (nombre ? ' - ' + nombre : ''), true); }}
        const indAdd = ensureIndicatorFix(); if(indAdd){{ indAdd.style.display='block'; indAdd.textContent='✅ Guardado temporal en registro masivo: ' + checkedCountFix() + ' marcado(s) / ' + loteMasivoFix.length + ' detectado(s).'; }}
        renderLoteFix(); const inp = document.getElementById('dni_consumo'); const out = document.getElementById('nombre_trabajador'); if(inp) inp.value=''; if(out) out.value=''; setTimeout(()=>inp?.focus(),100); return true;
      }};
      async function validarDniFix(dni){{ const r = await fetch('/api/trabajador/' + encodeURIComponent(dni) + '?_=' + Date.now(), {{cache:'no-store', credentials:'same-origin'}}); return await r.json(); }}
      window.buscarTrabajadorConsumo = async function(force=false){{
        const enLote = document.getElementById('modo_lote')?.checked;
        if(!validarResponsableFix()){{
          const inp0 = document.getElementById('dni_consumo');
          const out0 = document.getElementById('nombre_trabajador');
          const info0 = document.getElementById('info_trabajador_consumo');
          if(inp0) inp0.value='';
          if(out0) out0.value='';
          if(info0){{ info0.style.display='none'; info0.innerHTML=''; }}
          return;
        }}
        const inp = document.getElementById('dni_consumo'); const out = document.getElementById('nombre_trabajador'); if(!inp || !out) return;
        const dni = onlyDni(inp.value); inp.value = dni; if(dni.length < 8){{ out.value=''; return; }}
        out.value='Validando DNI...';
        try{{ const d = await validarDniFix(dni); if(d && d.ok){{ out.value = d.nombre || ''; const info = document.getElementById('info_trabajador_consumo'); if(info){{ info.style.display='block'; info.innerHTML='<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:10px"><div><b>Trabajador</b><br>'+(d.nombre||'-')+'</div><div><b>DNI</b><br>'+dni+'</div><div><b>Área</b><br>'+(d.area||'-')+'</div><div><b>Estado</b><br><span class="badge ok">Activo</span></div></div>'; }} if(enLote) window.agregarDniLote(dni, d.nombre || '', d.area || ''); else {{ try{{ beepOk(); }}catch(e){{}} }} }} else {{ out.value='DNI no encontrado'; toastFix('DNI no encontrado: '+dni, false); }} }}catch(e){{ out.value='Error validando DNI'; toastFix('Error consultando trabajador.', false); }}
      }};
      window.dniInputHandler = function(){{
        const inp = document.getElementById('dni_consumo'); if(!inp) return;
        inp.value = onlyDni(inp.value);
        clearTimeout(window.__fixDniTimer);
        if(inp.value.length > 0 && !responsableFix()){{
          const out = document.getElementById('nombre_trabajador');
          if(out) out.value='PRIMERO COLOCA RESPONSABLE';
          inp.value='';
          validarResponsableFix();
          return;
        }}
        if(inp.value.length === 8){{ window.__fixDniTimer = setTimeout(()=>window.buscarTrabajadorConsumo(false), 60); }}
      }};
      window.procesarDniQR = async function(texto){{ if(!validarResponsableFix()) return; const dni = onlyDni(texto); if(dni.length !== 8){{ toastFix('QR/barras inválido: no contiene DNI de 8 dígitos.', false); return; }} const inp = document.getElementById('dni_consumo'); if(inp) inp.value=dni; await window.buscarTrabajadorConsumo(true); }};
      const oldAbrir = window.abrirScannerQR; window.abrirScannerQR = function(){{ if(!validarResponsableFix()) return false; return oldAbrir ? oldAbrir() : false; }};
      window.agregarActualAlLote = async function(){{ if(!validarResponsableFix()) return; await window.buscarTrabajadorConsumo(true); }};
      window.validarAntesEnviar = function(e){{
        const enLote = document.getElementById('modo_lote')?.checked; if(!validarResponsableFix()){{ if(e) e.preventDefault(); return false; }}
        if(enLote){{ const marcados = loteMasivoFix.filter(x => x.checked !== false).map(x => x.dni); if(marcados.length === 0){{ if(e) e.preventDefault(); toastFix('No hay trabajadores marcados para guardar. Escanea/digita y deja check marcado.', false); return false; }} if(!confirm('Se guardarán ' + marcados.length + ' trabajador(es) marcados. ¿Confirmas REGISTRO DE CONSUMO?')){{ if(e) e.preventDefault(); return false; }} const hidden = document.getElementById('dni_lote'); if(hidden) hidden.value = marcados.join('\n'); try{{ sessionStorage.setItem('limpiar_lote_tras_envio_fix', '1'); }}catch(ex){{}} const ind = ensureIndicatorFix(); if(ind){{ ind.style.display='block'; ind.textContent='⏳ Guardando registro masivo: 0 de ' + marcados.length + '...'; }} const btn = document.getElementById('btn_submit_consumo'); if(btn){{ btn.disabled=true; btn.textContent='GUARDANDO MASIVO...'; }} setTimeout(()=>{{ if(ind) ind.textContent='⏳ Enviando y grabando ' + marcados.length + ' trabajador(es) en base de datos...'; }},150); return true; }}
        return true;
      }};

      // ===== AUTO-GUARDADO MASIVO REAL: guarda al detectar DNI válido =====
      let autoGuardandoFix = false;
      let autoGuardadosFix = 0;
      function ensureAutoPanelFix(){{
        let p = document.getElementById('auto_guardado_panel');
        const form = document.getElementById('form_consumo');
        if(!p && form){{
          p = document.createElement('div');
          p.id = 'auto_guardado_panel';
          p.innerHTML = '<div>✅ Registros automáticos guardados: <span id="auto_guardado_count">0</span></div><div class="mini">Cada DNI válido se guarda en CONSUMOS DE LA FECHA y el campo DNI queda limpio para el siguiente.</div>';
          const info = document.getElementById('info_trabajador_consumo');
          if(info && info.parentNode) info.parentNode.insertBefore(p, info.nextSibling);
          else form.insertBefore(p, form.firstChild);
        }}
        return p;
      }}
      function limpiarDniParaSiguienteFix(){{
        const inp = document.getElementById('dni_consumo');
        const out = document.getElementById('nombre_trabajador');
        const info = document.getElementById('info_trabajador_consumo');
        if(inp) inp.value = '';
        if(out) out.value = '';
        if(info){{ info.style.display='none'; info.innerHTML=''; }}
        setTimeout(()=>inp?.focus(), 80);
      }}
      function prependConsumoGuardadoFix(rowHtml){{
        const tbody = document.getElementById('tbody_consumos_principal');
        if(!tbody) return;
        const sin = document.getElementById('fila_sin_registros');
        if(sin) sin.remove();
        tbody.insertAdjacentHTML('afterbegin', rowHtml);
      }}
      async function registrarConsumoAutomaticoFix(dni){{
        if(autoGuardandoFix) return;
        if(!validarResponsableFix()){{ limpiarDniParaSiguienteFix(); return; }}
        dni = onlyDni(dni);
        if(dni.length !== 8) return;
        autoGuardandoFix = true;
        const ind = ensureIndicatorFix();
        if(ind){{ ind.style.display='block'; ind.textContent='⏳ Guardando automáticamente DNI ' + dni + '...'; }}
        try{{
          const form = document.getElementById('form_consumo');
          const fd = new FormData(form || document.createElement('form'));
          fd.set('dni', dni);
          fd.set('modo_lote', '0');
          const r = await fetch('/api/registrar_consumo_auto', {{method:'POST', body:fd, credentials:'same-origin', cache:'no-store'}});
          const data = await r.json().catch(()=>({{ok:false,msg:'Error leyendo respuesta del servidor'}}));
          if(data.ok){{
            prependConsumoGuardadoFix(data.row_html || '');
            autoGuardadosFix += 1;
            const p = ensureAutoPanelFix();
            const c = document.getElementById('auto_guardado_count');
            if(c) c.textContent = autoGuardadosFix;
            if(p) p.style.display = 'block';
            if(ind){{ ind.style.display='block'; ind.textContent = data.msg || ('✅ Guardado automático: ' + dni); }}
            try{{ beepOk(); }}catch(e){{}}
            toastFix(data.msg || ('Guardado automático: ' + dni), true);
            limpiarDniParaSiguienteFix();
            try{{ loteMasivoFix = []; localStorage.removeItem(LS_KEY); renderLoteFix(); }}catch(e){{}}
          }}else{{
            if(ind){{ ind.style.display='block'; ind.textContent = '❌ ' + (data.msg || 'No se pudo guardar'); }}
            toastFix(data.msg || 'No se pudo guardar el consumo.', false);
            limpiarDniParaSiguienteFix();
          }}
        }}catch(e){{
          if(ind){{ ind.style.display='block'; ind.textContent='❌ Error de conexión al guardar.'; }}
          toastFix('Error de conexión al guardar consumo.', false);
          limpiarDniParaSiguienteFix();
        }}finally{{
          setTimeout(()=>{{ autoGuardandoFix=false; }}, 250);
        }}
      }}
      window.buscarTrabajadorConsumo = async function(force=false){{
        if(!validarResponsableFix()){{
          limpiarDniParaSiguienteFix();
          return;
        }}
        const inp = document.getElementById('dni_consumo');
        const out = document.getElementById('nombre_trabajador');
        if(!inp || !out) return;
        const dni = onlyDni(inp.value);
        inp.value = dni;
        if(dni.length < 8){{ out.value=''; return; }}
        out.value='Validando DNI...';
        try{{
          const d = await validarDniFix(dni);
          if(d && d.ok){{
            out.value = d.nombre || '';
            const info = document.getElementById('info_trabajador_consumo');
            if(info){{ info.style.display='block'; info.innerHTML='<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:10px"><div><b>Trabajador</b><br>'+(d.nombre||'-')+'</div><div><b>DNI</b><br>'+dni+'</div><div><b>Área</b><br>'+(d.area||'-')+'</div><div><b>Estado</b><br><span class="badge ok">Activo</span></div></div>'; }}
            await registrarConsumoAutomaticoFix(dni);
          }}else{{
            out.value='DNI no encontrado';
            toastFix('DNI no encontrado: '+dni, false);
            limpiarDniParaSiguienteFix();
          }}
        }}catch(e){{
          out.value='Error validando DNI';
          toastFix('Error consultando trabajador.', false);
          limpiarDniParaSiguienteFix();
        }}
      }};
      window.dniInputHandler = function(){{
        const inp = document.getElementById('dni_consumo'); if(!inp) return;
        inp.value = onlyDni(inp.value);
        clearTimeout(window.__fixDniTimer);
        if(inp.value.length > 0 && !responsableFix()){{
          const out = document.getElementById('nombre_trabajador');
          if(out) out.value='PRIMERO COLOCA RESPONSABLE';
          inp.value='';
          validarResponsableFix();
          return;
        }}
        if(inp.value.length === 8){{ window.__fixDniTimer = setTimeout(()=>window.buscarTrabajadorConsumo(false), 80); }}
      }};
      window.procesarDniQR = async function(texto){{
        if(!validarResponsableFix()){{ limpiarDniParaSiguienteFix(); return; }}
        const dni = onlyDni(texto);
        if(dni.length !== 8){{ toastFix('QR/barras inválido: no contiene DNI de 8 dígitos.', false); return; }}
        const inp = document.getElementById('dni_consumo'); if(inp) inp.value=dni;
        await window.buscarTrabajadorConsumo(true);
      }};

      document.addEventListener('DOMContentLoaded', function(){{
        loadLoteFix(); try{{ if(sessionStorage.getItem('limpiar_lote_tras_envio_fix') === '1'){{ localStorage.removeItem(LS_KEY); sessionStorage.removeItem('limpiar_lote_tras_envio_fix'); loteMasivoFix=[]; }} }}catch(ex){{}} ensureIndicatorFix(); const form = document.getElementById('form_consumo'); if(form){{ form.onsubmit = window.validarAntesEnviar; }}
        function syncResponsibleLock(){{
          const has = !!responsableFix();
          const inp = document.getElementById('dni_consumo');
          const b1 = document.querySelector('button[onclick="buscarTrabajadorConsumo(true)"]');
          const b2 = document.getElementById('btn_qr');
          const chk0 = document.getElementById('modo_lote');
          if(inp){{ inp.placeholder = has ? 'Digite DNI o escanee QR/barras' : 'PRIMERO COLOCA RESPONSABLE'; }}
          if(b1){{ b1.title = has ? '' : 'Primero coloca RESPONSABLE'; }}
          if(b2){{ b2.title = has ? '' : 'Primero coloca RESPONSABLE'; }}
          if(!has){{
            if(inp) inp.value='';
            const out=document.getElementById('nombre_trabajador'); if(out) out.value='';
            const info=document.getElementById('info_trabajador_consumo'); if(info){{info.style.display='none'; info.innerHTML='';}}
          }} else {{
            if(chk0) chk0.checked = true;
          }}
          bloquearControlesPorResponsable();
          renderLoteFix();
        }}
        setTimeout(syncResponsibleLock, 80);
        const r = document.querySelector('#form_consumo [name="responsable"]'); if(r){{ r.addEventListener('input', function(){{ this.value=this.value.toUpperCase(); syncResponsibleLock(); renderTablaPreviewFix(); }}); }}
        const chk = document.getElementById('modo_lote'); if(chk){{ chk.addEventListener('change', function(){{ if(this.checked && !responsableFix()){{ validarResponsableFix(); }} renderLoteFix(); }}); }}
        ['tipo','comedor','fundo','cantidad','precio_unitario'].forEach(n => {{ const el=document.querySelector(`#form_consumo [name="${{n}}"]`); if(el) el.addEventListener('change', renderTablaPreviewFix); }});
        renderLoteFix();
      }});
    }})();
    </script>

    <br>
    {filtros}

    <div class="card">
      <div class="table-head">
        <h3>Consumos de la fecha {fecha_peru_txt(fecha)}</h3>
        <a class="btn btn-blue" href="{url_for('exportar_consumos', fecha=fecha)}">Exportar Excel</a>
      </div>
      <div class="table-wrap">
        <table id="tabla_consumos_principal">
          <thead><tr><th>Sel.</th><th>Fecha</th><th>Hora</th><th>DNI</th><th>Trabajador</th><th>Área</th><th>Tipo</th><th>Comedor</th><th>Fundo</th><th>Responsable</th><th>Cant.</th><th>P. Unit.</th><th>Total</th><th>Estado</th><th>Quitar</th></tr></thead>
          <tbody id="tbody_consumos_principal">
          {tabla}
          </tbody>
        </table>
      </div>
    </div>
    """
    return render_page(html, "consumos")



@app.route("/api/registrar_consumo_auto", methods=["POST"])
@login_required
@roles_required("admin", "rrhh", "comedor")
def api_registrar_consumo_auto():
    """Registro automático por DNI para modo masivo."""
    fecha = request.form.get("fecha") or hoy_iso()
    if fecha != hoy_iso():
        return jsonify({"ok": False, "msg": "Solo se puede registrar consumo en la fecha actual de hoy."}), 400
    if dia_cerrado(fecha):
        return jsonify({"ok": False, "msg": "El día ya está cerrado. No se puede registrar consumos."}), 400
    bloqueado, msg_bloq = registro_bloqueado()
    if bloqueado and session.get("role") != "admin":
        return jsonify({"ok": False, "msg": msg_bloq}), 400
    responsable = clean_text(request.form.get("responsable")).upper()
    if not responsable:
        return jsonify({"ok": False, "msg": "Primero registra el RESPONSABLE antes de detectar DNI."}), 400
    dni = clean_dni(request.form.get("dni"))
    if len(dni) != 8:
        return jsonify({"ok": False, "msg": "DNI inválido. Debe tener 8 dígitos."}), 400
    trabajador = q_one("SELECT * FROM trabajadores WHERE dni=? AND activo=1", (dni,))
    if not trabajador:
        return jsonify({"ok": False, "msg": f"DNI no encontrado o trabajador inactivo: {dni}"}), 404
    tipo = request.form.get("tipo", "Almuerzo")
    if tipo not in ["Almuerzo", "Dieta"]:
        tipo = "Almuerzo"
    comedor = request.form.get("comedor", "Comedor 01")
    fundo = request.form.get("fundo", "Kawsay Allpa")
    cantidad = int(float(request.form.get("cantidad") or 1))
    precio = float(request.form.get("precio_unitario") or 10)
    total = cantidad * precio
    obs = clean_text(request.form.get("observacion"))
    es_adicional = 1 if request.form.get("adicional") == "1" and session.get("role") == "admin" else 0
    if not es_adicional and q_one("SELECT id,hora FROM consumos WHERE fecha=? AND dni=? AND COALESCE(adicional,0)=0", (fecha, dni)):
        return jsonify({"ok": False, "msg": f"NO DUPLICADO: el DNI {dni} ya tiene consumo registrado hoy."}), 409
    hora = hora_now()
    try:
        new_id = q_exec("""
            INSERT INTO consumos(fecha,hora,dni,trabajador,empresa,area,tipo,cantidad,precio_unitario,total,observacion,comedor,fundo,responsable,adicional,estado,creado_por)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (fecha, hora, dni, trabajador["nombre"], trabajador["empresa"], trabajador["area"], tipo, cantidad, precio, total, obs, comedor, fundo, responsable, es_adicional, "PENDIENTE", session["user"]))
    except Exception:
        return jsonify({"ok": False, "msg": f"NO DUPLICADO: el DNI {dni} ya tiene consumo registrado para el día {fecha_peru_txt(fecha)}."}), 409
    if new_id:
        row = q_one("SELECT * FROM consumos WHERE id=?", (new_id,))
    else:
        row = q_one("SELECT * FROM consumos WHERE fecha=? AND dni=? ORDER BY id DESC", (fecha, dni))
    html = f"""
    <tr class="fila-db-consumo consumo-recien-guardado">
      <td>✅</td><td>{row['fecha']}</td><td>{row['hora']}</td><td>{row['dni']}</td><td>{row['trabajador']}</td><td>{row['area']}</td>
      <td>{row['tipo']}{' + Adic.' if row['adicional'] else ''}</td><td>{row['comedor']}</td><td>{row['fundo']}</td><td>{row['responsable'] or '-'}</td>
      <td>{row['cantidad']}</td><td>{money(row['precio_unitario'])}</td><td>{money(row['total'])}</td><td><span class="badge warn">{row['estado']}</span></td>
      <td><form method="post" action="{url_for('quitar_consumo')}" style="display:flex;gap:6px;align-items:center"><input type="hidden" name="id" value="{row['id']}"><input name="clave" placeholder="Clave" style="width:85px;padding:8px"><button class="btn-red" style="padding:8px 10px">Quitar</button></form></td>
    </tr>
    """
    audit_event("REGISTRO_CONSUMO_AUTO", "consumos", row["id"], f"DNI {dni} - responsable {responsable}")
    return jsonify({"ok": True, "msg": f"✅ Guardado automático: {dni} - {trabajador['nombre']}", "row_html": html, "dni": dni, "nombre": trabajador["nombre"], "area": trabajador["area"], "id": row["id"]})


@app.route("/quitar_consumo", methods=["POST"])
@login_required
@roles_required("admin", "rrhh", "comedor")
def quitar_consumo():
    id_ = request.form.get("id")
    c = q_one("SELECT * FROM consumos WHERE id=?", (id_,))
    if not c:
        flash("Consumo no encontrado.", "error")
        return redirect(url_for("consumos"))

    # NIVEL PRO: si ya está ENTREGADO, solo el administrador puede quitarlo.
    if c["estado"] == "ENTREGADO" and session.get("role") != "admin":
        audit_event("INTENTO_QUITAR_ENTREGADO_BLOQUEADO", "consumos", id_, f"DNI {c['dni']} - creado_por {c['creado_por']}")
        flash("Bloqueado: el pedido ya fue ENTREGADO. Solo un administrador puede quitarlo.", "error")
        return redirect(request.referrer or url_for("consumos"))

    clave = request.form.get("clave")
    if session.get("role") != "admin" and not require_remove_key(clave):
        flash("Clave incorrecta. No se quitó el consumo.", "error")
        return redirect(request.referrer or url_for("consumos"))

    if session.get("role") != "admin" and c["creado_por"] != session.get("user"):
        flash("Solo puedes quitar consumos registrados por tu usuario. El administrador puede quitar todos.", "error")
        return redirect(request.referrer or url_for("consumos"))

    audit_event("QUITAR_CONSUMO", "consumos", id_, f"DNI {c['dni']} - estado {c['estado']} - total {c['total']}")
    q_exec("DELETE FROM consumos WHERE id=?", (id_,))
    flash("Consumo quitado correctamente.", "ok")
    return redirect(request.referrer or url_for("consumos"))


@app.route("/api/entregas_pedidos")
@login_required
@roles_required("admin", "rrhh", "comedor")
def api_entregas_pedidos():
    fecha = request.args.get("fecha") or hoy_iso()
    dni = clean_dni(request.args.get("dni"))
    if dni:
        rows = q_all("SELECT * FROM consumos WHERE fecha=? AND dni=? ORDER BY hora,id", (fecha, dni))
    else:
        rows = q_all("SELECT * FROM consumos WHERE fecha=? ORDER BY CASE estado WHEN 'PENDIENTE' THEN 0 ELSE 1 END, hora, id", (fecha,))
    pedidos = []
    for i, r in enumerate(rows, 1):
        pedidos.append({
            "id": r["id"], "n": i, "hora": r["hora"], "dni": r["dni"], "trabajador": r["trabajador"], "tipo": r["tipo"],
            "cantidad": r["cantidad"], "observacion": r["observacion"] or "-",
            "estado": r["estado"], "pendiente": r["estado"] == "PENDIENTE"
        })
    return jsonify({"ok": True, "pedidos": pedidos, "count": len(pedidos)})


@app.route("/api/entregar_dni_auto", methods=["POST"])
@login_required
@roles_required("admin", "rrhh", "comedor")
def api_entregar_dni_auto():
    fecha = request.form.get("fecha") or hoy_iso()
    dni = clean_dni(request.form.get("dni"))
    responsable = clean_text(request.form.get("responsable") or session.get("user", "")).upper()
    if fecha != hoy_iso():
        return jsonify({"ok": False, "msg": "Solo se puede entregar en la fecha actual de hoy."}), 400
    if dia_cerrado(fecha):
        return jsonify({"ok": False, "msg": "Día cerrado. No se pueden entregar más pedidos."}), 400
    if not responsable:
        return jsonify({"ok": False, "msg": "Primero coloca RESPONSABLE DE ENTREGA."}), 400
    if len(dni) != 8:
        return jsonify({"ok": False, "msg": "DNI inválido. Debe tener 8 dígitos."}), 400
    trabajador = q_one("SELECT * FROM trabajadores WHERE dni=? AND activo=1", (dni,))
    if not trabajador:
        return jsonify({"ok": False, "msg": f"DNI no encontrado o trabajador inactivo: {dni}"}), 404
    pendientes = q_all("SELECT * FROM consumos WHERE fecha=? AND dni=? AND estado='PENDIENTE' ORDER BY hora,id", (fecha, dni))
    todos = q_all("SELECT * FROM consumos WHERE fecha=? AND dni=? ORDER BY hora,id", (fecha, dni))
    if not todos:
        return jsonify({"ok": False, "msg": f"{dni} - {trabajador['nombre']} no tiene consumo registrado hoy."}), 404
    if not pendientes:
        return jsonify({"ok": False, "msg": f"{dni} - {trabajador['nombre']} ya figura ENTREGADO o sin pendiente."}), 409
    entregado_en = now_app().strftime("%Y-%m-%d %H:%M:%S")
    for r in pendientes:
        q_exec("UPDATE consumos SET estado='ENTREGADO', entregado_por=?, entregado_en=? WHERE id=? AND estado='PENDIENTE'",
               (responsable or session["user"], entregado_en, r["id"]))
        audit_event("ENTREGA_AUTO_DNI", "consumos", r["id"], f"DNI {dni} - responsable {responsable}")
    rows = q_all("SELECT * FROM consumos WHERE fecha=? AND dni=? ORDER BY hora,id", (fecha, dni))
    row_html = "".join([f"""
      <tr class="entrega-ok-row">
        <td><input type="checkbox" name="ids" value="{r['id']}" {'disabled' if r['estado']!='PENDIENTE' else 'checked'}></td>
        <td>{i}</td><td>{r['hora']}</td><td>{r['dni']}</td><td>{r['trabajador']}</td><td>{r['tipo']}</td><td>{r['cantidad']}</td>
        <td>{r['observacion'] or '-'}</td><td><span class="badge {'ok' if r['estado']=='ENTREGADO' else 'warn'}">{r['estado']}</span></td>
      </tr>""" for i, r in enumerate(rows, 1)])
    return jsonify({
        "ok": True,
        "msg": f"✅ ENTREGADO: {dni} - {trabajador['nombre']} ({len(pendientes)} pedido(s))",
        "dni": dni,
        "nombre": trabajador["nombre"],
        "area": trabajador["area"],
        "entregados": len(pendientes),
        "row_html": row_html,
        "count": len(rows)
    })

@app.route("/entregas", methods=["GET", "POST"])
@login_required
@roles_required("admin", "rrhh", "comedor")
def entregas():
    fecha = request.values.get("fecha") or hoy_iso()
    dni = clean_dni(request.values.get("dni"))

    if request.method == "POST":
        if dia_cerrado(fecha):
            flash("Día cerrado. No se pueden entregar más pedidos.", "error")
            return redirect(url_for("entregas", fecha=fecha, dni=dni))
        responsable = clean_text(request.form.get("responsable_entrega") or session.get("user", "")).upper()
        ids = request.form.getlist("ids")
        if request.form.get("entregar_todos") == "1":
            if dni:
                ids = [str(r["id"]) for r in q_all("SELECT id FROM consumos WHERE fecha=? AND dni=? AND estado='PENDIENTE'", (fecha, dni))]
            else:
                ids = [str(r["id"]) for r in q_all("SELECT id FROM consumos WHERE fecha=? AND estado='PENDIENTE'", (fecha,))]
        if not ids:
            flash("No hay pedidos pendientes seleccionados para entregar.", "error")
            return redirect(url_for("entregas", dni=dni, fecha=fecha))
        for id_ in ids:
            q_exec("UPDATE consumos SET estado='ENTREGADO', entregado_por=?, entregado_en=? WHERE id=? AND estado='PENDIENTE'",
                   (responsable or session["user"], now_app().strftime("%Y-%m-%d %H:%M:%S"), id_))
            audit_event("ENTREGAR_PEDIDO", "consumos", id_, f"DNI {dni} - responsable {responsable}")
        flash(f"Pedidos entregados: {len(ids)}", "ok")
        return redirect(url_for("entregas", dni=dni, fecha=fecha))

    trabajador = q_one("SELECT * FROM trabajadores WHERE dni=? AND activo=1", (dni,)) if dni else None
    pedidos = q_all("SELECT * FROM consumos WHERE fecha=? AND dni=? ORDER BY hora,id", (fecha, dni)) if dni else q_all("SELECT * FROM consumos WHERE fecha=? ORDER BY CASE estado WHEN 'PENDIENTE' THEN 0 ELSE 1 END, hora,id", (fecha,))

    info = ""
    if dni and trabajador:
        info = f"""
        <div class="card" style="margin-top:12px;padding:14px" id="info_trabajador_entrega">
          <div style="display:grid;grid-template-columns:1fr 1fr auto;gap:15px">
            <div><b>Trabajador</b><br>{trabajador['nombre']}</div>
            <div><b>Área</b><br>{trabajador['area']}</div>
            <div><b>Estado</b><br><span class="badge ok">Activo</span></div>
          </div>
        </div>
        """
    elif dni:
        info = '<div class="flash error" id="info_trabajador_entrega">DNI no encontrado o trabajador inactivo.</div>'
    else:
        info = '<div id="info_trabajador_entrega" style="display:none"></div>'

    tabla = "".join([
        f"""
        <tr>
          <td><input type="checkbox" name="ids" value="{r['id']}" {'disabled' if r['estado']!='PENDIENTE' else 'checked'}></td>
          <td>{i}</td><td>{r['hora']}</td><td>{r['dni']}</td><td>{r['trabajador']}</td><td>{r['tipo']}</td><td>{r['cantidad']}</td>
          <td>{r['observacion'] or '-'}</td>
          <td><span class="badge {'ok' if r['estado']=='ENTREGADO' else 'warn'}">{r['estado']}</span></td>
        </tr>
        """ for i, r in enumerate(pedidos, 1)
    ]) or "<tr><td colspan='9'>Sin pedidos para este DNI hoy.</td></tr>"

    html = topbar("Entrega de Pedidos", "Lectura individual y masiva por DNI igual que Consumos") + f"""
    <div class="card">
      <h3 style="margin-top:0">Entrega rápida por DNI</h3>
      <div class="entrega-pro-panel">
        <form method="get" class="form-grid two" id="form_entrega_busqueda">
          <input type="date" id="fecha_entrega" name="fecha" value="{fecha}">
          <input id="dni_entrega" name="dni" value="{dni}" placeholder="DNI del trabajador" inputmode="numeric" autocomplete="off" maxlength="8" autofocus oninput="dniEntregaHandler()">
          <input id="nombre_trabajador_entrega" readonly placeholder="Nombre del trabajador" value="{trabajador['nombre'] if trabajador else ''}">
          <input id="responsable_entrega" name="responsable_entrega" placeholder="RESPONSABLE DE ENTREGA" value="{session.get('user','').upper()}" oninput="this.value=this.value.toUpperCase()">
          <label class="label-lote-final" style="grid-column:1/-1">
            <input type="checkbox" id="modo_lote_entrega" checked>
            Entrega masiva automática: cada DNI válido se ENTREGA al instante y queda listado abajo.
          </label>
          <button type="button" class="btn-blue" onclick="buscarTrabajadorEntrega(true)">🔎 Validar / entregar DNI</button>
          <button type="button" class="btn-blue" onclick="refrescarEntregas()">🔄 Actualizar / refrescar</button>
          <button type="button" id="btn_qr_entrega" class="btn-orange" onclick="abrirScannerEntrega()">📷 Scanner QR / Barras</button>
          <button type="button" class="btn-red" onclick="limpiarEntregaRapida()">Limpiar</button>
        </form>
        {info}
        <div class="entrega-pro-status">
          <div>✅ Entregas guardadas: <b id="entrega_auto_count">0</b></div>
          <div>📌 Último DNI: <b id="entrega_ultimo_dni">-</b></div>
          <div>👤 Último trabajador: <b id="entrega_ultimo_nombre">-</b></div>
        </div>
        <div id="estado_entrega_auto" style="display:none;padding:12px;border-radius:12px;background:#dcfce7;color:#166534;font-weight:950"></div>
        <div id="qr_entrega_box" style="display:none;margin-top:12px;padding:12px;border-radius:16px;background:#061a2d;color:white">
          <div id="qr_entrega_reader" style="width:100%;max-width:380px;margin:auto"></div>
          <button type="button" class="btn-red" onclick="cerrarScannerEntrega()" style="margin-top:10px">Cerrar scanner</button>
        </div>
      </div>
    </div>

    <br>
    <div class="card">
      <div class="table-head">
        <h3>Pedidos del día ({fecha_peru_txt(fecha)})</h3>
        <span id="contador_pedidos" class="badge ok">{len(pedidos)} pedido(s)</span>
      </div>
      <form method="post">
        <input type="hidden" name="fecha" value="{fecha}">
        <input type="hidden" name="dni" value="{dni}">
        <input type="hidden" name="responsable_entrega" id="responsable_entrega_post" value="{session.get('user','').upper()}">
        <div class="table-wrap">
          <table>
            <thead><tr><th></th><th>#</th><th>Hora</th><th>DNI</th><th>Trabajador</th><th>Tipo</th><th>Cantidad</th><th>Observación</th><th>Estado</th></tr></thead>
            <tbody id="pedidos_body">{tabla}</tbody>
          </table>
        </div>
        <br>
        <button name="entregar_seleccionado" value="1">Entregar seleccionado</button>
        <button name="entregar_todos" value="1" class="btn-blue">Entregar todos pendientes</button>
      </form>
      <p class="muted small">Lectura rápida activa. Al digitar o escanear 8 dígitos, valida el DNI y entrega automáticamente los pendientes.</p>
    </div>
    <script>
    let entregaTimer=null, entregaBusy=false, entregaCount=0, qrEntrega=null;
    function onlyDniEntrega(v){{ const d=String(v||'').replace(/\D/g,''); return d.length>8 ? d.slice(-8) : d; }}
    function entregaToast(msg, ok=true){{
      const d=document.createElement('div'); d.textContent=msg;
      d.style.cssText='position:fixed;left:10px;right:10px;top:14px;z-index:999999;padding:13px;border-radius:13px;text-align:center;font-weight:950;color:white;background:'+(ok?'#166534':'#991b1b')+';box-shadow:0 12px 30px rgba(0,0,0,.35)';
      document.body.appendChild(d); setTimeout(()=>d.remove(),2300);
      try{{ if(navigator.vibrate) navigator.vibrate(ok?90:[80,50,80]); const C=window.AudioContext||window.webkitAudioContext; const c=new C(); const o=c.createOscillator(); const g=c.createGain(); o.connect(g); g.connect(c.destination); o.frequency.value=ok?980:220; g.gain.value=.08; o.start(); setTimeout(()=>{{o.stop();c.close();}},140); }}catch(e){{}}
    }}
    function responsableEntrega(){{ return String(document.getElementById('responsable_entrega')?.value||'').trim().toUpperCase(); }}
    function setEstadoEntrega(msg, ok=true){{ const e=document.getElementById('estado_entrega_auto'); if(e){{ e.style.display='block'; e.style.background=ok?'#dcfce7':'#fee2e2'; e.style.color=ok?'#166534':'#991b1b'; e.textContent=msg; }} }}
    function limpiarEntregaRapida(){{ const i=document.getElementById('dni_entrega'); const n=document.getElementById('nombre_trabajador_entrega'); if(i){{i.value='';i.focus();}} if(n)n.value=''; setEstadoEntrega('Listo para nueva lectura.', true); }}
    async function buscarTrabajadorEntrega(force=false){{
      if(entregaBusy) return;
      const inp=document.getElementById('dni_entrega'); const nom=document.getElementById('nombre_trabajador_entrega');
      const fecha=document.getElementById('fecha_entrega')?.value||''; const responsable=responsableEntrega();
      if(!responsable){{ entregaToast('Primero coloca RESPONSABLE DE ENTREGA.', false); if(inp) inp.value=''; return; }}
      const dni=onlyDniEntrega(inp?.value||''); if(inp) inp.value=dni; if(dni.length<8) return;
      entregaBusy=true; if(nom) nom.value='Validando y entregando...'; setEstadoEntrega('⏳ Validando DNI y entregando pedidos pendientes...', true);
      try{{
        const fd=new FormData(); fd.append('dni',dni); fd.append('fecha',fecha); fd.append('responsable',responsable);
        const res=await fetch('/api/entregar_dni_auto', {{method:'POST', body:fd}}); const data=await res.json();
        if(data.ok){{
          if(nom) nom.value=data.nombre||''; entregaCount += Number(data.entregados||1);
          document.getElementById('entrega_auto_count').textContent=entregaCount;
          document.getElementById('entrega_ultimo_dni').textContent=data.dni||dni;
          document.getElementById('entrega_ultimo_nombre').textContent=data.nombre||'-';
          const body=document.getElementById('pedidos_body'); if(body && data.row_html) body.innerHTML=data.row_html;
          const cont=document.getElementById('contador_pedidos'); if(cont) cont.textContent=(data.count||0)+' pedido(s)';
          setEstadoEntrega(data.msg, true); entregaToast(data.msg, true);
        }}else{{ if(nom) nom.value=''; setEstadoEntrega(data.msg||'No se pudo entregar.', false); entregaToast(data.msg||'No se pudo entregar.', false); }}
      }}catch(e){{ setEstadoEntrega('Error de conexión al entregar.', false); entregaToast('Error de conexión al entregar.', false); }}
      finally{{ setTimeout(()=>{{ entregaBusy=false; if(inp){{ inp.value=''; inp.focus(); }} }},220); }}
    }}
    function dniEntregaHandler(){{ const inp=document.getElementById('dni_entrega'); if(!inp) return; inp.value=onlyDniEntrega(inp.value); clearTimeout(entregaTimer); if(inp.value.length===8) entregaTimer=setTimeout(()=>buscarTrabajadorEntrega(false),70); }}
    async function refrescarEntregas(){{
      const dni=document.getElementById('dni_entrega')?.value||''; const fecha=document.getElementById('fecha_entrega')?.value||'';
      try{{ const res=await fetch(`/api/entregas_pedidos?dni=${{encodeURIComponent(dni)}}&fecha=${{encodeURIComponent(fecha)}}`); const data=await res.json(); const body=document.getElementById('pedidos_body'); const contador=document.getElementById('contador_pedidos'); if(contador) contador.textContent=`${{data.count}} pedido(s)`; if(!body) return; if(!data.pedidos||data.pedidos.length===0){{ body.innerHTML='<tr><td colspan="9">Sin pedidos para este DNI hoy.</td></tr>'; return; }} body.innerHTML=data.pedidos.map(p=>`<tr><td><input type="checkbox" name="ids" value="${{p.id}}" ${{p.pendiente?'checked':'disabled'}}></td><td>${{p.n}}</td><td>${{p.hora}}</td><td>${{p.dni||dni||'-'}}</td><td>${{p.trabajador||'-'}}</td><td>${{p.tipo}}</td><td>${{p.cantidad}}</td><td>${{p.observacion}}</td><td><span class="badge ${{p.estado==='ENTREGADO'?'ok':'warn'}}">${{p.estado}}</span></td></tr>`).join(''); }}catch(e){{console.warn(e)}}
    }}
    async function abrirScannerEntrega(){{
      const box=document.getElementById('qr_entrega_box'); if(box) box.style.display='block';
      if(typeof Html5Qrcode==='undefined'){{ entregaToast('No cargó librería de cámara. Digita el DNI o recarga.', false); return; }}
      try{{ cerrarScannerEntrega(); qrEntrega=new Html5Qrcode('qr_entrega_reader'); await qrEntrega.start({{facingMode:'environment'}}, {{fps:12, qrbox:260}}, txt=>{{ const dni=onlyDniEntrega(txt); if(dni.length===8){{ document.getElementById('dni_entrega').value=dni; buscarTrabajadorEntrega(true); }} }}); }}catch(e){{ entregaToast('No se pudo abrir cámara. Revisa permisos del navegador.', false); }}
    }}
    function cerrarScannerEntrega(){{ try{{ if(qrEntrega){{ qrEntrega.stop().catch(()=>{{}}); qrEntrega.clear(); qrEntrega=null; }} }}catch(e){{}} const box=document.getElementById('qr_entrega_box'); if(box) box.style.display='none'; }}
    document.addEventListener('DOMContentLoaded',()=>{{ const r=document.getElementById('responsable_entrega'); const rp=document.getElementById('responsable_entrega_post'); if(r&&rp) r.addEventListener('input',()=>rp.value=r.value.toUpperCase()); }});
    </script>
    """
    return render_page(html, "entregas")

@app.route("/carga_masiva", methods=["GET", "POST"])
@login_required
@roles_required("admin", "rrhh", "comedor")
def carga_masiva():
    if request.method == "POST":
        if dia_cerrado():
            flash("Día cerrado. No se puede cargar consumos.", "error")
            return redirect(url_for("carga_masiva"))

        f = request.files.get("excel")
        if not f or not f.filename.lower().endswith((".xlsx", ".xls")):
            flash("Sube un archivo Excel válido.", "error")
            return redirect(url_for("carga_masiva"))

        try:
            df = pd.read_excel(f, dtype=str, engine="openpyxl" if f.filename.lower().endswith(".xlsx") else None).fillna("")
            df.columns = normalize_columns(df.columns)
        except Exception:
            flash("No se pudo leer el Excel. Guarda el archivo como .xlsx y vuelve a cargarlo.", "error")
            return redirect(url_for("carga_masiva"))

        if "DNI" not in df.columns:
            flash("Falta la columna DNI. Usa la plantilla.", "error")
            return redirect(url_for("carga_masiva"))

        total = len(df)
        creados = 0
        errores = 0

        for _, r in df.iterrows():
            dni = clean_dni(col_value(r, "DNI"))
            trabajador = q_one("SELECT * FROM trabajadores WHERE dni=? AND activo=1", (dni,))
            if not trabajador:
                errores += 1
                continue

            fecha_raw = clean_text(r.get("FECHA"))
            if fecha_raw:
                try:
                    fecha = pd.to_datetime(fecha_raw).date().isoformat()
                except Exception:
                    fecha = hoy_iso()
            else:
                fecha = hoy_iso()

            if dia_cerrado(fecha):
                errores += 1
                continue

            if q_one("SELECT id FROM consumos WHERE fecha=? AND dni=? AND COALESCE(adicional,0)=0", (fecha, dni)):
                errores += 1
                continue

            tipo = clean_text(r.get("TIPO")) or "Almuerzo"
            if tipo not in ["Almuerzo", "Dieta"]:
                tipo = "Almuerzo"
            comedor = clean_text(r.get("COMEDOR")) or "Comedor 01"
            fundo = clean_text(r.get("FUNDO")) or "Kawsay Allpa"
            responsable = clean_text(r.get("RESPONSABLE"))
            cantidad = int(float(r.get("CANTIDAD") or 1))
            precio = float(r.get("PRECIO_UNITARIO") or r.get("PRECIO") or 10)
            obs = clean_text(r.get("OBSERVACION"))
            q_exec("""
                INSERT INTO consumos(fecha,hora,dni,trabajador,empresa,area,tipo,cantidad,precio_unitario,total,observacion,comedor,fundo,responsable,adicional,estado,creado_por)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (fecha, hora_now(), dni, trabajador["nombre"], trabajador["empresa"], trabajador["area"], tipo, cantidad, precio, cantidad*precio, obs, comedor, fundo, responsable, 0, "PENDIENTE", session["user"]))
            creados += 1

        q_exec("INSERT INTO importaciones(archivo,total,creados,errores,usuario) VALUES(?,?,?,?,?)",
               (f.filename, total, creados, errores, session["user"]))
        flash(f"Carga terminada: {creados} creados, {errores} errores.", "ok" if errores == 0 else "error")
        return redirect(url_for("carga_masiva"))

    hist = q_all("SELECT * FROM importaciones ORDER BY id DESC LIMIT 10")
    tabla = "".join([
        f"<tr><td>{r['fecha_hora']}</td><td>{r['archivo']}</td><td>{r['total']}</td><td>{r['creados']}</td><td>{r['errores']}</td><td>{r['usuario']}</td><td>⬇️</td></tr>"
        for r in hist
    ]) or "<tr><td colspan='7'>Sin historial de importaciones.</td></tr>"

    html = topbar("Carga Masiva de Consumos", "Importa consumos desde un archivo Excel") + f"""
    <div class="card">
      <form method="post" enctype="multipart/form-data">
        <input type="file" name="excel" accept=".xlsx,.xls" required>
        <br><br>
        <button class="btn-orange">Importar consumos</button>
        <a class="btn btn-blue" href="{url_for('plantilla_consumos')}">Descargar plantilla Excel</a>
      </form>
    </div>

    <br>
    <div class="card">
      <h3 style="margin-top:0">Historial de importaciones</h3>
      <div class="table-wrap">
        <table>
          <tr><th>Fecha</th><th>Archivo</th><th>Total</th><th>Creados</th><th>Errores</th><th>Usuario</th><th></th></tr>
          {tabla}
        </table>
      </div>
    </div>
    """
    return render_page(html, "carga")


@app.route("/trabajadores", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def trabajadores():
    if request.method == "POST" and request.form.get("manual") == "1":
        dni = clean_dni(request.form.get("dni"))
        nombre = clean_text(request.form.get("nombre"))
        empresa = clean_text(request.form.get("empresa")) or "PRIZE"
        cargo = clean_text(request.form.get("cargo"))
        area = clean_text(request.form.get("area"))
        if len(dni) != 8 or not nombre:
            flash("Ingresa un DNI de 8 dígitos y nombre válido.", "error")
            return redirect(url_for("trabajadores"))

        existe = q_one("SELECT id FROM trabajadores WHERE dni=?", (dni,))
        if existe:
            q_exec("UPDATE trabajadores SET empresa=?,nombre=?,cargo=?,area=?,activo=1,actualizado=CURRENT_TIMESTAMP WHERE dni=?",
                   (empresa, nombre, cargo, area, dni))
        else:
            q_exec("INSERT INTO trabajadores(empresa,dni,nombre,cargo,area,activo) VALUES(?,?,?,?,?,1)",
                   (empresa, dni, nombre, cargo, area))
        flash("Trabajador guardado correctamente.", "ok")
        return redirect(url_for("trabajadores"))

    if request.method == "POST" and "excel" in request.files:
        f = request.files.get("excel")
        try:
            if not f or not f.filename:
                flash("Selecciona un archivo Excel para importar.", "error")
                return redirect(url_for("trabajadores"))
            if not f.filename.lower().endswith((".xlsx", ".xls")):
                flash("Sube un archivo Excel válido (.xlsx o .xls).", "error")
                return redirect(url_for("trabajadores"))

            registros_dict, total_filas, omitidos = leer_trabajadores_excel_stream(f)

            if not registros_dict:
                flash("No se importó nada: no encontré filas válidas con DNI de 8 dígitos y NOMBRE. Descarga la plantilla y vuelve a intentar.", "error")
                return redirect(url_for("trabajadores"))

            # REEMPLAZO TOTAL OPTIMIZADO:
            # Carga TODO el Excel válido, pero en una sola transacción y por lotes.
            # Esto evita que Render mate el proceso por abrir miles de conexiones.
            creados = reemplazar_trabajadores_batch(list(registros_dict.values()))

            q_exec("INSERT INTO importaciones(archivo,total,creados,errores,usuario) VALUES(?,?,?,?,?)",
                   (f.filename, total_filas, creados, omitidos, session.get("user", "")))
            flash(f"Base de trabajadores reemplazada correctamente: {creados} trabajadores cargados desde todo el Excel. Omitidos: {omitidos}.", "ok")
            return redirect(url_for("trabajadores"))
        except Exception as e:
            app.logger.exception("Error importando trabajadores")
            flash("No se pudo importar trabajadores. El Excel debe tener como mínimo DNI y NOMBRE/TRABAJADOR. También acepta EMPRESA, CARGO y AREA si existen. Detalle: " + str(e)[:180], "error")
            return redirect(url_for("trabajadores"))

    buscar = clean_text(request.args.get("buscar"))
    total_activos = q_one("SELECT COUNT(*) AS total FROM trabajadores WHERE activo=1")
    total_activos = int(total_activos["total"] if total_activos else 0)
    total_inactivos = q_one("SELECT COUNT(*) AS total FROM trabajadores WHERE activo=0")
    total_inactivos = int(total_inactivos["total"] if total_inactivos else 0)
    if buscar:
        b = f"%{buscar}%"
        rows = q_all("""
            SELECT * FROM trabajadores
            WHERE dni LIKE ? OR nombre LIKE ? OR cargo LIKE ? OR area LIKE ? OR empresa LIKE ?
            ORDER BY nombre
        """, (b, b, b, b, b))
    else:
        rows = q_all("SELECT * FROM trabajadores ORDER BY nombre")

    tabla = "".join([
        f"<tr><td>{r['empresa']}</td><td>{r['dni']}</td><td>{r['nombre']}</td><td>{r['cargo']}</td><td>{r['area']}</td><td><span class='badge ok'>Activo</span></td></tr>"
        for r in rows
    ]) or "<tr><td colspan='6'>Sin trabajadores encontrados.</td></tr>"

    html = topbar("Trabajadores", "Base de trabajadores activos para validar DNI") + f"""
    <div class="kpi-grid" style="grid-template-columns:repeat(3,minmax(180px,1fr))!important">
      <div class="card kpi-card">
        <div class="icon-circle ic-green">👥</div>
        <div>
          <div class="label">Trabajadores activos</div>
          <div class="num">{total_activos}</div>
          <div class="sub">Disponibles para validar DNI</div>
        </div>
      </div>
      <div class="card kpi-card">
        <div class="icon-circle ic-blue">🔎</div>
        <div>
          <div class="label">Resultado mostrado</div>
          <div class="num">{len(rows)}</div>
          <div class="sub">Según filtro actual</div>
        </div>
      </div>
      <div class="card kpi-card">
        <div class="icon-circle ic-orange">⛔</div>
        <div>
          <div class="label">Inactivos</div>
          <div class="num">{total_inactivos}</div>
          <div class="sub">No validan consumo</div>
        </div>
      </div>
    </div>

    <div class="card">
      <h3 style="margin-top:0">Registro manual</h3>
      <form method="post" class="form-grid" id="form_consumo" onsubmit="return validarAntesEnviar(event)">
        <input type="hidden" name="manual" value="1">
        <input name="empresa" value="PRIZE" placeholder="Empresa">
        <input name="dni" placeholder="DNI" required>
        <input name="nombre" placeholder="Apellidos y nombres" required>
        <input name="cargo" placeholder="Cargo">
        <input name="area" placeholder="Área">
        <button>Guardar</button>
      </form>
    </div>

    <br>
    <div class="card">
      <h3 style="margin-top:0">Carga masiva trabajadores</h3>
      <p class="muted small"><b>Importante:</b> al importar, la base de trabajadores se REEMPLAZA por la información del Excel.</p>
      <form method="post" enctype="multipart/form-data" class="form-grid">
        <input type="file" name="excel" accept=".xlsx,.xls" required>
        <button class="btn-orange">Importar y reemplazar trabajadores</button>
        <a class="btn btn-blue" href="{url_for('plantilla_trabajadores')}">Descargar plantilla</a>
      </form>
    </div>

    <br>
    <div class="card">
      <div class="table-head">
        <h3>Base de trabajadores</h3>
      </div>

      <form method="get" action="{url_for('trabajadores')}" class="form-grid" style="grid-template-columns:1fr auto auto;margin-bottom:14px">
        <input name="buscar" value="{buscar}" placeholder="Buscar por DNI, nombre, cargo, área o empresa">
        <button class="btn-blue">Buscar</button>
        <a class="btn" href="{url_for('trabajadores')}">Actualizar</a>
      </form>

      <div class="table-wrap">
        <table>
          <tr><th>Empresa</th><th>DNI</th><th>Nombre</th><th>Cargo</th><th>Área</th><th>Estado</th></tr>
          {tabla}
        </table>
      </div>
    </div>
    """
    return render_page(html, "trabajadores")


@app.route("/cierre_dia", methods=["GET", "POST"])
@login_required
@roles_required("admin", "comedor", "rrhh")
def cierre_dia():
    fecha = hoy_iso()
    cerrado = dia_cerrado(fecha)

    if request.method == "POST":
        if cerrado:
            flash("Este día ya fue cerrado.", "error")
            return redirect(url_for("cierre_dia"))

        correo = clean_text(request.form.get("correo"))
        pedidos = q_all("SELECT * FROM consumos WHERE fecha=? ORDER BY area,trabajador", (fecha,))
        df = pd.DataFrame([dict(p) for p in pedidos])
        if df.empty:
            df = pd.DataFrame(columns=["fecha","hora","dni","trabajador","empresa","area","tipo","cantidad","precio_unitario","total","estado","creado_por"])

        resumen_area = df.groupby(["area","estado"], as_index=False).agg(cantidad=("cantidad","sum"), total=("total","sum")) if not df.empty else pd.DataFrame()
        resumen_usuario = df.groupby(["creado_por"], as_index=False).agg(consumos=("dni","count"), total=("total","sum")) if not df.empty else pd.DataFrame()

        filename = f"cierre_comedor_{fecha.replace('-','_')}.xlsx"
        path = os.path.join(REPORT_DIR, filename)
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="DETALLE_DIA", index=False)
            resumen_area.to_excel(writer, sheet_name="RESUMEN_AREA", index=False)
            resumen_usuario.to_excel(writer, sheet_name="RESUMEN_USUARIOS", index=False)

        total_consumos = len(pedidos)
        total_entregados = sum(1 for p in pedidos if p["estado"] == "ENTREGADO")
        total_pendientes = total_consumos - total_entregados
        total_importe = sum(float(p["total"] or 0) for p in pedidos)

        estado_correo = send_report_email(
            correo,
            f"Cierre comedor PRIZE {fecha_peru_txt(fecha)}",
            f"Se adjunta cierre del día. Consumos: {total_consumos}. Entregados: {total_entregados}. Pendientes: {total_pendientes}. Total: {money(total_importe)}",
            path
        )

        q_exec("""
            INSERT INTO cierres(fecha,cerrado_por,total_consumos,total_entregados,total_pendientes,total_importe,archivo_excel,correo_destino,correo_estado)
            VALUES(?,?,?,?,?,?,?,?,?)
        """, (fecha, session["user"], total_consumos, total_entregados, total_pendientes, total_importe, filename, correo, estado_correo))

        flash(f"Día cerrado. Reporte generado: {filename}. Correo: {estado_correo}", "ok")
        return redirect(url_for("cierre_dia"))

    stats = q_one("""
        SELECT COUNT(*) c, COALESCE(SUM(total),0) t,
        SUM(CASE WHEN estado='ENTREGADO' THEN 1 ELSE 0 END) e
        FROM consumos WHERE fecha=?
    """, (fecha,))
    usuarios = q_all("SELECT creado_por, COUNT(*) c, COALESCE(SUM(total),0) t FROM consumos WHERE fecha=? GROUP BY creado_por", (fecha,))
    ultimo = q_one("SELECT hora FROM consumos WHERE fecha=? ORDER BY hora DESC,id DESC LIMIT 1", (fecha,))
    cerrado_html = ""
    if cerrado:
        cerrado_html = f"""
        <div class="card">
          <span class="badge off">DÍA CERRADO</span>
          <p>Archivo generado: <b>{cerrado['archivo_excel']}</b></p>
          <a class="btn btn-blue" href="{url_for('descargar_cierre', filename=cerrado['archivo_excel'])}">Descargar reporte</a>
        </div>
        """
    usuarios_html = "".join([
        f"<div class='user-row'><span>👤 <b>{u['creado_por'] or 'sin usuario'}</b></span><span>{u['c']} consumos</span><span>{money(u['t'])}</span></div>"
        for u in usuarios
    ]) or "<div class='muted'>Sin usuarios con registros hoy.</div>"

    form = "" if cerrado else f"""
    <form method="post">
      <label><b>Correo destino</b></label><br><br>
      <input name="correo" value="{os.getenv('REPORTE_DESTINO','administracion@prize.pe')}" placeholder="correo@empresa.com">
      <br><br>
      <label><input type="checkbox" checked> Incluir archivo Excel</label>
      <br><br>
      <button class="btn-orange" style="width:100%">Cerrar día y enviar reporte</button>
    </form>
    """

    admin_extra = ""
    if session.get("role") == "admin":
        admin_extra = f"""
        <div class='admin-actions'>
          <a class='btn btn-orange' href='{url_for('cerrar_dia_manual')}'>🔒 Cerrar día</a>
          <a class='btn btn-blue' href='{url_for('abrir_dia_manual')}'>🔓 Abrir día</a>
          <a class='btn' href='{url_for('exportar_concesionaria')}'>Archivo concesionaria</a>
          <a class='btn btn-orange' href='{url_for('reporte_entrega')}'>Reporte entrega/pago</a>
        </div>
        """
    html = topbar("Cierre de Día y Reportes", "Consolida y envía el reporte del día por correo") + admin_extra + f"""
    <div class="card">
      <span class="badge {'off' if cerrado else 'ok'}">🟢 {'DÍA CERRADO' if cerrado else 'DÍA ABIERTO'}</span>
      <span style="margin-left:18px" class="muted">Fecha actual: {fecha_peru_txt(fecha)}</span>

      <div class="mini-kpis">
        <div class="card"><span class="muted small">Total consumos</span><b>{stats['c']}</b></div>
        <div class="card"><span class="muted small">Total facturado</span><b>{money(stats['t'])}</b></div>
        <div class="card"><span class="muted small">Usuarios que registraron</span><b>{len(usuarios)}</b></div>
        <div class="card"><span class="muted small">Último registro</span><b>{ultimo['hora'] if ultimo else '--:--'}</b></div>
      </div>
    </div>

    <br>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:18px">
      <div class="card">
        <h3 style="margin-top:0">Usuarios que registraron hoy</h3>
        {usuarios_html}
      </div>
      <div class="card">
        <h3 style="margin-top:0">Enviar reporte por correo</h3>
        {form}
        {cerrado_html}
      </div>
    </div>
    """
    return render_page(html, "cierre")


@app.route("/reportes")
@login_required
def reportes():
    fecha_inicio = request.args.get("fecha_inicio") or request.args.get("fecha") or hoy_iso()
    fecha_fin = request.args.get("fecha_fin") or fecha_inicio
    buscar = clean_text(request.args.get("buscar"))

    cond, params = rango_sql(fecha_inicio, fecha_fin)
    where = cond
    final_params = list(params)

    if buscar:
        where += " AND (cerrado_por LIKE ? OR correo_estado LIKE ? OR archivo_excel LIKE ? OR correo_destino LIKE ?)"
        b = f"%{buscar}%"
        final_params += [b, b, b, b]

    rows = q_all(f"SELECT * FROM cierres WHERE {where} ORDER BY fecha DESC LIMIT 200", tuple(final_params))

    tabla = "".join([
        f"""
        <tr>
          <td>{fecha_peru_txt(r['fecha'])}</td>
          <td>{r['total_consumos']}</td>
          <td>{money(r['total_importe'])}</td>
          <td>{r['cerrado_por']}</td>
          <td>{r['correo_destino'] or '-'}</td>
          <td>{r['correo_estado'] or '-'}</td>
          <td>{('<a class="btn btn-blue" href="' + url_for('descargar_cierre', filename=r['archivo_excel']) + '">Descargar</a>') if r['archivo_excel'] else '-'}</td>
        </tr>
        """
        for r in rows
    ]) or "<tr><td colspan='7'>Sin reportes en el rango seleccionado.</td></tr>"

    html = topbar("Reportes", "Historial de cierres y reportes generados") + filtro_bar(url_for("reportes"), fecha_inicio, fecha_fin, buscar) + f"""
    <div class="card">
      <div class="table-head">
        <h3>Reportes generados</h3>
      </div>
      <div class="table-wrap">
        <table>
          <tr>
            <th>Fecha</th>
            <th>Consumos</th>
            <th>Total</th>
            <th>Cerrado por</th>
            <th>Correo destino</th>
            <th>Estado correo</th>
            <th>Archivo</th>
          </tr>
          {tabla}
        </table>
      </div>
    </div>
    """
    return render_page(html, "reportes")


@app.route("/configuracion", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def configuracion():
    if request.method == "POST":
        cfg_set("bloqueo_activo", "1" if request.form.get("bloqueo_activo") else "0")
        cfg_set("hora_inicio", request.form.get("hora_inicio") or "00:00")
        cfg_set("hora_fin", request.form.get("hora_fin") or "23:59")
        cfg_set("clave_quitar", request.form.get("clave_quitar") or "1234")
        flash("Configuración actualizada.", "ok")
        return redirect(url_for("configuracion"))

    usuarios = q_all("SELECT id, username, role, active, COALESCE(password_plain,'') AS password_plain FROM usuarios ORDER BY username")
    usuarios_html = "".join([
        f"<tr><td>{u['username']}</td><td>{u['role']}</td><td><span class='badge {'ok' if u['active'] else 'off'}'>{'Activo' if u['active'] else 'Bloqueado'}</span></td></tr>"
        for u in usuarios
    ])

    html = topbar("Configuración", "Bloqueo por horario, clave para quitar y usuarios") + f"""
    <div class="card">
      <h3 style="margin-top:0">Bloqueo de registro por horario</h3>
      <form method="post" class="form-grid" id="form_configuracion">
        <label style="font-weight:900"><input type="checkbox" name="bloqueo_activo" {'checked' if cfg_get('bloqueo_activo','0')=='1' else ''}> Activar bloqueo para usuarios</label>
        <input type="time" name="hora_inicio" value="{cfg_get('hora_inicio','00:00')}">
        <input type="time" name="hora_fin" value="{cfg_get('hora_fin','23:59')}">
        <input name="clave_quitar" value="{cfg_get('clave_quitar','1234')}" placeholder="Clave para quitar consumo">
        <button>Guardar configuración</button>
      </form>
      <p class="muted small">Con bloqueo activo, los usuarios registran solo dentro del horario. Admin puede registrar adicionales.</p>
    </div>

    <br>
    <div class="card">
      <div class="table-head"><h3>Usuarios y claves</h3><a class="btn btn-blue" href="{url_for('usuarios_admin')}">Crear usuarios</a></div>
      <div class="table-wrap"><table><tr><th>Usuario</th><th>Rol</th><th>Estado</th></tr>{usuarios_html}</table></div>
    </div>
    """
    return render_page(html, "config")


@app.route("/usuarios", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def usuarios_admin():
    if request.method == "POST":
        username = clean_text(request.form.get("username"))
        password = request.form.get("password") or ""
        role = asegurar_rol_usuario(request.form.get("role") or "comedor")
        active = 1 if request.form.get("active") else 0
        if not username or not password:
            flash("Usuario y clave son obligatorios.", "error")
            return redirect(url_for("usuarios_admin"))

        existe = q_one("SELECT id FROM usuarios WHERE username=?", (username,))
        if existe:
            q_exec("UPDATE usuarios SET password_hash=?, password_plain=?, role=?, active=? WHERE username=?",
                   (generate_password_hash(password), password, role, active, username))
            send_admin_user_notice(username, role, "actualizado")
            audit_event("USUARIO_ACTUALIZADO", "usuarios", username, f"Rol: {role}")
            flash("Usuario actualizado y guardado correctamente.", "ok")
        else:
            q_exec("INSERT INTO usuarios(username,password_hash,password_plain,role,active) VALUES(?,?,?,?,?)",
                   (username, generate_password_hash(password), password, role, active))
            send_admin_user_notice(username, role, "creado")
            audit_event("USUARIO_CREADO", "usuarios", username, f"Rol: {role}")
            flash("Usuario creado y guardado correctamente.", "ok")
        return redirect(url_for("usuarios_admin"))

    usuarios = q_all("SELECT id, username, role, active, COALESCE(password_plain,'') AS password_plain FROM usuarios ORDER BY id ASC")
    total_usuarios = len(usuarios)
    tabla = "".join([
        f"""
        <tr data-user-row data-user="{(u['username'] or '').lower()}" data-role="{(u['role'] or '').lower()}">
          <td>{i}</td>
          <td><b>{u['username']}</b></td>
          <td>{'Administrador total' if u['role']=='admin' else 'Usuario operativo'}</td>
          <td>
            <div class="pass-cell">
              <input class="pass-view" type="password" value="{u['password_plain'] or 'No registrada'}" readonly>
              <button type="button" class="eye-btn" onclick="togglePass(this)" title="Ver / ocultar contraseña">👁️</button>
            </div>
          </td>
          <td><span class='badge {'ok' if u['active'] else 'off'}'>{'Activo' if u['active'] else 'Bloqueado'}</span></td>
          <td>
            <form method='post' action='{url_for('eliminar_usuario', username=u['username'])}' onsubmit="return confirm('¿Eliminar este usuario?');" style='display:inline'>
              <button class='btn-orange' style='padding:8px 12px' {'disabled' if u['username'] in ['adm','adm1','adm2'] or u['username']==session.get('user') else ''}>Eliminar</button>
            </form>
          </td>
        </tr>
        """
        for i, u in enumerate(usuarios, 1)
    ])
    html = topbar("Crear usuarios y claves", "Solo administrador") + f"""
    <div class="card">
      <h3 style="margin-top:0">Crear / actualizar usuario</h3>
      <form method="post" class="form-grid" id="form_usuarios_admin">
        <input name="username" placeholder="Usuario" required>
        <input name="password" placeholder="Clave" required>
        <select name="role">
          <option value="comedor">Usuario operativo: Consumos / Entregas / Cerrar día</option>
          <option value="admin">Administrador total: crear y eliminar usuarios</option>
        </select>
        <label style="font-weight:900"><input type="checkbox" name="active" checked> Activo</label>
        <button>Guardar usuario</button>
      </form>
    </div>
    <br>
    <div class="card users-card">
      <div class="table-head" style="gap:14px;align-items:center;flex-wrap:wrap">
        <h3 style="margin:0">Usuarios registrados</h3>
        <span class="users-count">Total: {total_usuarios} usuario(s)</span>
        <input id="buscarUsuario" class="user-search" placeholder="🔎 Buscar usuario dinámicamente..." oninput="filtrarUsuarios()">
      </div>
      <div class="table-wrap users-scroll">
        <table id="tablaUsuarios">
          <tr><th>#</th><th>Usuario</th><th>Nivel</th><th>Contraseña</th><th>Estado</th><th>Acción</th></tr>{tabla}
        </table>
      </div>
      <p class="muted small">El usuario <b>adm</b> tiene clave <b>@123</b>. Los usuarios adm, adm1 y adm2 quedan como administradores totales.</p>
    </div>
    <script>
      function filtrarUsuarios(){{
        const q = (document.getElementById('buscarUsuario').value || '').toLowerCase().trim();
        let visibles = 0;
        document.querySelectorAll('[data-user-row]').forEach(tr => {{
          const texto = (tr.dataset.user + ' ' + tr.dataset.role + ' ' + tr.innerText.toLowerCase());
          const show = texto.includes(q);
          tr.style.display = show ? '' : 'none';
          if(show) visibles++;
        }});
        const badge = document.querySelector('.users-count');
        if(badge) badge.textContent = 'Total visible: ' + visibles + ' usuario(s)';
      }}
      function togglePass(btn){{
        const inp = btn.parentElement.querySelector('.pass-view');
        inp.type = inp.type === 'password' ? 'text' : 'password';
        btn.textContent = inp.type === 'password' ? '👁️' : '🙈';
      }}
    </script>
    """
    return render_page(html, "config")


@app.route("/usuarios/eliminar/<username>", methods=["POST"])
@login_required
@roles_required("admin")
def eliminar_usuario(username):
    username = clean_text(username)
    if username in ("adm", "adm1", "adm2"):
        flash("No se puede eliminar adm, adm1 ni adm2 porque son administradores principales.", "error")
        return redirect(url_for("usuarios_admin"))

    if username == session.get("user"):
        flash("No puedes eliminar el usuario con el que estás conectado.", "error")
        return redirect(url_for("usuarios_admin"))

    user = q_one("SELECT * FROM usuarios WHERE username=?", (username,))
    if not user:
        flash("Usuario no encontrado.", "error")
        return redirect(url_for("usuarios_admin"))

    if user["role"] == "admin":
        total_admins = q_one("SELECT COUNT(*) c FROM usuarios WHERE role='admin' AND active=1")["c"]
        if total_admins <= 2:
            flash("No se puede eliminar: deben quedar mínimo 2 administradores activos (adm1 y adm2).", "error")
            return redirect(url_for("usuarios_admin"))

    q_exec("DELETE FROM usuarios WHERE username=?", (username,))
    flash(f"Usuario {username} eliminado correctamente.", "ok")
    return redirect(url_for("usuarios_admin"))


# =========================
# DESCARGAS
# =========================
@app.route("/plantilla_consumos")
@login_required
def plantilla_consumos():
    df = pd.DataFrame([{
        "FECHA": hoy_iso(),
        "DNI": "74324033",
        "COMEDOR": "Comedor 01",
        "TIPO": "Almuerzo",
        "FUNDO": "Kawsay Allpa",
        "RESPONSABLE": "Nombre responsable",
        "CANTIDAD": 1,
        "PRECIO_UNITARIO": 10,
        "OBSERVACION": "Pedido desde Forms / QR DNI"
    }])
    output = BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="plantilla_carga_consumos.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/plantilla_trabajadores")
@login_required
def plantilla_trabajadores():
    df = pd.DataFrame([{
        "EMPRESA": "PRIZE",
        "DNI": "74324033",
        "NOMBRE": "AZABACHE LUJAN, OMAR EDUARDO",
        "CARGO": "OPERARIO",
        "AREA": "PRODUCCION"
    }])
    output = BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="plantilla_trabajadores.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/exportar_consumos")
@login_required
def exportar_consumos():
    # Exporta SOLO la data de un día. Por defecto, el día actual.
    fecha = request.args.get("fecha") or request.args.get("fecha_inicio") or hoy_iso()
    fecha = clean_text(fecha) or hoy_iso()
    rows = q_all("SELECT * FROM consumos WHERE fecha=? ORDER BY hora DESC,id DESC", (fecha,))
    df = pd.DataFrame([dict(r) for r in rows])
    output = BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"consumos_comedor_prize_{fecha}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/descargar_cierre/<path:filename>")
@login_required
def descargar_cierre(filename):
    safe = os.path.basename(filename)
    path = os.path.join(REPORT_DIR, safe)
    return send_file(path, as_attachment=True)


# =========================
# INICIO
# =========================
init_db()

if __name__ == "__main__":
    port = int(os.getenv("PORT", "5000"))
    app.run(host="0.0.0.0", port=port, debug=os.getenv("FLASK_DEBUG", "0") == "1")
